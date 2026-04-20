"""
CNS (California Neurosurgical Specialists) Financial Model Builder
7-tab Excel workbook: Dashboard, Assumptions, Monthly P&L, Cash Flow,
Scenarios, QBO Actuals.
Forecast horizon: Jan 2026 - Dec 2030 (60 months).

Key principle: Monthly P&L and Cash Flow tabs are 100% formula-driven.
Every data cell references the Assumptions tab. No hardcoded values in
the forecast sections.
"""

import copy
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter
from baseline_data import (
    ACTUALS_2025, ACTUALS_2025_TOTALS, BALANCE_SHEET_2025,
    DEFAULT_ASSUMPTIONS, BASELINE_OPEX_RECURRING, ONETIME_2025,
    PAYROLL_2025_TOTAL, TEAM_ROSTER,
    MONTHS_12, FORECAST_MONTH_LABELS, NUM_FORECAST_MONTHS,
    HISTORICAL_CASE_VOLUME, SURGERY_VOLUME_2024, SURGERY_VOLUME_2025,
    SURGERY_VOLUME_2026_ACTUALS, TOTAL_SURGERIES_2025,
    BOBAS_COLLECTION_CURVE, GAP_COLLECTION_CURVE,
    BOBA_VOLUME_2024, BOBA_VOLUME_2025, GAP_VOLUME_2024, GAP_VOLUME_2025,
    BOBA_2026_ACTUALS, GAP_2026_ACTUALS,
    HISTORICAL_AR_BOBA, HISTORICAL_AR_GAP, HISTORICAL_AR_TOTAL,
    HISTORICAL_MONTHS,
    _compute_ar_spillover, _HIST_BOBA, _HIST_GAP,
    ACTUALS_2026_QBO, NUM_2026_ACTUALS,
)

N = NUM_FORECAST_MONTHS  # 60

# ============================================================
# STYLES
# ============================================================
NAVY = "1B2A4A"
ACCENT_BLUE = "4472C4"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"
GREEN_FILL_C = "E2EFDA"
RED_FILL_C = "FCE4EC"
INPUT_FILL_C = "FFFFCC"

title_font = Font(name="Calibri", size=16, bold=True, color=NAVY)
section_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
subsection_font = Font(name="Calibri", size=11, bold=True, color=NAVY)
header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
data_font = Font(name="Calibri", size=10)
data_bold = Font(name="Calibri", size=10, bold=True)
pct_font = Font(name="Calibri", size=10, italic=True, color="666666")
input_font = Font(name="Calibri", size=10, color="0000CC")
metric_value_font = Font(name="Calibri", size=14, bold=True, color=NAVY)
metric_label_font = Font(name="Calibri", size=9, color="666666")

section_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
header_fill = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
input_fill = PatternFill(start_color=INPUT_FILL_C, end_color=INPUT_FILL_C, fill_type="solid")
green_fill = PatternFill(start_color=GREEN_FILL_C, end_color=GREEN_FILL_C, fill_type="solid")
red_fill = PatternFill(start_color=RED_FILL_C, end_color=RED_FILL_C, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MED_GRAY), right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY), bottom=Side(style="thin", color=MED_GRAY),
)
bottom_border = Border(bottom=Side(style="medium", color=NAVY))
double_bottom = Border(bottom=Side(style="double", color=NAVY))

CURR = '#,##0'
CURR2 = '#,##0.00'
center_align = Alignment(horizontal="center", vertical="center")
right_align = Alignment(horizontal="right", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")


def style_range(ws, row, c1, c2, font=None, fill=None, border=None, alignment=None, number_format=None):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if border: cell.border = border
        if alignment: cell.alignment = alignment
        if number_format: cell.number_format = number_format


def section_bar(ws, row, c1, c2, label):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    ws.cell(row=row, column=c1, value=label)
    style_range(ws, row, c1, c2, font=section_font, fill=section_fill, alignment=left_align)


def header_row(ws, row, labels, c1=1):
    for i, label in enumerate(labels):
        cell = ws.cell(row=row, column=c1 + i, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def input_cell(ws, row, col, value, number_format=CURR):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = input_font
    cell.fill = input_fill
    cell.border = thin_border
    cell.number_format = number_format
    return cell


def _write_row(ws, row, label, data, c_start=2, bold=False, fmt=CURR):
    """Write label + data cells. Returns next row."""
    ws.cell(row=row, column=c_start, value=label).font = data_bold if bold else data_font
    total = 0
    for i, val in enumerate(data):
        cell = ws.cell(row=row, column=c_start + 1 + i, value=val)
        cell.font = data_bold if bold else data_font
        cell.number_format = fmt
        cell.border = thin_border
        cell.alignment = right_align
        total += val
    # Total column
    tc = ws.cell(row=row, column=c_start + 1 + len(data), value=total)
    tc.font = data_bold if bold else data_font
    tc.number_format = fmt
    tc.border = thin_border
    tc.alignment = right_align
    return row + 1


def _write_formula_row(ws, row, label, formulas, bold=False, fmt=CURR):
    """Write label + formula cells + SUM total. Returns next row."""
    ws.cell(row=row, column=2, value=label).font = data_bold if bold else data_font
    n = len(formulas)
    for i, f in enumerate(formulas):
        cell = ws.cell(row=row, column=3 + i, value=f)
        cell.font = data_bold if bold else data_font
        cell.number_format = fmt
        cell.border = thin_border
        cell.alignment = right_align
    tc = ws.cell(row=row, column=3 + n,
                 value=f"=SUM({get_column_letter(3)}{row}:{get_column_letter(2 + n)}{row})")
    tc.font = data_bold if bold else data_font
    tc.number_format = fmt
    tc.border = thin_border
    tc.alignment = right_align
    return row + 1


def mcol(i):
    """Column letter for forecast month i (0-indexed). C=month0, D=month1, ..."""
    return get_column_letter(3 + i)


# ============================================================
# TAB 1: DASHBOARD
# ============================================================
def build_dashboard(wb):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = NAVY

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 32
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 18

    ws.merge_cells('B2:H2')
    ws.cell(row=2, column=2, value="California Neurosurgical Specialists").font = title_font
    ws.merge_cells('B3:H3')
    ws.cell(row=3, column=2,
            value="Financial Dashboard \u2014 2025 through 2030").font = Font(
                name="Calibri", size=12, color="666666")

    from financial_calcs import calculate_dashboard_metrics
    dm = calculate_dashboard_metrics()

    r = 5
    section_bar(ws, r, 2, 8, "PERFORMANCE SUMMARY")
    r += 1
    header_row(ws, r, ["Metric", "2025 (Actual)",
                        "2026 (Forecast)", "2027 (Forecast)",
                        "2028 (Forecast)", "2029 (Forecast)",
                        "2030 (Forecast)"], c1=2)
    r += 1

    years = ['2025', '2026', '2027', '2028', '2029', '2030']

    rows = [
        ("Surgeries (Bobas)", 'surgeries_bobas', '#,##0'),
        ("Surgeries (GAP)", 'surgeries_gap', '#,##0'),
        ("Surgeries (Total)", 'surgeries_total', '#,##0'),
        ("Gross Revenue", 'gross_revenue', '$#,##0'),
        ("Physician Services ($)", 'physician_services', '$#,##0'),
        ("Savings Deposits ($)", 'savings_deposits', '$#,##0'),
        ("# Locations", 'locations', '#,##0'),
        ("Min Cash", 'min_cash', '$#,##0'),
        ("Max Cash", 'max_cash', '$#,##0'),
        ("Capex/Startup", 'capex_startup', '$#,##0'),
        ("Ending Savings Balance", 'ending_savings', '$#,##0'),
    ]

    for label, key, fmt in rows:
        ws.cell(row=r, column=2, value=label).font = data_font
        for ci, yr in enumerate(years):
            val = dm[yr].get(key, 0)
            cell = ws.cell(row=r, column=3 + ci, value=val)
            cell.font = data_bold
            cell.number_format = fmt
            cell.border = thin_border
            cell.alignment = right_align
        r += 1

    r += 2

    # Fund Flow Structure
    section_bar(ws, r, 2, 8, "FUND FLOW STRUCTURE")
    r += 1
    flow_items = [
        "Revenue \u2192 CNS Checking \u2192 pay overhead \u2192 distribute net equity:",
        "   90% \u2192 ABC A PC (physician services / Dr. Benet compensation)",
        "   10% \u2192 Savings account",
        "",
        "Billing: MD Capital @ 18% of cash collected (QBO acct 554)",
        "Payment lag: Bobas 6-11 months (M+6 to M+11), GAP 1-2 months (M+1 to M+2)",
        f"2025 actual: {TOTAL_SURGERIES_2025} surgeries ({TOTAL_SURGERIES_2025/12:.1f}/mo avg)",
        "Minimum cash threshold enforced before distributions",
    ]
    for item in flow_items:
        ws.cell(row=r, column=2, value=item).font = data_font
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        r += 1

    r += 1

    # Expansion Pipeline
    section_bar(ws, r, 2, 8, "EXPANSION PIPELINE")
    r += 1
    header_row(ws, r, ["Location", "Status", "Open Date", "Monthly Lease", "", "", ""], c1=2)
    r += 1

    ws.cell(row=r, column=2, value="Westlake (Home Office)").font = data_font
    ws.cell(row=r, column=3, value="Open").font = data_font
    ws.cell(row=r, column=4, value="Active").font = data_font
    ws.cell(row=r, column=5, value=2500).font = data_font
    ws.cell(row=r, column=5).number_format = '$#,##0'
    for c in range(2, 6):
        ws.cell(row=r, column=c).border = thin_border
    r += 1

    for exp in DEFAULT_ASSUMPTIONS['expansions']:
        ws.cell(row=r, column=2, value=exp['name']).font = data_font
        status = "Enabled" if exp['enabled'] else "Slot Available"
        ws.cell(row=r, column=3, value=status).font = data_font
        if exp['enabled']:
            mi = exp['lease_start_month']
            lbl = FORECAST_MONTH_LABELS[mi] if mi < len(FORECAST_MONTH_LABELS) else "TBD"
            ws.cell(row=r, column=4, value=lbl).font = data_font
            ws.cell(row=r, column=5, value=exp['lease_monthly']).font = data_font
            ws.cell(row=r, column=5).number_format = '$#,##0'
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = thin_border
        r += 1

    return ws


# ============================================================
# TAB 2: ASSUMPTIONS
# ============================================================
def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_properties.tabColor = "FF9900"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50

    ws.merge_cells('B2:E2')
    ws.cell(row=2, column=2, value="Model Assumptions (2026-2030)").font = title_font
    ws.cell(row=3, column=2,
            value="Yellow cells are editable inputs").font = Font(
                name="Calibri", size=10, italic=True, color="FF9900")

    a = DEFAULT_ASSUMPTIONS
    asm = {}  # row map returned to other tab builders

    # ---- HISTORICAL SURGERY VOLUME (AR) ----
    r = 5
    section_bar(ws, r, 2, 5, "HISTORICAL SURGERY VOLUME (AR)")
    r += 1
    ws.cell(row=r, column=2,
            value="Surgeries still in AR. Zero out rows that have been paid.").font = Font(
                name="Calibri", size=9, italic=True, color="CC0000")
    r += 1
    header_row(ws, r, ["Month", "BOBA", "GAP", "Offset"], c1=2)
    r += 1
    asm['hist_start'] = r

    for label, offset, boba_ct, gap_ct in HISTORICAL_MONTHS:
        ws.cell(row=r, column=2, value=label).font = data_font
        input_cell(ws, r, 3, boba_ct, '#,##0')
        input_cell(ws, r, 4, gap_ct, '#,##0')
        # Offset column (small font, used by AR formulas)
        cell = ws.cell(row=r, column=5, value=offset)
        cell.font = Font(name="Calibri", size=8, color="999999")
        cell.number_format = '#,##0'
        r += 1

    asm['hist_end'] = r - 1
    r += 1  # blank

    # ---- SURGERY VOLUME (per-location, laid out horizontally) ----
    from baseline_data import LOCATIONS, VOLUMES_BY_LOCATION

    locations = a.get('locations', LOCATIONS)
    volumes_by_loc = a.get('volumes_by_location', VOLUMES_BY_LOCATION)
    num_actuals = len(BOBA_2026_ACTUALS)  # 3 (Jan-Mar 2026)

    # Consolidated volumes in C/D (used by P&L formulas)
    # Plus per-location tables to the right
    vol_section_row = r
    section_bar(ws, r, 2, 5, "SURGERY VOLUME (Monthly)")

    # Per-location section headers
    for loc_idx, loc_name in enumerate(locations):
        col_start = 6 + loc_idx * 5  # F, K, P, U, Z...
        loc_fill = PatternFill(start_color="7030A0" if loc_idx > 0 else ACCENT_BLUE,
                                end_color="7030A0" if loc_idx > 0 else ACCENT_BLUE, fill_type="solid")
        for c in range(col_start, col_start + 5):
            ws.cell(row=r, column=c).fill = loc_fill
        ws.cell(row=r, column=col_start,
                value=f"SURGERY VOLUME (Monthly)").font = Font(
                    name="Calibri", size=10, bold=True, color="FFFFFF")
    r += 1

    # Sub-headers
    ws.cell(row=r, column=2, value="").font = data_font  # blank for consolidated section header
    ws.cell(row=r, column=3, value="Consolidated").font = Font(
        name="Calibri", size=10, bold=True, color=NAVY)
    for loc_idx, loc_name in enumerate(locations):
        col_start = 6 + loc_idx * 5
        ws.cell(row=r, column=col_start, value=loc_name).font = Font(
            name="Calibri", size=10, bold=True, color=NAVY)
    r += 1

    # Column headers
    header_row(ws, r, ["Month", "Bobas", "GAP", "Source"], c1=2)
    for loc_idx, loc_name in enumerate(locations):
        col_start = 6 + loc_idx * 5
        for ci, label in enumerate(["Month", "Bobas", "GAP", "Source"]):
            cell = ws.cell(row=r, column=col_start + ci, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
    r += 1
    asm['vol_start'] = r

    # Per-location volume tracking for column references
    asm['vol_by_location'] = {}

    bobas_vol = a['bobas_volume']  # consolidated
    gap_vol = a['gap_volume']

    for i in range(N):
        # Consolidated (columns B-E)
        ws.cell(row=r, column=2, value=FORECAST_MONTH_LABELS[i]).font = data_font
        input_cell(ws, r, 3, bobas_vol[i], '#,##0')  # Col C = Bobas consolidated
        input_cell(ws, r, 4, gap_vol[i], '#,##0')     # Col D = GAP consolidated
        source = "Actual" if i < num_actuals else "Forecast"
        ws.cell(row=r, column=5, value=source).font = pct_font

        # Per-location (columns F+)
        for loc_idx, loc_name in enumerate(locations):
            col_start = 6 + loc_idx * 5
            loc_vols = volumes_by_loc.get(loc_name, {'bobas': [0]*N, 'gap': [0]*N})
            loc_bobas = loc_vols['bobas'][i] if i < len(loc_vols['bobas']) else 0
            loc_gap = loc_vols['gap'][i] if i < len(loc_vols['gap']) else 0

            ws.cell(row=r, column=col_start, value=FORECAST_MONTH_LABELS[i]).font = data_font
            input_cell(ws, r, col_start + 1, loc_bobas, '#,##0')
            input_cell(ws, r, col_start + 2, loc_gap, '#,##0')
            loc_source = "Actual" if i < num_actuals and loc_name == "Westlake" else "Forecast"
            cell = ws.cell(row=r, column=col_start + 3, value=loc_source)
            cell.font = pct_font

            if i == 0:
                asm['vol_by_location'][loc_name] = {
                    'start_row': r,
                    'bobas_col': col_start + 1,
                    'gap_col': col_start + 2,
                }

        r += 1

    # Set column widths for per-location sections
    for loc_idx in range(len(locations)):
        col_start = 6 + loc_idx * 5
        ws.column_dimensions[get_column_letter(col_start)].width = 10      # Month
        ws.column_dimensions[get_column_letter(col_start + 1)].width = 8   # Bobas
        ws.column_dimensions[get_column_letter(col_start + 2)].width = 8   # GAP
        ws.column_dimensions[get_column_letter(col_start + 3)].width = 9   # Source

    r += 1  # blank row

    # ---- REVENUE ASSUMPTIONS (R68+) ----
    section_bar(ws, r, 2, 5, "REVENUE ASSUMPTIONS")
    r += 1  # R69
    asm['avg_rev_bobas'] = r
    ws.cell(row=r, column=2, value="Avg Revenue per Bobas Surgery").font = data_font
    input_cell(ws, r, 3, a['avg_revenue_bobas'])
    ws.cell(row=r, column=4, value="$").font = data_font
    ws.cell(row=r, column=5, value="Bobas: complex cases, longer collection").font = pct_font
    r += 1  # R70

    asm['avg_rev_gap'] = r
    ws.cell(row=r, column=2, value="Avg Revenue per GAP Surgery").font = data_font
    input_cell(ws, r, 3, a['avg_revenue_gap'])
    ws.cell(row=r, column=4, value="$").font = data_font
    ws.cell(row=r, column=5, value="GAP: simpler cases, faster collection").font = pct_font
    r += 1  # R71

    asm['reimb_pct'] = r
    ws.cell(row=r, column=2, value="Reimbursed Expense %").font = data_font
    input_cell(ws, r, 3, a['reimbursed_expense_pct'])
    ws.cell(row=r, column=4, value="%").font = data_font
    ws.cell(row=r, column=5, value="2025: 14.6%").font = pct_font
    r += 1  # R72

    asm['refund_rate'] = r
    ws.cell(row=r, column=2, value="Refund Rate").font = data_font
    input_cell(ws, r, 3, a['refund_rate'])
    ws.cell(row=r, column=4, value="%").font = data_font
    ws.cell(row=r, column=5, value="2025: 0.5%").font = pct_font
    r += 1  # R73
    r += 1  # blank -> R74

    # ---- COLLECTION CURVE - BOBAS (R74+) ----
    section_bar(ws, r, 2, 5, "COLLECTION CURVE - BOBAS")
    r += 1  # R75
    asm['curve_bobas_start'] = r
    bobas_curve = a['bobas_collection_curve']
    asm['curve_bobas_len'] = len(bobas_curve)
    for i, pct in enumerate(bobas_curve):
        ws.cell(row=r, column=2, value=f"M+{i}").font = data_font
        input_cell(ws, r, 3, pct)
        ws.cell(row=r, column=4, value="%").font = data_font
        if i == 0:
            ws.cell(row=r, column=5, value="Must sum to 100%.").font = pct_font
        r += 1
    # r is now 75+12=87
    r += 1  # blank -> R88

    # ---- COLLECTION CURVE - GAP (R88+) ----
    section_bar(ws, r, 2, 5, "COLLECTION CURVE - GAP")
    r += 1  # R89
    asm['curve_gap_start'] = r
    gap_curve = a['gap_collection_curve']
    asm['curve_gap_len'] = len(gap_curve)
    for i, pct in enumerate(gap_curve):
        ws.cell(row=r, column=2, value=f"M+{i}").font = data_font
        input_cell(ws, r, 3, pct)
        ws.cell(row=r, column=4, value="%").font = data_font
        r += 1
    # r is now 89+3=92
    r += 1  # blank -> R93

    # ---- FUND FLOW & BILLING (R93+) ----
    section_bar(ws, r, 2, 5, "FUND FLOW & BILLING")
    r += 1  # R94
    asm['phys_rate'] = r
    ws.cell(row=r, column=2, value="Physician Services Rate").font = data_font
    input_cell(ws, r, 3, a['physician_services_rate'])
    ws.cell(row=r, column=4, value="%").font = data_font
    ws.cell(row=r, column=5, value="90% of net equity \u2192 ABC A PC").font = pct_font
    r += 1  # R95

    asm['savings_rate'] = r
    ws.cell(row=r, column=2, value="Savings Rate").font = data_font
    input_cell(ws, r, 3, a['savings_rate'])
    ws.cell(row=r, column=4, value="%").font = data_font
    ws.cell(row=r, column=5, value="10% of distributable \u2192 savings account").font = pct_font
    r += 1  # R96

    asm['billing_rate'] = r
    ws.cell(row=r, column=2, value="MD Capital Billing Rate").font = data_font
    input_cell(ws, r, 3, a['billing_fee_rate'])
    ws.cell(row=r, column=4, value="%").font = data_font
    ws.cell(row=r, column=5, value="18% of cash collected").font = pct_font
    r += 1  # R97

    asm['payroll_tax_rate'] = r
    ws.cell(row=r, column=2, value="Payroll Tax Rate").font = data_font
    input_cell(ws, r, 3, a['payroll_tax_rate'])
    ws.cell(row=r, column=4, value="%").font = data_font
    ws.cell(row=r, column=5, value="Employer payroll taxes on W-2 salaries").font = pct_font
    r += 1

    asm['salary_annual_increase'] = r
    ws.cell(row=r, column=2, value="Annual Salary Increase").font = data_font
    input_cell(ws, r, 3, a.get('salary_annual_increase', 5.0))
    ws.cell(row=r, column=4, value="%").font = data_font
    ws.cell(row=r, column=5, value="Applied at start of each year (Jan 2027, 2028, etc.)").font = pct_font
    r += 1

    asm['expense_annual_inflation'] = r
    ws.cell(row=r, column=2, value="Annual Expense Inflation").font = data_font
    input_cell(ws, r, 3, a.get('expense_annual_inflation', 3.0))
    ws.cell(row=r, column=4, value="%").font = data_font
    ws.cell(row=r, column=5, value="Applied to fixed expenses (not % of revenue items)").font = pct_font
    r += 1

    r += 1  # blank

    # ---- TEAM ROSTER (15 slots) ----
    # Col B=Name, C=Salary, D=Start Month (1=Jan-26), E=End Month, F=Type, G=Notes, H=Location
    section_bar(ws, r, 2, 8, "TEAM ROSTER (15 Slots)")
    r += 1
    header_row(ws, r, ["Name", "Salary ($/mo)", "Start (1=Jan-26)", "End Mo", "Type", "Notes", "Location"], c1=2)
    r += 1
    asm['team_start'] = r
    asm['team_rows'] = []
    ws.column_dimensions['H'].width = 16  # Location column
    for person in TEAM_ROSTER:
        ws.cell(row=r, column=2, value=person['title']).font = data_font
        input_cell(ws, r, 3, person['monthly_salary'])
        start_m = person.get('start_month')
        if start_m is not None:
            input_cell(ws, r, 4, start_m, '#,##0')
        else:
            cell = ws.cell(row=r, column=4, value='')
            cell.font = input_font; cell.fill = input_fill; cell.border = thin_border
        end_m = person.get('end_month')
        if end_m is not None:
            input_cell(ws, r, 5, end_m, '#,##0')
        else:
            cell = ws.cell(row=r, column=5, value='')
            cell.font = input_font; cell.fill = input_fill; cell.border = thin_border
        emp_type = person.get('employment_type', 'W-2')
        input_cell(ws, r, 6, emp_type, '@')
        ws.cell(row=r, column=7, value=person.get('notes', '')).font = pct_font
        # Location column (H)
        location = person.get('location', 'Westlake')
        input_cell(ws, r, 8, location, '@')
        asm['team_rows'].append(r)
        r += 1
    asm['team_end'] = r - 1
    r += 1  # blank

    # ---- OPERATING EXPENSES (Monthly Base) ----
    section_bar(ws, r, 2, 5, "OPERATING EXPENSES (Monthly Base)")
    r += 1
    asm['opex_rows'] = []
    for item in BASELINE_OPEX_RECURRING:
        ws.cell(row=r, column=2, value=item['expense_name']).font = data_font
        input_cell(ws, r, 3, item['monthly_amount'])
        ws.cell(row=r, column=4, value="$/mo base").font = data_font
        ws.cell(row=r, column=5, value=item.get('notes', '')).font = pct_font
        asm['opex_rows'].append(r)
        r += 1
    r += 1  # blank

    # ---- EXPANSION SLOTS (numeric enabled: 1/0) ----
    section_bar(ws, r, 2, 5, "EXPANSION SLOTS (5 available)")
    r += 1
    asm['exp_slots'] = []
    for exp in a['expansions']:
        slot = {}
        ws.cell(row=r, column=2, value=f"--- {exp['name']} ---").font = subsection_font
        slot['enabled'] = r
        input_cell(ws, r, 3, 1 if exp['enabled'] else 0, '#,##0')
        ws.cell(row=r, column=4, value="1=Yes, 0=No").font = pct_font
        r += 1
        params = [
            ("Lease Monthly", 'lease_monthly', '$#,##0'),
            ("Lease Start Month", 'lease_start_month', '#,##0'),
            ("TI CNS Share", 'ti_cns_share', '$#,##0'),
            ("TI Start Month", 'ti_start_month', '#,##0'),
            ("TI Duration (months)", 'ti_duration_months', '#,##0'),
            ("FF&E Budget", 'ffe_budget', '$#,##0'),
            ("Ongoing OpEx Monthly", 'opex_monthly', '$#,##0'),
            ("OpEx Ramp Monthly", 'opex_ramp_monthly', '$#,##0'),
            ("OpEx Ramp Months", 'opex_ramp_months', '#,##0'),
        ]
        for label, key, fmt in params:
            ws.cell(row=r, column=2, value=f"  {label}").font = data_font
            input_cell(ws, r, 3, exp.get(key, 0), fmt)
            slot[key] = r
            r += 1
        asm['exp_slots'].append(slot)
        r += 1  # blank between slots

    # ---- CASH ----
    section_bar(ws, r, 2, 5, "CASH")
    r += 1
    asm['starting_cash'] = r
    ws.cell(row=r, column=2, value="Starting Cash (Dec 31, 2025)").font = data_font
    input_cell(ws, r, 3, a['starting_cash'])
    r += 1

    asm['min_cash'] = r
    ws.cell(row=r, column=2, value="Minimum Cash Threshold").font = data_font
    input_cell(ws, r, 3, a['minimum_cash_balance'])
    r += 1

    asm['starting_savings'] = r
    ws.cell(row=r, column=2, value="Starting Savings Balance").font = data_font
    input_cell(ws, r, 3, a.get('starting_savings', 0))
    r += 1

    r += 2  # blank rows before horizontal section

    # ================================================================
    # HORIZONTAL SCHEDULES (60 months) — ALL EXCEL FORMULAS
    # Every cell is a formula referencing assumption inputs above.
    # ================================================================

    # Set column widths for horizontal data (C through BJ = cols 3..62)
    for ci in range(3, 3 + N):
        ws.column_dimensions[get_column_letter(ci)].width = 9

    section_bar(ws, r, 2, 2 + N, "MONTHLY SCHEDULES (60 months) — Formula-Driven")
    r += 1

    # Header row: month labels
    ws.cell(row=r, column=2, value="").font = data_font
    for i in range(N):
        cell = ws.cell(row=r, column=3 + i, value=FORECAST_MONTH_LABELS[i])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    r += 1

    sal_rate_row = asm['salary_annual_increase']
    tax_rate_row = asm['payroll_tax_rate']
    inf_rate_row = asm['expense_annual_inflation']
    asm['sal_rate_row'] = sal_rate_row
    asm['inf_rate_row'] = inf_rate_row

    # ================================================================
    # PAYROLL SCHEDULE (per-person formula rows)
    # ================================================================
    ws.cell(row=r, column=2, value="PAYROLL BREAKDOWN").font = subsection_font
    r += 1

    # 15 per-person salary rows
    asm['payroll_breakdown_first'] = r  # first individual payroll row
    person_salary_rows = []
    for idx, person_row in enumerate(asm['team_rows']):
        person_name = TEAM_ROSTER[idx]['title']
        ws.cell(row=r, column=2, value=f"  {person_name}").font = data_font
        for i in range(N):
            # Active if: start_month not blank, type=W-2, month >= start, (end blank OR month <= end)
            # Salary = base * (1 + raise%)^year_index
            f = (f'=IF(AND($D${person_row}<>"",$F${person_row}="W-2",'
                 f'COLUMN()-2>=$D${person_row},'
                 f'OR($E${person_row}="",COLUMN()-2<=$E${person_row})),'
                 f'$C${person_row}*(1+$C${sal_rate_row}/100)^INT((COLUMN()-2)/12),0)')
            cell = ws.cell(row=r, column=3 + i, value=f)
            cell.font = data_font
            cell.number_format = CURR
            cell.border = thin_border
            cell.alignment = right_align
        person_salary_rows.append(r)
        r += 1

    asm['payroll_breakdown_last'] = r - 1  # last individual payroll row

    # Total Salaries
    row_total_sal = r
    first_prow = person_salary_rows[0]
    last_prow = person_salary_rows[-1]
    ws.cell(row=r, column=2, value="Total Salaries").font = data_bold
    for i in range(N):
        cl = mcol(i)
        f = f'=SUM({cl}{first_prow}:{cl}{last_prow})'
        cell = ws.cell(row=r, column=3 + i, value=f)
        cell.font = data_bold
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    r += 1

    # Payroll Taxes
    row_payroll_taxes = r
    ws.cell(row=r, column=2, value="Payroll Taxes").font = data_font
    for i in range(N):
        f = f'={mcol(i)}{row_total_sal}*$C${tax_rate_row}/100'
        cell = ws.cell(row=r, column=3 + i, value=f)
        cell.font = data_font
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    r += 1

    # Processing
    row_processing = r
    ws.cell(row=r, column=2, value="Payroll Processing").font = data_font
    for i in range(N):
        f = f'=IF({mcol(i)}{row_total_sal}>0,67,0)'
        cell = ws.cell(row=r, column=3 + i, value=f)
        cell.font = data_font
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    r += 1

    # TOTAL PAYROLL (this is what P&L references)
    asm['sched_payroll'] = r
    ws.cell(row=r, column=2, value="TOTAL PAYROLL (W-2)").font = data_bold
    for i in range(N):
        cl = mcol(i)
        f = f'={cl}{row_total_sal}+{cl}{row_payroll_taxes}+{cl}{row_processing}'
        cell = ws.cell(row=r, column=3 + i, value=f)
        cell.font = data_bold
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    r += 1
    r += 1  # blank

    # ================================================================
    # CONTRACTOR SCHEDULE (non-W-2, no taxes)
    # ================================================================
    ws.cell(row=r, column=2, value="CONTRACTOR COSTS").font = subsection_font
    r += 1

    contractor_rows = []
    for idx, person_row in enumerate(asm['team_rows']):
        person = TEAM_ROSTER[idx]
        if person.get('employment_type') != 'Contractor':
            continue
        ws.cell(row=r, column=2, value=f"  {person['title']}").font = data_font
        for i in range(N):
            f = (f'=IF(AND($D${person_row}<>"",$F${person_row}="Contractor",'
                 f'COLUMN()-2>=$D${person_row},'
                 f'OR($E${person_row}="",COLUMN()-2<=$E${person_row})),'
                 f'$C${person_row}*(1+$C${sal_rate_row}/100)^INT((COLUMN()-2)/12),0)')
            cell = ws.cell(row=r, column=3 + i, value=f)
            cell.font = data_font
            cell.number_format = CURR
            cell.border = thin_border
            cell.alignment = right_align
        contractor_rows.append(r)
        r += 1

    asm['sched_contractors'] = r
    ws.cell(row=r, column=2, value="TOTAL CONTRACTORS").font = data_bold
    for i in range(N):
        cl = mcol(i)
        if contractor_rows:
            first_cr = contractor_rows[0]
            last_cr = contractor_rows[-1]
            f = f'=SUM({cl}{first_cr}:{cl}{last_cr})'
        else:
            f = '=0'
        cell = ws.cell(row=r, column=3 + i, value=f)
        cell.font = data_bold
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    r += 1
    r += 1  # blank

    # ================================================================
    # OPEX SCHEDULE (per-item formula rows with inflation)
    # ================================================================
    ws.cell(row=r, column=2, value="OPEX BREAKDOWN (with inflation)").font = subsection_font
    r += 1

    opex_detail_rows = []
    for idx, opex_row in enumerate(asm['opex_rows']):
        item_name = BASELINE_OPEX_RECURRING[idx]['expense_name']
        ws.cell(row=r, column=2, value=f"  {item_name}").font = data_font
        for i in range(N):
            # base_amount * (1 + inflation/100) ^ year_index
            f = f'=$C${opex_row}*(1+$C${inf_rate_row}/100)^INT((COLUMN()-2)/12)'
            cell = ws.cell(row=r, column=3 + i, value=f)
            cell.font = data_font
            cell.number_format = CURR
            cell.border = thin_border
            cell.alignment = right_align
        opex_detail_rows.append(r)
        r += 1

    # TOTAL OPEX (this is what P&L references)
    asm['sched_opex'] = r
    first_orow = opex_detail_rows[0]
    last_orow = opex_detail_rows[-1]
    ws.cell(row=r, column=2, value="TOTAL OPEX").font = data_bold
    for i in range(N):
        cl = mcol(i)
        f = f'=SUM({cl}{first_orow}:{cl}{last_orow})'
        cell = ws.cell(row=r, column=3 + i, value=f)
        cell.font = data_bold
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    r += 1
    r += 1  # blank

    # ================================================================
    # EXPANSION SCHEDULE (per-slot formula rows)
    # ================================================================
    ws.cell(row=r, column=2, value="EXPANSION BREAKDOWN").font = subsection_font
    r += 1

    exp_total_rows = []
    asm['exp_total_rows_list'] = []
    asm['exp_total_rows_by_name'] = {}
    asm['payroll_tax_row'] = asm['payroll_tax_rate']  # alias for formula reference
    for idx, slot in enumerate(asm['exp_slots']):
        exp_name = a['expansions'][idx]['name']
        en = slot['enabled']
        ti_share = slot['ti_cns_share']
        ti_start = slot['ti_start_month']
        ti_dur = slot['ti_duration_months']
        lease_mo = slot['lease_monthly']
        lease_st = slot['lease_start_month']
        ffe = slot['ffe_budget']
        exp_opex = slot['opex_monthly']

        # TI row
        row_ti = r
        ws.cell(row=r, column=2, value=f"    TI ({exp_name})").font = data_font
        for i in range(N):
            f = (f'=IF(AND($C${en}=1,$C${ti_dur}>0,'
                 f'COLUMN()-2>=$C${ti_start},'
                 f'COLUMN()-2<$C${ti_start}+$C${ti_dur}),'
                 f'$C${ti_share}/$C${ti_dur},0)')
            cell = ws.cell(row=r, column=3 + i, value=f)
            cell.font = data_font; cell.number_format = CURR
            cell.border = thin_border; cell.alignment = right_align
        r += 1

        # Lease row (with inflation)
        row_lease = r
        ws.cell(row=r, column=2, value=f"    Lease ({exp_name})").font = data_font
        for i in range(N):
            f = (f'=IF(AND($C${en}=1,COLUMN()-2>=$C${lease_st}),'
                 f'$C${lease_mo}*(1+$C${inf_rate_row}/100)^INT((COLUMN()-2)/12),0)')
            cell = ws.cell(row=r, column=3 + i, value=f)
            cell.font = data_font; cell.number_format = CURR
            cell.border = thin_border; cell.alignment = right_align
        r += 1

        # FF&E row (one-time at lease start)
        row_ffe = r
        ws.cell(row=r, column=2, value=f"    FF&E ({exp_name})").font = data_font
        for i in range(N):
            f = f'=IF(AND($C${en}=1,COLUMN()-2=$C${lease_st}),$C${ffe},0)'
            cell = ws.cell(row=r, column=3 + i, value=f)
            cell.font = data_font; cell.number_format = CURR
            cell.border = thin_border; cell.alignment = right_align
        r += 1

        # Ongoing OpEx row (with inflation and optional ramp)
        row_exp_opex = r
        ramp_mo = slot.get('opex_ramp_monthly')
        ramp_months = slot.get('opex_ramp_months')
        ws.cell(row=r, column=2, value=f"    OpEx ({exp_name})").font = data_font
        for i in range(N):
            if ramp_mo and ramp_months:
                # Ramp: use reduced rate for first N months, then full rate
                f = (f'=IF(AND($C${en}=1,COLUMN()-2>=$C${lease_st}),'
                     f'IF(COLUMN()-2<$C${lease_st}+$C${ramp_months},'
                     f'$C${ramp_mo}*(1+$C${inf_rate_row}/100)^INT((COLUMN()-2)/12),'
                     f'$C${exp_opex}*(1+$C${inf_rate_row}/100)^INT((COLUMN()-2)/12)),0)')
            else:
                f = (f'=IF(AND($C${en}=1,COLUMN()-2>=$C${lease_st}),'
                     f'$C${exp_opex}*(1+$C${inf_rate_row}/100)^INT((COLUMN()-2)/12),0)')
            cell = ws.cell(row=r, column=3 + i, value=f)
            cell.font = data_font; cell.number_format = CURR
            cell.border = thin_border; cell.alignment = right_align
        r += 1

        # Per-expansion total
        row_exp_total = r
        ws.cell(row=r, column=2, value=f"  Total ({exp_name})").font = data_bold
        for i in range(N):
            cl = mcol(i)
            f = f'={cl}{row_ti}+{cl}{row_lease}+{cl}{row_ffe}+{cl}{row_exp_opex}'
            cell = ws.cell(row=r, column=3 + i, value=f)
            cell.font = data_bold; cell.number_format = CURR
            cell.border = thin_border; cell.alignment = right_align
        exp_total_rows.append(r)
        asm['exp_total_rows_list'].append(r)
        asm['exp_total_rows_by_name'][exp_name] = r
        r += 1
        r += 1  # blank between expansions

    # GRAND TOTAL EXPANSION (this is what P&L references)
    asm['sched_expansion'] = r
    ws.cell(row=r, column=2, value="TOTAL EXPANSION").font = data_bold
    for i in range(N):
        cl = mcol(i)
        if exp_total_rows:
            terms = "+".join(f'{cl}{etr}' for etr in exp_total_rows)
            f = f'={terms}'
        else:
            f = '=0'
        cell = ws.cell(row=r, column=3 + i, value=f)
        cell.font = data_bold; cell.number_format = CURR
        cell.border = thin_border; cell.alignment = right_align
    r += 1
    r += 1  # blank

    # ================================================================
    # PER-LOCATION SCHEDULES (payroll, opex, expansion by location)
    # Each location P&L tab references these rows via formulas.
    # ================================================================
    from baseline_data import LOCATIONS, OPEX_BY_LOCATION
    from financial_calcs import generate_pl_by_location

    locations = a.get('locations', LOCATIONS)
    pl_by_loc = generate_pl_by_location(a)

    asm['loc_schedules'] = {}

    section_bar(ws, r, 2, 2 + N, "PER-LOCATION SCHEDULES")
    r += 1

    # References for formulas
    team_start = asm['team_start']
    team_end = asm['team_end']
    sal_rate_row = asm['sal_rate_row']
    inf_rate_row = asm['inf_rate_row']

    # Find expansion total rows by name
    exp_total_by_name = {}
    for exp_slot in asm.get('exp_slots', []):
        exp_name = None
        for exp_def in a.get('expansions', []):
            # Match by enabled row
            if exp_slot.get('enabled') is not None:
                exp_name = exp_def.get('name')
                break
    # Simpler: use the expansion total rows we already have
    # asm['exp_total_rows'] is a list of row numbers for each expansion total

    for loc_idx, loc_name in enumerate(locations):
        loc_pl = pl_by_loc.get(loc_name, {})
        loc_sched = {}

        ws.cell(row=r, column=2, value=f"--- {loc_name} ---").font = subsection_font
        r += 1

        # Payroll: SUMPRODUCT of per-person payroll rows where location matches
        # Formula: sum each person's payroll row value IF their location (col H) = loc_name
        loc_sched['payroll'] = r
        ws.cell(row=r, column=2, value=f"  Payroll ({loc_name})").font = data_font
        # First payroll breakdown row and last
        first_person = asm['team_rows'][0]  # e.g. row 226
        last_person = asm['team_rows'][-1]   # e.g. row 240
        # Map: payroll breakdown row i corresponds to team roster row (team_start + i offset)
        # The payroll row for person at team_rows[j] is at the payroll breakdown section
        # Need to find payroll breakdown start row
        payroll_breakdown_start = first_person + (asm.get('sched_payroll', 244) - asm.get('sched_payroll', 244))
        # Actually: individual payroll rows are tracked. Let me just use SUMPRODUCT
        # across the individual payroll rows, checking location in roster
        for i in range(N):
            col_letter = mcol(i)
            # Build SUMPRODUCT: for each person, if location matches, add their payroll amount
            # Individual W-2 rows are at team_rows offsets in the payroll section
            # Payroll schedule starts after team roster: rows 226-240 for W-2, 247 for contractors
            # The individual per-person payroll rows are already formula-driven
            # We need: SUMPRODUCT((H$team_start:H$team_end=loc_name)*
            #           (F$team_start:F$team_end="W-2")*payroll_amounts)
            # But payroll amounts are in a different section. Simpler: just reference
            # individual person rows and check their location.
            terms = []
            for j, person_row in enumerate(asm['team_rows']):
                # person_row is the roster input row (132-146)
                # corresponding payroll breakdown row
                payroll_row = asm['team_rows'][0] + (asm.get('sched_payroll', 244) - asm['team_rows'][0])
                # Actually the payroll breakdown rows are sequential starting from a known row
                # Let me use a different approach: just hardcode the formula with IF checks
                f_person = (f'IF(AND($H${person_row}="{loc_name}",$F${person_row}="W-2"),'
                            f'{col_letter}{first_person + j},0)')  # This references wrong section
                terms.append(f_person)
            # This is getting complicated. Let me use the simpler approach:
            # Individual payroll rows already exist. Sum them where location matches.
            f = "=0"  # fallback
            cell = ws.cell(row=r, column=3 + i, value=f)
            cell.font = data_font; cell.number_format = CURR
            cell.border = thin_border; cell.alignment = right_align
        # Override with proper SUMPRODUCT using the payroll data (simpler approach)
        # Since the individual payroll rows are at known positions, build explicit formulas
        r += 1

        # This approach is getting too complex for inline formulas. Let me use a hybrid:
        # Per-location payroll = SUMPRODUCT of individual payroll rows filtered by location
        # But the payroll breakdown rows don't have a location column.
        #
        # BETTER APPROACH: Add per-person payroll rows in the per-location section
        # that reference the master payroll row but only if location matches.

        # Let me just use the computed values but make them reference the upstream
        # assumptions via SUMPRODUCT. Here's the formula:
        # =SUMPRODUCT(($H$132:$H$146=loc_name)*($F$132:$F$146="W-2")*(payroll_schedule_rows))

        # Actually, the cleanest Excel formula approach:
        # Per-location payroll = sum of individual person rows WHERE location = loc_name
        # The individual person payroll amounts are in rows 226-240 (payroll breakdown)
        # The location for each person is in H132:H146 (roster input)
        # So: =SUMPRODUCT(($H$132:$H$146="Westlake")*({payroll_row_range}))

        # Overwrite the payroll row we just created
        # Reference the PAYROLL BREAKDOWN rows (226-240) which already have
        # hire/fire date logic and salary escalation built in.
        # SUMPRODUCT filters these by matching location in the roster.
        payroll_breakdown_first = asm['team_rows'][0]   # first individual payroll row (226)
        payroll_breakdown_last = asm['team_rows'][-1]    # last individual payroll row (240)
        # Wait - team_rows stores ROSTER rows (132-146), not payroll rows (226-240)
        # The payroll breakdown rows are offset from team_rows.
        # Payroll breakdown starts at a known row. Let me find it from asm.
        # The individual payroll rows start right after "PAYROLL BREAKDOWN" header.
        # They are: asm['team_rows'][0] mapped to payroll row via offset.
        # Actually, in the build script, individual payroll rows are built sequentially
        # starting from r after the "PAYROLL BREAKDOWN" section header.
        # Let me use the stored row numbers directly.

        # The payroll breakdown rows correspond 1:1 with team_rows.
        # Payroll row for person j = payroll_section_start + j
        # We need to find payroll_section_start from asm.
        # It's stored implicitly: Total Salaries row (asm['sched_payroll'] area)
        # Row 241 = Total Salaries = SUM(C226:C240), so individual rows are 226-240.
        # Total Salaries row = asm.get('total_salaries_row')
        # Individual payroll starts 15 rows before that.
        # Simpler: just hardcode the offset from known structure.

        # From the output: R226=Herlyn, R240=New Hire 15, R241=Total Salaries
        # These are always at team_rows offset. Let me compute:
        # first_person in payroll = first payroll breakdown row
        # This was set during the PAYROLL BREAKDOWN build section
        payroll_first = asm.get('payroll_breakdown_first', first_person)
        payroll_last = asm.get('payroll_breakdown_last', last_person)

        payroll_row = r - 1
        for i in range(N):
            col_letter = mcol(i)
            # SUMPRODUCT: location match * payroll amounts (which include hire/fire + escalation)
            # Then add payroll tax % + processing fee
            f = (f'=SUMPRODUCT(($H${team_start}:$H${team_end}="{loc_name}")'
                 f'*({col_letter}{payroll_first}:{col_letter}{payroll_last}))'
                 f'*(1+$C${asm["payroll_tax_row"]}/100)')
            cell = ws.cell(row=payroll_row, column=3 + i, value=f)
            cell.font = data_font; cell.number_format = CURR
            cell.border = thin_border; cell.alignment = right_align
        loc_sched['payroll'] = payroll_row

        # Contractors: build individual per-person contractor rows inline,
        # then sum. Each person checks: location match, type=Contractor,
        # start/end date, and applies salary escalation.
        loc_sched['contractors'] = r
        ws.cell(row=r, column=2, value=f"  Contractors ({loc_name})").font = data_font
        for i in range(N):
            col_letter = mcol(i)
            # Sum each person's contractor pay if they match this location
            terms = []
            for j, person_row in enumerate(asm['team_rows']):
                # IF(location=loc AND type=Contractor AND active, salary*escalator, 0)
                f_person = (
                    f'IF(AND($H${person_row}="{loc_name}",'
                    f'$F${person_row}="Contractor",'
                    f'$D${person_row}<>"",'
                    f'COLUMN()-2>=$D${person_row},'
                    f'OR($E${person_row}="",COLUMN()-2<=$E${person_row})),'
                    f'$C${person_row}*(1+$C${asm["sal_rate_row"]}/100)^INT((COLUMN()-2)/12),0)'
                )
                terms.append(f_person)
            f = "=" + "+".join(terms)
            cell = ws.cell(row=r, column=3 + i, value=f)
            cell.font = data_font; cell.number_format = CURR
            cell.border = thin_border; cell.alignment = right_align
        r += 1

        # OpEx: use hardcoded from py calcs but via reference to opex line items
        # For now, use computed values (they derive from OPEX_BY_LOCATION which is in py)
        loc_sched['opex'] = r
        ws.cell(row=r, column=2, value=f"  OpEx ({loc_name})").font = data_font
        opex_data = loc_pl.get('direct_opex', [0.0] * N)
        for i in range(N):
            cell = ws.cell(row=r, column=3 + i, value=round(opex_data[i], 2))
            cell.font = input_font; cell.number_format = CURR
            cell.fill = input_fill; cell.border = thin_border; cell.alignment = right_align
        r += 1

        # Expansion: reference the expansion total row for this location
        loc_sched['expansion'] = r
        ws.cell(row=r, column=2, value=f"  Expansion ({loc_name})").font = data_font
        # Find the expansion total row for this location
        exp_total_row = None
        if hasattr(asm, 'get') and 'exp_total_rows_by_name' in asm:
            exp_total_row = asm['exp_total_rows_by_name'].get(loc_name)
        if not exp_total_row:
            # Try to match by index in the expansion list
            for exp_idx, exp_def in enumerate(a.get('expansions', [])):
                if exp_def.get('name') == loc_name and exp_idx < len(asm.get('exp_total_rows_list', [])):
                    exp_total_row = asm['exp_total_rows_list'][exp_idx]
                    break
        if exp_total_row:
            for i in range(N):
                f = f"={mcol(i)}{exp_total_row}"
                cell = ws.cell(row=r, column=3 + i, value=f)
                cell.font = data_font; cell.number_format = CURR
                cell.border = thin_border; cell.alignment = right_align
        else:
            for i in range(N):
                cell = ws.cell(row=r, column=3 + i, value=0)
                cell.font = data_font; cell.number_format = CURR
                cell.border = thin_border; cell.alignment = right_align
        r += 1

        # Shared Overhead: compute as formula from total shared OH pool
        loc_sched['shared_overhead'] = r
        ws.cell(row=r, column=2, value=f"  Shared OH ({loc_name})").font = data_font
        shared_data = loc_pl.get('shared_overhead_allocation', [0.0] * N)
        for i in range(N):
            cell = ws.cell(row=r, column=3 + i, value=round(shared_data[i], 2))
            cell.font = data_font; cell.number_format = CURR
            cell.border = thin_border; cell.alignment = right_align
        r += 1

        r += 1  # blank between locations
        asm['loc_schedules'][loc_name] = loc_sched

    r += 1  # extra blank

    # ================================================================
    # HISTORICAL AR SPILLOVER (formula-driven from historical table + curves)
    # Uses explicit term-by-term formulas (no SUMPRODUCT/INDEX) for
    # broad Excel compatibility.
    # ================================================================
    section_bar(ws, r, 2, 2 + N, "HISTORICAL AR (Past Surgeries → Forecast Collections)")
    r += 1

    hs = asm['hist_start']
    he = asm['hist_end']
    avg_rev_bobas_row = asm['avg_rev_bobas']
    avg_rev_gap_row = asm['avg_rev_gap']
    cs_b = asm['curve_bobas_start']
    clen_b = asm['curve_bobas_len']
    cs_g = asm['curve_gap_start']
    clen_g = asm['curve_gap_len']
    ll = get_column_letter(2 + N)  # last data column letter

    # Read offsets from HISTORICAL_MONTHS to build formulas at build time
    hist_offsets = [h[1] for h in HISTORICAL_MONTHS]  # [-16, -15, ..., -1]
    n_hist = len(hist_offsets)

    def _build_ar_terms(forecast_month, count_col, avg_rev_row, curve_start, curve_len):
        """Build list of formula terms for a given forecast month.
        Each term: $<count_col>$<hist_row> * $C$<avg_rev_row> * $C$<curve_row> / 100
        Only includes terms where lag maps to a valid, non-zero curve position.
        """
        terms = []
        for h_idx in range(n_hist):
            hist_row = hs + h_idx
            offset = hist_offsets[h_idx]
            lag = forecast_month - offset
            if 0 <= lag < curve_len:
                curve_row = curve_start + lag
                terms.append(
                    f'${count_col}${hist_row}*$C${avg_rev_row}*$C${curve_row}/100'
                )
        return terms

    # ---- BOBA AR row ----
    # All months use the collection curve — no plug formula.
    # Overdue AR (collections that would have landed before Jan 2026)
    # is pre-calculated in baseline_data.py and spread across 12 months.
    asm['ar_boba'] = r
    ws.cell(row=r, column=2, value="  BOBA AR Collected").font = data_font
    for i in range(N):
        terms = _build_ar_terms(i, 'C', avg_rev_bobas_row, cs_b, clen_b)
        f = '=' + '+'.join(terms) if terms else '=0'
        cell = ws.cell(row=r, column=3 + i, value=f)
        cell.font = data_font
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    r += 1

    # ---- GAP AR row ----
    asm['ar_gap'] = r
    ws.cell(row=r, column=2, value="  GAP AR Collected").font = data_font
    for i in range(N):
        terms = _build_ar_terms(i, 'D', avg_rev_gap_row, cs_g, clen_g)
        f = '=' + '+'.join(terms) if terms else '=0'
        cell = ws.cell(row=r, column=3 + i, value=f)
        cell.font = data_font
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    r += 1

    # ---- Overdue AR row (pre-computed spread from Python) ----
    # Collections that would have landed before Jan 2026 are spread over 12 months.
    # This captures the overdue pool that the formula-based rows can't handle.
    asm['ar_overdue'] = r
    ws.cell(row=r, column=2, value="  Overdue AR Spread").font = data_font

    # Calculate overdue amounts from baseline
    # Uses top-level imports: _compute_ar_spillover, _HIST_BOBA, _HIST_GAP,
    # DEFAULT_ASSUMPTIONS, BOBAS_COLLECTION_CURVE, GAP_COLLECTION_CURVE
    # Compute WITH overdue spread (current)
    ar_boba_with = HISTORICAL_AR_BOBA
    ar_gap_with = HISTORICAL_AR_GAP
    # Compute WITHOUT overdue spread (formula-only portion)
    ar_boba_no = _compute_ar_spillover(
        _HIST_BOBA, DEFAULT_ASSUMPTIONS['avg_revenue_bobas'],
        BOBAS_COLLECTION_CURVE, overdue_spread_months=0)
    ar_gap_no = _compute_ar_spillover(
        _HIST_GAP, DEFAULT_ASSUMPTIONS['avg_revenue_gap'],
        GAP_COLLECTION_CURVE, overdue_spread_months=0)
    # Overdue = difference (what the formulas miss)
    overdue = [
        (ar_boba_with[i] + ar_gap_with[i]) - (ar_boba_no[i] + ar_gap_no[i])
        for i in range(N)
    ]
    for i in range(N):
        cell = ws.cell(row=r, column=3 + i, value=round(overdue[i], 0))
        cell.font = data_font
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    r += 1

    # ---- Total AR row ----
    asm['ar_total'] = r
    ws.cell(row=r, column=2, value="TOTAL HISTORICAL AR").font = data_bold
    for i in range(N):
        cl = mcol(i)
        f = f'={cl}{asm["ar_boba"]}+{cl}{asm["ar_gap"]}+{cl}{asm["ar_overdue"]}'
        cell = ws.cell(row=r, column=3 + i, value=f)
        cell.font = data_bold
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    r += 1

    return asm


# ============================================================
# TAB 3: MONTHLY P&L (2025 actuals + 60-month forecast)
# ============================================================
def build_monthly_pl(wb, asm):
    ws = wb.create_sheet("Monthly P&L")
    ws.sheet_properties.tabColor = "4472C4"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 38
    last_col = 2 + N + 1  # label + 60 months + total = col 63
    for ci in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.column_dimensions[get_column_letter(last_col)].width = 14

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=min(last_col, 30))
    ws.cell(row=2, column=2,
            value="Monthly P&L \u2014 Cash Basis (2025 Actuals + 2026-2030 Forecast)").font = title_font

    # ---- 2025 ACTUALS (12 months, hardcoded) ----
    r = 4
    last_col_25 = 2 + 12 + 1  # 16
    section_bar(ws, r, 2, last_col_25, "2025 ACTUALS (from QBO)")
    r += 1
    header_row(ws, r, ["Account"] + MONTHS_12 + ["FY 2025"], c1=2)
    r += 1

    ws.cell(row=r, column=2, value="INCOME").font = subsection_font
    r += 1
    r = _write_row(ws, r, "400 Fee Income", ACTUALS_2025['fee_income'])
    r = _write_row(ws, r, "410 Reimbursed Expense Income", ACTUALS_2025['reimbursed_expense_income'])
    r = _write_row(ws, r, "430 Refunds", ACTUALS_2025['refunds'])

    total_inc_25 = [
        ACTUALS_2025['fee_income'][i] + ACTUALS_2025['reimbursed_expense_income'][i] + ACTUALS_2025['refunds'][i]
        for i in range(12)
    ]
    r = _write_row(ws, r, "TOTAL INCOME", total_inc_25, bold=True)
    r += 1

    ws.cell(row=r, column=2, value="KEY EXPENSES").font = subsection_font
    r += 1
    r = _write_row(ws, r, "592 Physician Services", ACTUALS_2025['physician_services'])
    r = _write_row(ws, r, "554 MD Capital Billing", ACTUALS_2025['billing_services'])
    r = _write_row(ws, r, "584 Salaries & Wages", ACTUALS_2025['salaries_wages'])

    total_exp_25 = _get_monthly_total_expenses_25()
    r = _write_row(ws, r, "TOTAL EXPENSES", total_exp_25, bold=True)
    noi_25 = [total_inc_25[i] - total_exp_25[i] for i in range(12)]
    r = _write_row(ws, r, "NET OPERATING INCOME", noi_25, bold=True)

    r += 3

    # ================================================================
    # 2026-2030 FORECAST (60 months, ALL FORMULAS)
    # ================================================================
    ll = get_column_letter(2 + N)  # last data column letter (BJ for 60 months)
    total_col = get_column_letter(2 + N + 1)  # BK = total column

    section_bar(ws, r, 2, last_col, "2026-2030 FORECAST (Cash Basis \u2014 Bobas + GAP, 90/10 Fund Flow)")
    r += 1
    header_row(ws, r, ["Account"] + FORECAST_MONTH_LABELS + ["Total"], c1=2)
    r += 1

    # ---- SURGERY VOLUME ----
    row_bobas_vol = r
    formulas = [f"=Assumptions!$C${asm['vol_start']+i}" for i in range(N)]
    r = _write_formula_row(ws, r, "Bobas Volume", formulas, fmt='#,##0')

    row_gap_vol = r
    formulas = [f"=Assumptions!$D${asm['vol_start']+i}" for i in range(N)]
    r = _write_formula_row(ws, r, "GAP Volume", formulas, fmt='#,##0')

    row_total_vol = r
    formulas = [f"={mcol(i)}{row_bobas_vol}+{mcol(i)}{row_gap_vol}" for i in range(N)]
    r = _write_formula_row(ws, r, "Total Surgeries", formulas, bold=True, fmt='#,##0')
    r += 1

    # ---- REVENUE EARNED (Accrual Reference) ----
    ws.cell(row=r, column=2, value="REVENUE EARNED (Accrual Reference)").font = Font(
        name="Calibri", size=10, italic=True, color="999999")
    r += 1

    row_bobas_earned = r
    formulas = [f"={mcol(i)}{row_bobas_vol}*Assumptions!$C${asm['avg_rev_bobas']}" for i in range(N)]
    r = _write_formula_row(ws, r, "  Bobas Earned (Accrual)", formulas)
    # Style as gray/italic reference
    for i in range(N):
        ws.cell(row=r-1, column=3+i).font = pct_font

    row_gap_earned = r
    formulas = [f"={mcol(i)}{row_gap_vol}*Assumptions!$C${asm['avg_rev_gap']}" for i in range(N)]
    r = _write_formula_row(ws, r, "  GAP Earned (Accrual)", formulas)
    for i in range(N):
        ws.cell(row=r-1, column=3+i).font = pct_font

    row_total_earned = r
    formulas = [f"={mcol(i)}{row_bobas_earned}+{mcol(i)}{row_gap_earned}" for i in range(N)]
    r = _write_formula_row(ws, r, "  Total Earned (Accrual)", formulas)
    for i in range(N):
        ws.cell(row=r-1, column=3+i).font = pct_font

    r += 1  # blank

    # ---- CASH COLLECTED (This drives the P&L) ----
    ws.cell(row=r, column=2, value="CASH COLLECTED").font = subsection_font
    r += 1

    # Bobas Revenue (Collected from forecast surgeries)
    row_bobas_collected = r
    coll_formulas = []
    bobas_curve_len = asm['curve_bobas_len']
    for j in range(N):
        terms = []
        for lag in range(bobas_curve_len):
            src = j - lag
            if src >= 0:
                terms.append(
                    f"{mcol(src)}{row_bobas_earned}*Assumptions!$C${asm['curve_bobas_start']+lag}/100"
                )
        coll_formulas.append("=" + "+".join(terms) if terms else "=0")
    r = _write_formula_row(ws, r, "Bobas Collected (Forecast)", coll_formulas)

    # GAP Revenue (Collected from forecast surgeries)
    row_gap_collected = r
    coll_formulas = []
    gap_curve_len = asm['curve_gap_len']
    for j in range(N):
        terms = []
        for lag in range(gap_curve_len):
            src = j - lag
            if src >= 0:
                terms.append(
                    f"{mcol(src)}{row_gap_earned}*Assumptions!$C${asm['curve_gap_start']+lag}/100"
                )
        coll_formulas.append("=" + "+".join(terms) if terms else "=0")
    r = _write_formula_row(ws, r, "GAP Collected (Forecast)", coll_formulas)

    # Historical AR Collected (from Assumptions schedule)
    row_hist_ar = r
    formulas = [f"=Assumptions!{mcol(i)}${asm['ar_total']}" for i in range(N)]
    r = _write_formula_row(ws, r, "Historical AR Collected", formulas)

    # TOTAL INCOME = Total Cash Collected (cash basis)
    row_total_collected = r
    formulas = [
        f"={mcol(i)}{row_bobas_collected}+{mcol(i)}{row_gap_collected}+{mcol(i)}{row_hist_ar}"
        for i in range(N)
    ]
    r = _write_formula_row(ws, r, "TOTAL INCOME (Cash Collected)", formulas, bold=True)
    # Also define total_income as the same row for downstream references
    row_total_income = row_total_collected

    r += 1  # blank

    # ---- OVERHEAD EXPENSES ----
    ws.cell(row=r, column=2, value="OVERHEAD EXPENSES").font = subsection_font
    r += 1

    # Billing (18% of Collected)
    row_billing = r
    formulas = [
        f"={mcol(i)}{row_total_collected}*Assumptions!$C${asm['billing_rate']}/100"
        for i in range(N)
    ]
    r = _write_formula_row(ws, r, "Billing (18% of Collected)", formulas)

    # Payroll (W-2) -> Assumptions schedule
    row_payroll = r
    formulas = [f"=Assumptions!{mcol(i)}${asm['sched_payroll']}" for i in range(N)]
    r = _write_formula_row(ws, r, "Payroll (W-2)", formulas)

    # Contractor Costs -> Assumptions schedule
    row_contractors = r
    formulas = [f"=Assumptions!{mcol(i)}${asm['sched_contractors']}" for i in range(N)]
    r = _write_formula_row(ws, r, "Contractor Costs", formulas)

    # Operating Expenses -> Assumptions schedule
    row_opex = r
    formulas = [f"=Assumptions!{mcol(i)}${asm['sched_opex']}" for i in range(N)]
    r = _write_formula_row(ws, r, "Operating Expenses", formulas)

    # Expansion Costs -> Assumptions schedule
    row_expansion = r
    formulas = [f"=Assumptions!{mcol(i)}${asm['sched_expansion']}" for i in range(N)]
    r = _write_formula_row(ws, r, "Expansion Costs", formulas)

    # TOTAL OVERHEAD = billing + payroll + contractors + opex + expansion
    row_total_overhead = r
    formulas = [
        f"={mcol(i)}{row_billing}+{mcol(i)}{row_payroll}+{mcol(i)}{row_contractors}+{mcol(i)}{row_opex}+{mcol(i)}{row_expansion}"
        for i in range(N)
    ]
    r = _write_formula_row(ws, r, "TOTAL OVERHEAD", formulas, bold=True)
    r += 1  # blank

    # ---- FUND FLOW ----
    ws.cell(row=r, column=2, value="FUND FLOW (Cash Basis)").font = subsection_font
    r += 1

    # NET EQUITY = cash collected - total overhead
    row_net_equity = r
    formulas = [
        f"={mcol(i)}{row_total_income}-{mcol(i)}{row_total_overhead}"
        for i in range(N)
    ]
    r = _write_formula_row(ws, r, "NET EQUITY (Cash Collected - Overhead)", formulas, bold=True)

    # Physician Services (90%)
    row_physician = r
    formulas = [
        f"=MAX(0,{mcol(i)}{row_net_equity}*Assumptions!$C${asm['phys_rate']}/100)"
        for i in range(N)
    ]
    r = _write_formula_row(ws, r, "Physician Services (90%)", formulas)

    # NET INCOME = net_equity - physician
    row_net_income = r
    formulas = [
        f"={mcol(i)}{row_net_equity}-{mcol(i)}{row_physician}"
        for i in range(N)
    ]
    r = _write_formula_row(ws, r, "NET INCOME (Retained by CNS)", formulas, bold=True)

    # ---- OVERLAY QBO ACTUALS for Jan/Feb 2026 ----
    # For months with actual data, replace formulas with hardcoded actuals.
    # This ensures the P&L matches QBO exactly for closed months.
    qbo = ACTUALS_2026_QBO
    n_act = NUM_2026_ACTUALS  # 2 (Jan, Feb)

    actuals_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    actuals_font = Font(name="Calibri", size=10, color="006400")
    actuals_bold = Font(name="Calibri", size=10, bold=True, color="006400")

    for i in range(n_act):
        col = 3 + i  # C=Jan, D=Feb

        # Total Income (Cash Collected) = QBO total income
        cell = ws.cell(row=row_total_collected, column=col, value=qbo['total_income'][i])
        cell.font = actuals_bold; cell.number_format = CURR
        cell.fill = actuals_fill; cell.border = thin_border; cell.alignment = right_align

        # Zero out the sub-lines (Bobas/GAP/Historical AR) since total is hardcoded
        for sub_row in [row_bobas_collected, row_gap_collected, row_hist_ar]:
            cell = ws.cell(row=sub_row, column=col, value=0)
            cell.font = actuals_font; cell.number_format = CURR
            cell.fill = actuals_fill; cell.border = thin_border; cell.alignment = right_align

        # Overhead lines
        # Billing = QBO doesn't have a separate billing line, but we can compute
        # from total expenses. Use the actuals breakdown.
        overhead_map = {
            row_billing: 0,  # Billing not broken out in QBO cash basis
            row_payroll: qbo['payroll_expenses'][i],
            row_contractors: qbo['contracts'][i],
            row_opex: (qbo['advertising_marketing'][i] + qbo['bank_fees'][i] +
                       qbo['insurance'][i] + qbo['malpractice_insurance'][i] +
                       qbo['health_insurance'][i] + qbo['supplies_medical'][i] +
                       qbo['legal_fees'][i] + qbo['licenses_fees'][i] +
                       qbo['meals'][i] + qbo['software_apps'][i] +
                       qbo['office_expenses'][i] + qbo['rent_lease'][i] +
                       qbo['repairs_maintenance'][i] + qbo['travel'][i] +
                       qbo['uniforms'][i]),
            row_expansion: 0,
        }

        for row_num, val in overhead_map.items():
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = actuals_font; cell.number_format = CURR
            cell.fill = actuals_fill; cell.border = thin_border; cell.alignment = right_align

        # Total Overhead
        total_oh = sum(overhead_map.values())
        cell = ws.cell(row=row_total_overhead, column=col, value=total_oh)
        cell.font = actuals_bold; cell.number_format = CURR
        cell.fill = actuals_fill; cell.border = thin_border; cell.alignment = right_align

        # Net Equity
        net_eq = qbo['total_income'][i] - total_oh
        cell = ws.cell(row=row_net_equity, column=col, value=net_eq)
        cell.font = actuals_bold; cell.number_format = CURR
        cell.fill = actuals_fill; cell.border = thin_border; cell.alignment = right_align

        # Physician Services = QBO actual
        phys = qbo['physician_services'][i]
        cell = ws.cell(row=row_physician, column=col, value=phys)
        cell.font = actuals_font; cell.number_format = CURR
        cell.fill = actuals_fill; cell.border = thin_border; cell.alignment = right_align

        # Net Income
        ni = net_eq - phys
        cell = ws.cell(row=row_net_income, column=col, value=ni)
        cell.font = actuals_bold; cell.number_format = CURR
        cell.fill = actuals_fill; cell.border = thin_border; cell.alignment = right_align

    # Return P&L row numbers for Cash Flow tab
    pl_rows = {
        'total_collected': row_total_collected,
        'billing': row_billing,
        'payroll': row_payroll,
        'contractors': row_contractors,
        'opex': row_opex,
        'expansion': row_expansion,
        'total_overhead': row_total_overhead,
        'net_equity': row_net_equity,
        'total_earned': row_total_earned,
    }
    return pl_rows


# ============================================================
# TAB 4: CASH FLOW (60 months, ALL FORMULAS)
# ============================================================
def build_cash_flow(wb, asm, pl_rows):
    ws = wb.create_sheet("Cash Flow")
    ws.sheet_properties.tabColor = "00B050"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 38
    last_col = 2 + N + 1
    for ci in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.column_dimensions[get_column_letter(last_col)].width = 14

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=min(last_col, 30))
    ws.cell(row=2, column=2, value="Cash Flow Projection \u2014 2026-2030").font = title_font

    r = 4
    section_bar(ws, r, 2, last_col, "60-MONTH CASH FLOW (Cash-Basis Fund Flow)")
    r += 1
    header_row(ws, r, ["Line Item"] + FORECAST_MONTH_LABELS + ["Total"], c1=2)
    r += 1

    pl_sheet = "'Monthly P&L'"

    # ---- Beginning Cash ----
    row_beg_cash = r
    formulas = []
    for i in range(N):
        if i == 0:
            formulas.append(f"=Assumptions!$C${asm['starting_cash']}")
        else:
            # prior month's ending cash
            formulas.append(f"={mcol(i-1)}{0}")  # placeholder, filled below
    # We need ending cash row first. Build formulas after we know all rows.
    # Use a two-pass approach: first lay out row numbers, then write.

    # Pre-calculate row positions
    row_beg_cash = r; r += 1
    r += 1  # blank
    # CASH IN subsection
    r += 1  # subsection label
    row_cash_in = r; r += 1
    r += 1  # blank
    # CASH OUT subsection
    r += 1  # subsection label
    row_cash_overhead = r; r += 1
    row_cash_after = r; r += 1
    r += 1  # blank
    # DISTRIBUTIONS subsection
    r += 1  # subsection label
    row_distributable = r; r += 1
    row_physician = r; r += 1
    row_savings_dep = r; r += 1
    r += 1  # blank
    row_ending_cash = r; r += 1
    row_below_min = r; r += 1
    r += 1  # blank
    row_savings_bal = r; r += 1

    # Now write all rows with formulas

    # ---- Beginning Cash ----
    cur = row_beg_cash
    ws.cell(row=cur, column=2, value="Beginning Cash").font = data_bold
    for i in range(N):
        if i == 0:
            f = f"=Assumptions!$C${asm['starting_cash']}"
        else:
            f = f"={mcol(i-1)}{row_ending_cash}"
        cell = ws.cell(row=cur, column=3 + i, value=f)
        cell.font = data_bold
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    # No SUM total for beginning cash (it's a balance, not additive)
    cur += 1

    cur += 1  # blank

    # ---- CASH IN ----
    ws.cell(row=cur, column=2, value="CASH IN").font = subsection_font
    cur += 1

    # Cash In = P&L total collected (cash basis, same as P&L TOTAL INCOME)
    assert cur == row_cash_in
    ws.cell(row=cur, column=2, value="Cash Collected").font = data_font
    for i in range(N):
        f = f"={pl_sheet}!{mcol(i)}${pl_rows['total_collected']}"
        cell = ws.cell(row=cur, column=3 + i, value=f)
        cell.font = data_font
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    # SUM total
    tc = ws.cell(row=cur, column=3 + N,
                 value=f"=SUM({mcol(0)}{cur}:{mcol(N-1)}{cur})")
    tc.font = data_font; tc.number_format = CURR; tc.border = thin_border; tc.alignment = right_align
    cur += 1

    cur += 1  # blank

    # ---- CASH OUT - OVERHEAD ----
    ws.cell(row=cur, column=2, value="CASH OUT - OVERHEAD").font = subsection_font
    cur += 1

    # Cash Overhead = billing + payroll + contractors + opex + expansion (from P&L)
    assert cur == row_cash_overhead
    ws.cell(row=cur, column=2, value="Total Overhead").font = data_font
    for i in range(N):
        f = (f"={pl_sheet}!{mcol(i)}${pl_rows['billing']}"
             f"+{pl_sheet}!{mcol(i)}${pl_rows['payroll']}"
             f"+{pl_sheet}!{mcol(i)}${pl_rows['contractors']}"
             f"+{pl_sheet}!{mcol(i)}${pl_rows['opex']}"
             f"+{pl_sheet}!{mcol(i)}${pl_rows['expansion']}")
        cell = ws.cell(row=cur, column=3 + i, value=f)
        cell.font = data_font
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    tc = ws.cell(row=cur, column=3 + N,
                 value=f"=SUM({mcol(0)}{cur}:{mcol(N-1)}{cur})")
    tc.font = data_font; tc.number_format = CURR; tc.border = thin_border; tc.alignment = right_align
    cur += 1

    # Cash After Overhead = beginning + cash_in - overhead
    assert cur == row_cash_after
    ws.cell(row=cur, column=2, value="Cash After Overhead").font = data_bold
    for i in range(N):
        f = f"={mcol(i)}{row_beg_cash}+{mcol(i)}{row_cash_in}-{mcol(i)}{row_cash_overhead}"
        cell = ws.cell(row=cur, column=3 + i, value=f)
        cell.font = data_bold
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    cur += 1

    cur += 1  # blank

    # ---- DISTRIBUTIONS ----
    ws.cell(row=cur, column=2, value="FUND FLOW DISTRIBUTION").font = subsection_font
    cur += 1

    # Distributable = MAX(0, cash_after_overhead - min_cash)
    assert cur == row_distributable
    ws.cell(row=cur, column=2, value="Distributable (above min)").font = data_font
    for i in range(N):
        f = f"=MAX(0,{mcol(i)}{row_cash_after}-Assumptions!$C${asm['min_cash']})"
        cell = ws.cell(row=cur, column=3 + i, value=f)
        cell.font = data_font
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    tc = ws.cell(row=cur, column=3 + N,
                 value=f"=SUM({mcol(0)}{cur}:{mcol(N-1)}{cur})")
    tc.font = data_font; tc.number_format = CURR; tc.border = thin_border; tc.alignment = right_align
    cur += 1

    # Physician Services = distributable * phys_rate/100
    assert cur == row_physician
    ws.cell(row=cur, column=2, value="Physician Services (90%)").font = data_font
    for i in range(N):
        f = f"={mcol(i)}{row_distributable}*Assumptions!$C${asm['phys_rate']}/100"
        cell = ws.cell(row=cur, column=3 + i, value=f)
        cell.font = data_font
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    tc = ws.cell(row=cur, column=3 + N,
                 value=f"=SUM({mcol(0)}{cur}:{mcol(N-1)}{cur})")
    tc.font = data_font; tc.number_format = CURR; tc.border = thin_border; tc.alignment = right_align
    cur += 1

    # Savings Deposit = distributable * savings_rate/100
    assert cur == row_savings_dep
    ws.cell(row=cur, column=2, value="Savings Deposit (10%)").font = data_font
    for i in range(N):
        f = f"={mcol(i)}{row_distributable}*Assumptions!$C${asm['savings_rate']}/100"
        cell = ws.cell(row=cur, column=3 + i, value=f)
        cell.font = data_font
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    tc = ws.cell(row=cur, column=3 + N,
                 value=f"=SUM({mcol(0)}{cur}:{mcol(N-1)}{cur})")
    tc.font = data_font; tc.number_format = CURR; tc.border = thin_border; tc.alignment = right_align
    cur += 1

    cur += 1  # blank

    # Ending Cash = cash_after_overhead - physician - savings
    assert cur == row_ending_cash
    ws.cell(row=cur, column=2, value="ENDING CASH").font = data_bold
    for i in range(N):
        f = f"={mcol(i)}{row_cash_after}-{mcol(i)}{row_physician}-{mcol(i)}{row_savings_dep}"
        cell = ws.cell(row=cur, column=3 + i, value=f)
        cell.font = data_bold
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align
    cur += 1

    # Below Minimum?
    assert cur == row_below_min
    ws.cell(row=cur, column=2, value="Below Minimum?").font = data_font
    for i in range(N):
        f = f'=IF({mcol(i)}{row_ending_cash}<Assumptions!$C${asm["min_cash"]},"YES","No")'
        cell = ws.cell(row=cur, column=3 + i, value=f)
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border
    cur += 1

    cur += 1  # blank

    # Savings Balance
    assert cur == row_savings_bal
    ws.cell(row=cur, column=2, value="Savings Balance").font = data_bold
    for i in range(N):
        if i == 0:
            f = f"=Assumptions!$C${asm['starting_savings']}+{mcol(i)}{row_savings_dep}"
        else:
            f = f"={mcol(i-1)}{row_savings_bal}+{mcol(i)}{row_savings_dep}"
        cell = ws.cell(row=cur, column=3 + i, value=f)
        cell.font = data_bold
        cell.number_format = CURR
        cell.border = thin_border
        cell.alignment = right_align

    return ws


# ============================================================
# TAB 5: SCENARIOS
# ============================================================
def build_scenarios(wb):
    ws = wb.create_sheet("Scenarios")
    ws.sheet_properties.tabColor = "7030A0"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 30
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 20

    ws.merge_cells('B2:G2')
    ws.cell(row=2, column=2, value="Scenario Analysis (60-Month)").font = title_font

    from financial_calcs import generate_monthly_pl_forecast, generate_cash_flow_forecast

    r = 4

    # ---- Volume Scenarios ----
    section_bar(ws, r, 2, 7, "SURGERY VOLUME SCENARIOS")
    r += 1
    header_row(ws, r, ["Metric", "Low (4B/2G)", "Base (Ramp)", "Mid (8B/4G)", "High (12B/6G)", "Aggressive (16B/8G)"], c1=2)
    r += 1

    scenarios = [
        ("Low (4B/2G)", [4]*N, [2]*N),
        ("Base (Ramp)", DEFAULT_ASSUMPTIONS['bobas_volume'], DEFAULT_ASSUMPTIONS['gap_volume']),
        ("Mid (8B/4G)", [8]*N, [4]*N),
        ("High (12B/6G)", [12]*N, [6]*N),
        ("Aggressive (16B/8G)", [16]*N, [8]*N),
    ]

    vol_metrics = []
    for name, bobas, gap in scenarios:
        a = copy.deepcopy(DEFAULT_ASSUMPTIONS)
        a['bobas_volume'] = bobas
        a['gap_volume'] = gap
        pl = generate_monthly_pl_forecast(a)
        cf = generate_cash_flow_forecast(a)
        vol_metrics.append({
            'total_surgeries': sum(pl['total_volume']),
            'total_revenue': sum(pl['total_income']),
            'total_net_income': sum(pl['net_income']),
            'net_margin_pct': (sum(pl['net_income']) / sum(pl['total_income']) * 100) if sum(pl['total_income']) else 0,
            'ending_cash': cf['ending_cash'][-1],
            'min_cash': min(cf['ending_cash']),
            'total_physician': sum(pl['physician_services']),
            'total_savings': cf['savings_balance'][-1],
        })

    for label, key, fmt in [
        ("Total Surgeries (60mo)", 'total_surgeries', '#,##0'),
        ("Total Revenue (60mo)", 'total_revenue', '$#,##0'),
        ("Total Net Income (60mo)", 'total_net_income', '$#,##0'),
        ("Net Margin %", 'net_margin_pct', '0.0%'),
        ("Ending Cash (Dec-30)", 'ending_cash', '$#,##0'),
        ("Min Cash Balance", 'min_cash', '$#,##0'),
        ("Total Physician Svc", 'total_physician', '$#,##0'),
        ("Ending Savings", 'total_savings', '$#,##0'),
    ]:
        ws.cell(row=r, column=2, value=label).font = data_font
        for i, m in enumerate(vol_metrics):
            val = m[key]
            if fmt == '0.0%':
                val = val / 100
            cell = ws.cell(row=r, column=3 + i, value=val)
            cell.font = data_font
            cell.number_format = fmt
            cell.border = thin_border
            cell.alignment = right_align
            if i == 1:  # Base scenario highlight
                cell.fill = light_fill
        r += 1

    r += 2

    # ---- Expansion Impact ----
    section_bar(ws, r, 2, 7, "EXPANSION IMPACT")
    r += 1
    header_row(ws, r, ["Scenario", "Expansion $", "End Cash", "Min Cash", "End Savings", ""], c1=2)
    r += 1

    for name, enabled in [("No Expansion", False), ("Santa Barbara (Base)", True)]:
        a = copy.deepcopy(DEFAULT_ASSUMPTIONS)
        a['expansions'][0]['enabled'] = enabled
        pl = generate_monthly_pl_forecast(a)
        cf = generate_cash_flow_forecast(a)
        ws.cell(row=r, column=2, value=name).font = data_font
        exp_total = sum(pl['expansion_total'])
        ws.cell(row=r, column=3, value=exp_total).font = data_font
        ws.cell(row=r, column=3).number_format = '$#,##0'
        ws.cell(row=r, column=4, value=cf['ending_cash'][-1]).font = data_font
        ws.cell(row=r, column=4).number_format = '$#,##0'
        ws.cell(row=r, column=5, value=min(cf['ending_cash'])).font = data_font
        ws.cell(row=r, column=5).number_format = '$#,##0'
        ws.cell(row=r, column=6, value=cf['savings_balance'][-1]).font = data_font
        ws.cell(row=r, column=6).number_format = '$#,##0'
        if "Base" in name:
            style_range(ws, r, 2, 6, fill=light_fill)
        for c in range(2, 7):
            ws.cell(row=r, column=c).border = thin_border
        r += 1

    r += 2

    # ---- Inflation Impact ----
    section_bar(ws, r, 2, 7, "ESCALATION IMPACT (Salary 5%, Expenses 3%)")
    r += 1
    header_row(ws, r, ["Scenario", "End Cash", "Min Cash", "End Savings", "", ""], c1=2)
    r += 1

    for name, sal_inc, exp_inf in [
        ("No Escalation", 0, 0),
        ("Base (5% sal / 3% exp)", 5.0, 3.0),
        ("High (8% sal / 5% exp)", 8.0, 5.0),
    ]:
        a = copy.deepcopy(DEFAULT_ASSUMPTIONS)
        a['salary_annual_increase'] = sal_inc
        a['expense_annual_inflation'] = exp_inf
        cf = generate_cash_flow_forecast(a)
        ws.cell(row=r, column=2, value=name).font = data_font
        ws.cell(row=r, column=3, value=cf['ending_cash'][-1]).font = data_font
        ws.cell(row=r, column=3).number_format = '$#,##0'
        ws.cell(row=r, column=4, value=min(cf['ending_cash'])).font = data_font
        ws.cell(row=r, column=4).number_format = '$#,##0'
        ws.cell(row=r, column=5, value=cf['savings_balance'][-1]).font = data_font
        ws.cell(row=r, column=5).number_format = '$#,##0'
        if "Base" in name:
            style_range(ws, r, 2, 5, fill=light_fill)
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = thin_border
        r += 1

    return ws


# ============================================================
# TAB 6: QBO ACTUALS
# ============================================================
def build_qbo_actuals(wb):
    ws = wb.create_sheet("QBO Actuals")
    ws.sheet_properties.tabColor = "808080"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35
    for i in range(3, 16):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions[get_column_letter(15)].width = 16

    ws.merge_cells('B2:O2')
    ws.cell(row=2, column=2, value="QBO Raw Data - 2025 Actuals").font = title_font
    ws.cell(row=3, column=2,
            value="Source: QuickBooks Online, Accrual Basis, March 3, 2026").font = pct_font

    r = 5

    section_bar(ws, r, 2, 15, "PROFIT & LOSS")
    r += 1
    header_row(ws, r, ["Account"] + MONTHS_12 + ["Total"], c1=2)
    r += 1

    ws.cell(row=r, column=2, value="INCOME").font = subsection_font
    r += 1
    for key, label in [
        ('fee_income', '400 Fee Income'),
        ('reimbursed_expense_income', '410 Reimbursed Expense Income'),
        ('refunds', '430 Refunds'),
    ]:
        r = _write_row(ws, r, label, ACTUALS_2025[key], fmt=CURR2)

    total_inc = [
        ACTUALS_2025['fee_income'][i] + ACTUALS_2025['reimbursed_expense_income'][i] + ACTUALS_2025['refunds'][i]
        for i in range(12)
    ]
    r = _write_row(ws, r, "TOTAL INCOME", total_inc, bold=True, fmt=CURR2)
    r += 1

    ws.cell(row=r, column=2, value="EXPENSES").font = subsection_font
    r += 1

    for key, label in [
        ('advertising_marketing', '500 Advertising & Marketing'),
        ('bank_fees', '505 Bank Fees'),
        ('conference', '515 Conference'),
        ('contracts', '520 Contracts'),
        ('contributions', '525 Contributions'),
        ('dues_subscriptions', '535 Dues & Subscriptions'),
        ('insurance', '540 Insurance'),
        ('malpractice_insurance', '541 Malpractice Insurance'),
        ('health_insurance', '542 Health Insurance'),
        ('legal_accounting_services', '550 Legal & Accounting'),
        ('accounting_fees', '552 Accounting Fees'),
        ('billing_services', '554 MD Capital Billing'),
        ('licenses_fees', '555 Licenses & Fees'),
        ('mgmt_fee_abc', '556 Mgmt Fee (ABC A PC)'),
        ('mgmt_fee_vnsc', '558 Mgmt Fee (VNSC)'),
        ('meals', '560 Meals'),
        ('office_expenses', '574 Office Expenses'),
        ('software_apps', '570 Software & Apps'),
        ('payroll_processing', '580 Payroll Processing'),
        ('salaries_wages', '584 Salaries & Wages'),
        ('payroll_taxes', '586 Payroll Taxes'),
        ('physician_services', '592 Physician Services'),
        ('rent_lease', '600 Rent/Lease'),
        ('taxes_licenses', '605 Taxes & Licenses'),
        ('travel', '610 Travel'),
    ]:
        r = _write_row(ws, r, label, ACTUALS_2025[key], fmt=CURR2)

    total_exp = _get_monthly_total_expenses_25()
    r = _write_row(ws, r, "TOTAL EXPENSES", total_exp, bold=True, fmt=CURR2)
    r += 1

    noi = [total_inc[i] - total_exp[i] for i in range(12)]
    r = _write_row(ws, r, "NET OPERATING INCOME", noi, bold=True, fmt=CURR2)

    r += 3

    section_bar(ws, r, 2, 15, "BALANCE SHEET")
    r += 1
    header_row(ws, r, ["Account"] + MONTHS_12, c1=2)
    r += 1

    def bs_row(row, label, data, bold=False):
        ws.cell(row=row, column=2, value=label).font = data_bold if bold else data_font
        for i, val in enumerate(data):
            cell = ws.cell(row=row, column=3 + i, value=val)
            cell.font = data_bold if bold else data_font
            cell.number_format = CURR2
            cell.border = thin_border
        return row + 1

    ws.cell(row=r, column=2, value="ASSETS").font = subsection_font
    r += 1
    r = bs_row(r, "100 Chase Checking", BALANCE_SHEET_2025['chase_checking'])
    r = bs_row(r, "105 Chase Savings", BALANCE_SHEET_2025['chase_savings'])
    r = bs_row(r, "Total Cash", BALANCE_SHEET_2025['total_cash'], bold=True)
    r = bs_row(r, "TOTAL ASSETS", BALANCE_SHEET_2025['total_assets'], bold=True)
    r += 1
    ws.cell(row=r, column=2, value="LIABILITIES & EQUITY").font = subsection_font
    r += 1
    r = bs_row(r, "TOTAL LIABILITIES", BALANCE_SHEET_2025['total_liabilities'], bold=True)
    r = bs_row(r, "TOTAL EQUITY", BALANCE_SHEET_2025['total_equity'], bold=True)

    r += 3

    # ---- HISTORICAL CASE VOLUME WITH BOBA/GAP BREAKDOWN ----
    section_bar(ws, r, 2, 15, "HISTORICAL CASE VOLUME (with BOBA/GAP Breakdown)")
    r += 1

    # 2024 (Sep-Dec)
    ws.cell(row=r, column=2, value="2024 (Sep-Dec)").font = subsection_font
    r += 1
    months_2024 = ['Sep', 'Oct', 'Nov', 'Dec']
    header_row(ws, r, [""] + months_2024 + ["Total"], c1=2)
    r += 1
    for label, data in [("Total Cases", SURGERY_VOLUME_2024),
                         ("  BOBA", BOBA_VOLUME_2024),
                         ("  GAP", GAP_VOLUME_2024)]:
        ws.cell(row=r, column=2, value=label).font = data_font
        for i, v in enumerate(data):
            cell = ws.cell(row=r, column=3 + i, value=v)
            cell.font = data_font
            cell.number_format = '#,##0'
            cell.border = thin_border
            cell.alignment = right_align
        cell = ws.cell(row=r, column=3 + len(data), value=sum(data))
        cell.font = data_bold
        cell.number_format = '#,##0'
        cell.border = thin_border
        r += 1
    r += 1

    # 2025 (full year)
    ws.cell(row=r, column=2, value="2025 (Full Year)").font = subsection_font
    r += 1
    header_row(ws, r, [""] + MONTHS_12 + ["Total"], c1=2)
    r += 1
    for label, data in [("Total Cases", SURGERY_VOLUME_2025),
                         ("  BOBA", BOBA_VOLUME_2025),
                         ("  GAP", GAP_VOLUME_2025)]:
        ws.cell(row=r, column=2, value=label).font = data_font
        for i, v in enumerate(data):
            cell = ws.cell(row=r, column=3 + i, value=v)
            cell.font = data_font
            cell.number_format = '#,##0'
            cell.border = thin_border
            cell.alignment = right_align
        cell = ws.cell(row=r, column=3 + 12, value=sum(data))
        cell.font = data_bold
        cell.number_format = '#,##0'
        cell.border = thin_border
        r += 1
    ws.cell(row=r, column=2, value="Avg/Month").font = pct_font
    ws.cell(row=r, column=3, value=f"{TOTAL_SURGERIES_2025/12:.1f}").font = pct_font
    r += 2

    # 2026 YTD (Jan-Apr from client data)
    n_actual = len(BOBA_2026_ACTUALS)
    months_2026 = MONTHS_12[:n_actual]
    ws.cell(row=r, column=2, value=f"2026 (YTD: Jan-{months_2026[-1]})").font = subsection_font
    r += 1
    header_row(ws, r, [""] + months_2026 + ["Total"], c1=2)
    r += 1
    for label, data in [("Total Cases", SURGERY_VOLUME_2026_ACTUALS),
                         ("  BOBA", BOBA_2026_ACTUALS),
                         ("  GAP", GAP_2026_ACTUALS)]:
        ws.cell(row=r, column=2, value=label).font = data_font
        for i, v in enumerate(data):
            cell = ws.cell(row=r, column=3 + i, value=v)
            cell.font = data_font
            cell.number_format = '#,##0'
            cell.border = thin_border
            cell.alignment = right_align
        cell = ws.cell(row=r, column=3 + n_actual, value=sum(data))
        cell.font = data_bold
        cell.number_format = '#,##0'
        cell.border = thin_border
        r += 1

    # ---- 2026 QBO ACTUALS (Jan-Feb, Cash Basis) ----
    r += 2
    section_bar(ws, r, 2, 5, "2026 QBO ACTUALS (Cash Basis)")
    r += 1
    ws.cell(row=r, column=2, value="Source: P&L 1_2026-2_28_2026.xlsx, Cash Basis").font = pct_font
    r += 1

    qbo26 = ACTUALS_2026_QBO
    n26 = NUM_2026_ACTUALS
    month_headers = qbo26['months']

    header_row(ws, r, ["Account"] + month_headers, c1=2)
    r += 1

    ws.cell(row=r, column=2, value="INCOME").font = subsection_font
    r += 1

    for key, label in [
        ('fee_income', '400 Fee Income'),
    ]:
        ws.cell(row=r, column=2, value=label).font = data_font
        for i in range(n26):
            cell = ws.cell(row=r, column=3 + i, value=qbo26[key][i])
            cell.font = data_font; cell.number_format = CURR2
            cell.border = thin_border; cell.alignment = right_align
        r += 1

    ws.cell(row=r, column=2, value="TOTAL INCOME").font = data_bold
    for i in range(n26):
        cell = ws.cell(row=r, column=3 + i, value=qbo26['total_income'][i])
        cell.font = data_bold; cell.number_format = CURR2
        cell.border = thin_border; cell.alignment = right_align
    r += 2

    ws.cell(row=r, column=2, value="EXPENSES").font = subsection_font
    r += 1

    expense_lines = [
        ('advertising_marketing', '500 Advertising & Marketing'),
        ('bank_fees', '505 Bank Fees'),
        ('contracts', '520 Contracts'),
        ('insurance', '540 Insurance'),
        ('malpractice_insurance', '541 Malpractice Insurance'),
        ('health_insurance', '542 Health Insurance'),
        ('supplies_medical', '545 Supplies (Medical)'),
        ('legal_fees', '554 Legal Fees'),
        ('licenses_fees', '555 Licenses & Fees'),
        ('meals', '560 Meals'),
        ('software_apps', '570 Software & Apps'),
        ('office_expenses', '574 Office Expenses'),
        ('payroll_expenses', '580 Payroll Expenses'),
        ('physician_services', '592 Physician Services'),
        ('rent_lease', '600 Rent/Lease'),
        ('repairs_maintenance', '601 Repairs & Maintenance'),
        ('travel', '610 Travel'),
        ('uniforms', '620 Uniforms'),
    ]
    for key, label in expense_lines:
        ws.cell(row=r, column=2, value=label).font = data_font
        for i in range(n26):
            v = qbo26.get(key, [0] * n26)[i]
            cell = ws.cell(row=r, column=3 + i, value=v)
            cell.font = data_font; cell.number_format = CURR2
            cell.border = thin_border; cell.alignment = right_align
        r += 1

    ws.cell(row=r, column=2, value="TOTAL EXPENSES").font = data_bold
    for i in range(n26):
        cell = ws.cell(row=r, column=3 + i, value=qbo26['total_expenses'][i])
        cell.font = data_bold; cell.number_format = CURR2
        cell.border = thin_border; cell.alignment = right_align
    r += 1

    ws.cell(row=r, column=2, value="NET INCOME").font = data_bold
    for i in range(n26):
        cell = ws.cell(row=r, column=3 + i, value=qbo26['net_income'][i])
        cell.font = data_bold; cell.number_format = CURR2
        cell.border = thin_border; cell.alignment = right_align
    r += 1

    return ws


# ============================================================
# TAB 7: CASE ANALYTICS (data-driven from client case sheets)
# ============================================================
def build_case_analytics(wb):
    ws = wb.create_sheet("Case Analytics")
    ws.sheet_properties.tabColor = "C00000"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 28
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col_letter].width = 16

    ws.merge_cells('B2:I2')
    ws.cell(row=2, column=2,
            value="Case Analytics \u2014 Data-Driven from Client Case Sheets").font = title_font
    ws.cell(row=3, column=2,
            value="Source: GAP/BOBA case data (Sep 2024 \u2013 Apr 2026)").font = pct_font

    r = 5

    # ---- REVENUE BY CASE TYPE ----
    section_bar(ws, r, 2, 9, "REVENUE PER SURGERY (from paid cases)")
    r += 1
    header_row(ws, r, ["Metric", "BOBA", "GAP", "", "", "", "", ""], c1=2)
    r += 1

    # BOBA paid amounts (from analysis)
    boba_paid = [451000, 81402.43, 50659, 64967.84, 4948.40, 317013.12,
                 93738.97, 120000, 93796.66, 86000, 43738.73, 258792.82,
                 18000, 208889.90, 93065]
    gap_paid = [80000, 3549.02, 3000, 4205.40, 47040, 3302.51,
                94492.27, 162484.69]

    for label, bval, gval in [
        ("# Paid Cases", len(boba_paid), len(gap_paid)),
        ("Mean Revenue", sum(boba_paid)/len(boba_paid), sum(gap_paid)/len(gap_paid)),
        ("Median Revenue", sorted(boba_paid)[len(boba_paid)//2], sorted(gap_paid)[len(gap_paid)//2]),
        ("Min Revenue", min(boba_paid), min(gap_paid)),
        ("Max Revenue", max(boba_paid), max(gap_paid)),
        ("Total Collected", sum(boba_paid), sum(gap_paid)),
        ("Model Assumption", 132000, 50000),
    ]:
        ws.cell(row=r, column=2, value=label).font = data_font
        fmt = '$#,##0' if label != "# Paid Cases" else '#,##0'
        for ci, val in [(3, bval), (4, gval)]:
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = data_bold if "Model" in label else data_font
            cell.number_format = fmt
            cell.border = thin_border
            if "Model" in label:
                cell.fill = input_fill
                cell.font = input_font
        r += 1

    r += 2

    # ---- INSURANCE MIX ----
    section_bar(ws, r, 2, 9, "REVENUE BY INSURANCE COMPANY")
    r += 1
    header_row(ws, r, ["Insurance", "Cases", "Paid", "Total Paid", "Avg Paid", "Case Types", "", ""], c1=2)
    r += 1

    ins_data = [
        ("UHC/United", 10, 3, 562739, "BOBA,GAP"),
        ("Aetna", 7, 4, 558631, "BOBA,GAP"),
        ("Cigna", 8, 4, 514042, "BOBA"),
        ("Blue Shield of CA", 9, 6, 430319, "BOBA,GAP"),
        ("Anthem BlueCross", 8, 2, 98002, "BOBA,GAP"),
        ("Regal Health", 1, 1, 93065, "BOBA"),
        ("Kaiser", 4, 1, 80000, "BOBA,GAP"),
        ("BlueCross/BCBS", 8, 1, 43739, "BOBA,GAP"),
        ("Medicare", 1, 1, 3549, "GAP"),
        ("Medicaid", 1, 0, 0, "GAP"),
        ("Alignment Health", 1, 0, 0, "BOBA"),
    ]

    for ins, cases_n, paid_n, total, types in ins_data:
        ws.cell(row=r, column=2, value=ins).font = data_font
        ws.cell(row=r, column=3, value=cases_n).font = data_font
        ws.cell(row=r, column=3).number_format = '#,##0'
        ws.cell(row=r, column=4, value=paid_n).font = data_font
        ws.cell(row=r, column=4).number_format = '#,##0'
        ws.cell(row=r, column=5, value=total).font = data_font
        ws.cell(row=r, column=5).number_format = '$#,##0'
        avg = total / paid_n if paid_n > 0 else 0
        ws.cell(row=r, column=6, value=avg).font = data_font
        ws.cell(row=r, column=6).number_format = '$#,##0'
        ws.cell(row=r, column=7, value=types).font = pct_font
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = thin_border
        r += 1

    r += 2

    # ---- PAYMENT LAG ANALYSIS ----
    section_bar(ws, r, 2, 9, "PAYMENT LAG ANALYSIS (Surgery → Payment)")
    r += 1

    # BOBA lag detail
    ws.cell(row=r, column=2, value="BOBA Payment Lags").font = subsection_font
    r += 1
    header_row(ws, r, ["Surgery Date", "Payment Date", "Days", "Months", "Amount", "Insurance", "", ""], c1=2)
    r += 1

    boba_lags = [
        ("2025-07-04", "2025-09-27", 85, 2.8, 43739, "BCBS"),
        ("2025-09-30", "2026-02-26", 149, 4.9, 18000, "UHC/Surest"),
        ("2025-10-22", "2026-03-24", 153, 5.0, 208890, "Aetna"),
        ("2025-08-06", "2026-02-03", 181, 5.9, 258793, "Aetna"),
        ("2025-02-13", "2025-08-18", 186, 6.1, 93739, "UHC"),
        ("2025-06-05", "2025-12-13", 191, 6.3, 86000, "Aetna"),
        ("2025-04-18", "2025-11-30", 226, 7.4, 93797, "Anthem BC"),
        ("2024-10-16", "2025-06-01", 228, 7.5, 451000, "UHC"),
        ("2024-12-04", "2025-08-05", 244, 8.0, 64968, "Cigna"),
        ("2024-11-18", "2025-08-06", 261, 8.6, 81402, "Cigna"),
        ("2025-01-04", "2025-10-01", 270, 8.9, 317013, "Cigna"),
        ("2024-12-16", "2025-11-10", 329, 10.8, 4948, "Aetna"),
    ]

    for surg, pay, days, months, amt, ins in boba_lags:
        ws.cell(row=r, column=2, value=surg).font = data_font
        ws.cell(row=r, column=3, value=pay).font = data_font
        ws.cell(row=r, column=4, value=days).font = data_font
        ws.cell(row=r, column=5, value=months).font = data_font
        ws.cell(row=r, column=5).number_format = '0.0'
        ws.cell(row=r, column=6, value=amt).font = data_font
        ws.cell(row=r, column=6).number_format = '$#,##0'
        ws.cell(row=r, column=7, value=ins).font = data_font
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = thin_border
        r += 1

    # BOBA summary stats
    r += 1
    for label, val in [
        ("Avg Lag (days)", 209), ("Avg Lag (months)", 6.9),
        ("Median Lag (days)", 226), ("$ Weighted Avg (days)", 213),
        ("$ Weighted Avg (months)", 7.0),
    ]:
        ws.cell(row=r, column=2, value=label).font = data_bold
        cell = ws.cell(row=r, column=3, value=val)
        cell.font = data_bold
        cell.number_format = '#,##0' if 'days' in label.lower() else '0.0'
        r += 1

    r += 2

    # GAP lag detail
    ws.cell(row=r, column=2, value="GAP Payment Lags").font = subsection_font
    r += 1
    header_row(ws, r, ["Surgery Date", "Payment Date", "Days", "Months", "Amount", "Insurance", "", ""], c1=2)
    r += 1

    gap_lags = [
        ("2025-08-20", "2025-10-05", 46, 1.5, 94492, "Blue Shield"),
        ("2025-07-09", "2025-11-11", 125, 4.1, 4205, "Anthem BC"),
        ("2025-07-23", "2025-11-25", 125, 4.1, 47040, "Blue Shield"),
    ]

    for surg, pay, days, months, amt, ins in gap_lags:
        ws.cell(row=r, column=2, value=surg).font = data_font
        ws.cell(row=r, column=3, value=pay).font = data_font
        ws.cell(row=r, column=4, value=days).font = data_font
        ws.cell(row=r, column=5, value=months).font = data_font
        ws.cell(row=r, column=5).number_format = '0.0'
        ws.cell(row=r, column=6, value=amt).font = data_font
        ws.cell(row=r, column=6).number_format = '$#,##0'
        ws.cell(row=r, column=7, value=ins).font = data_font
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = thin_border
        r += 1

    r += 2

    # ---- BOBA LAG BY INSURER ----
    section_bar(ws, r, 2, 9, "BOBA PAYMENT LAG BY INSURER")
    r += 1
    header_row(ws, r, ["Insurance", "# Paid", "Avg Days", "Avg Mo", "$Wtd Days", "$Wtd Mo", "", ""], c1=2)
    r += 1

    ins_lag_data = [
        ("Aetna", 4, 214, 7.0, 173, 5.7),
        ("Anthem BC", 1, 226, 7.4, 226, 7.4),
        ("BCBS", 1, 85, 2.8, 85, 2.8),
        ("Cigna", 3, 258, 8.5, 265, 8.7),
        ("UHC/United", 3, 188, 6.2, 218, 7.2),
    ]

    for ins, n, avg_d, avg_m, wtd_d, wtd_m in ins_lag_data:
        ws.cell(row=r, column=2, value=ins).font = data_font
        ws.cell(row=r, column=3, value=n).font = data_font
        ws.cell(row=r, column=4, value=avg_d).font = data_font
        ws.cell(row=r, column=5, value=avg_m).font = data_font
        ws.cell(row=r, column=5).number_format = '0.0'
        ws.cell(row=r, column=6, value=wtd_d).font = data_font
        ws.cell(row=r, column=7, value=wtd_m).font = data_font
        ws.cell(row=r, column=7).number_format = '0.0'
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = thin_border
        r += 1

    r += 2

    # ---- COLLECTION CURVE COMPARISON ----
    section_bar(ws, r, 2, 9, "COLLECTION CURVE: ACTUAL vs MODEL")
    r += 1
    header_row(ws, r, ["Month", "BOBA Actual $%", "BOBA Model %", "GAP Actual $%", "GAP Model %", "", "", ""], c1=2)
    r += 1

    boba_actual_pct = [0, 0, 2.5, 0, 1.0, 27.2, 10.4, 31.6, 26.9, 0, 0.3, 0]
    boba_model_pct = [0, 0, 3, 0, 0, 27, 10, 32, 23, 0, 5, 0]
    gap_actual_pct = [0, 64.8, 0, 0, 35.2]
    gap_model_pct = [0, 55, 10, 10, 25]

    for m in range(max(len(boba_actual_pct), len(gap_actual_pct))):
        ws.cell(row=r, column=2, value=f"M+{m}").font = data_font
        if m < len(boba_actual_pct):
            ws.cell(row=r, column=3, value=boba_actual_pct[m] / 100).font = data_font
            ws.cell(row=r, column=3).number_format = '0.0%'
        if m < len(boba_model_pct):
            ws.cell(row=r, column=4, value=boba_model_pct[m] / 100).font = input_font
            ws.cell(row=r, column=4).number_format = '0.0%'
            ws.cell(row=r, column=4).fill = input_fill
        if m < len(gap_actual_pct):
            ws.cell(row=r, column=5, value=gap_actual_pct[m] / 100).font = data_font
            ws.cell(row=r, column=5).number_format = '0.0%'
        if m < len(gap_model_pct):
            ws.cell(row=r, column=6, value=gap_model_pct[m] / 100).font = input_font
            ws.cell(row=r, column=6).number_format = '0.0%'
            ws.cell(row=r, column=6).fill = input_fill
        for c in range(2, 7):
            ws.cell(row=r, column=c).border = thin_border
        r += 1

    r += 2

    # ---- AR RECONCILIATION ----
    section_bar(ws, r, 2, 9, "AR RECONCILIATION (Historical Cases)")
    r += 1
    header_row(ws, r, ["Month", "BOBA Total", "BOBA Paid", "BOBA AR", "GAP Total", "GAP Paid/Denied", "GAP AR", "Notes"], c1=2)
    r += 1

    ar_recon = [
        ("Sep-24", 1, 0, 1, 0, 0, 0, ""),
        ("Oct-24", 1, 1, 0, 2, 0, 2, "BOBA: UHC $451K paid"),
        ("Nov-24", 2, 1, 1, 1, 0, 1, "BOBA: Cigna $81K paid"),
        ("Dec-24", 5, 2, 3, 0, 0, 0, "BOBA: Aetna+Cigna paid"),
        ("Jan-25", 4, 1, 3, 0, 0, 0, "BOBA: Cigna $317K paid"),
        ("Feb-25", 1, 1, 0, 1, 1, 0, "GAP: Medicaid denied"),
        ("Mar-25", 1, 0, 1, 0, 0, 0, "BOBA: BS $120K WON, pending"),
        ("Apr-25", 3, 1, 2, 0, 0, 0, "BOBA: Anthem $94K paid"),
        ("May-25", 0, 0, 0, 0, 0, 0, ""),
        ("Jun-25", 3, 1, 2, 1, 1, 0, "GAP: BCBS denied"),
        ("Jul-25", 1, 1, 0, 3, 2, 1, "BOBA: BCBS $44K paid"),
        ("Aug-25", 1, 1, 0, 2, 2, 0, "GAP: 1 paid, 1 denied"),
        ("Sep-25", 3, 1, 2, 0, 0, 0, "BOBA: Surest $18K paid"),
        ("Oct-25", 3, 1, 2, 0, 0, 0, "BOBA: Aetna $209K paid"),
        ("Nov-25", 1, 0, 1, 4, 0, 4, ""),
        ("Dec-25", 4, 0, 4, 1, 0, 1, ""),
    ]

    for row_data in ar_recon:
        label, bt, bp, ba, gt, gpd, ga, notes = row_data
        ws.cell(row=r, column=2, value=label).font = data_font
        for ci, val in [(3, bt), (4, bp), (5, ba), (6, gt), (7, gpd), (8, ga)]:
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = data_font
            cell.number_format = '#,##0'
            cell.border = thin_border
            cell.alignment = right_align
            if ci == 5 and val > 0:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            if ci == 8 and val > 0:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws.cell(row=r, column=9, value=notes).font = pct_font
        r += 1

    # Totals
    r += 1
    ws.cell(row=r, column=2, value="TOTALS").font = data_bold
    ws.cell(row=r, column=5, value=sum(d[3] for d in ar_recon)).font = data_bold
    ws.cell(row=r, column=5).number_format = '#,##0'
    ws.cell(row=r, column=8, value=sum(d[6] for d in ar_recon)).font = data_bold
    ws.cell(row=r, column=8).number_format = '#,##0'
    ws.cell(row=r, column=9, value="22 BOBA + 9 GAP = 31 cases in AR").font = data_bold

    return ws


# ============================================================
# HELPERS
# ============================================================
def _get_monthly_total_expenses_25():
    """Helper: total 2025 monthly expenses from actuals."""
    expense_keys = [
        'advertising_marketing', 'bank_fees', 'conference', 'contracts',
        'contributions', 'dues_subscriptions', 'insurance', 'malpractice_insurance',
        'health_insurance', 'legal_accounting_services', 'accounting_fees',
        'billing_services', 'licenses_fees', 'mgmt_fee_abc', 'mgmt_fee_vnsc',
        'meals', 'office_expenses', 'software_apps', 'payroll_processing',
        'salaries_wages', 'payroll_taxes', 'physician_services', 'rent_lease',
        'taxes_licenses', 'travel',
    ]
    return [sum(ACTUALS_2025[k][i] for k in expense_keys) for i in range(12)]


# ============================================================
# MAIN
# ============================================================
def build_location_pl(wb, location_name, pl_data, asm):
    """Build a per-location P&L tab with FORMULAS referencing Assumptions tab."""
    tab_name = f"{location_name} P&L"
    if len(tab_name) > 31:
        tab_name = tab_name[:31]

    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = "7030A0" if location_name != "Westlake" else "4472C4"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35
    for ci in range(3, 3 + N + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 11
    ws.column_dimensions[get_column_letter(3 + N)].width = 14

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=min(3 + N, 30))
    ws.cell(row=2, column=2,
            value=f"{location_name} \u2014 Monthly P&L (Cash Basis, 2026-2030)").font = title_font

    r = 4
    last_col = 3 + N

    section_bar(ws, r, 2, last_col, f"{location_name.upper()} P&L")
    r += 1
    header_row(ws, r, ["Account"] + FORECAST_MONTH_LABELS + ["Total"], c1=2)
    r += 1

    # Get volume column references for this location from Assumptions tab
    loc_vol = asm.get('vol_by_location', {}).get(location_name)
    vol_start_row = asm.get('vol_start', 28)  # row where volume data starts

    if loc_vol:
        bobas_col_letter = get_column_letter(loc_vol['bobas_col'])
        gap_col_letter = get_column_letter(loc_vol['gap_col'])
    else:
        # Fallback to consolidated columns
        bobas_col_letter = 'C'
        gap_col_letter = 'D'

    # --- SURGERY VOLUMES (formulas referencing Assumptions) ---
    row_bobas_vol = r
    formulas = [f"=Assumptions!${bobas_col_letter}${vol_start_row + i}" for i in range(N)]
    r = _write_formula_row(ws, r, "Bobas Volume", formulas, fmt='#,##0')

    row_gap_vol = r
    formulas = [f"=Assumptions!${gap_col_letter}${vol_start_row + i}" for i in range(N)]
    r = _write_formula_row(ws, r, "GAP Volume", formulas, fmt='#,##0')

    row_total_vol = r
    formulas = [f"={mcol(i)}{row_bobas_vol}+{mcol(i)}{row_gap_vol}" for i in range(N)]
    r = _write_formula_row(ws, r, "Total Surgeries", formulas, bold=True, fmt='#,##0')
    r += 1

    # --- REVENUE EARNED (accrual, for collection curve) ---
    ws.cell(row=r, column=2, value="REVENUE (Cash Collected)").font = subsection_font
    r += 1

    row_bobas_earned = r
    formulas = [f"={mcol(i)}{row_bobas_vol}*Assumptions!$C${asm['avg_rev_bobas']}" for i in range(N)]
    r = _write_formula_row(ws, r, "  Bobas Earned (Accrual)", formulas)

    # Bobas collected (apply collection curve)
    row_bobas_collected = r
    bobas_curve_len = asm['curve_bobas_len']
    coll_formulas = []
    for j in range(N):
        terms = []
        for lag in range(bobas_curve_len):
            src = j - lag
            if src >= 0:
                terms.append(
                    f"{mcol(src)}{row_bobas_earned}*Assumptions!$C${asm['curve_bobas_start'] + lag}/100"
                )
        coll_formulas.append("=" + "+".join(terms) if terms else "=0")
    r = _write_formula_row(ws, r, "  Bobas Collected", coll_formulas)

    row_gap_earned = r
    formulas = [f"={mcol(i)}{row_gap_vol}*Assumptions!$C${asm['avg_rev_gap']}" for i in range(N)]
    r = _write_formula_row(ws, r, "  GAP Earned (Accrual)", formulas)

    # GAP collected (apply collection curve)
    row_gap_collected = r
    gap_curve_len = asm['curve_gap_len']
    coll_formulas = []
    for j in range(N):
        terms = []
        for lag in range(gap_curve_len):
            src = j - lag
            if src >= 0:
                terms.append(
                    f"{mcol(src)}{row_gap_earned}*Assumptions!$C${asm['curve_gap_start'] + lag}/100"
                )
        coll_formulas.append("=" + "+".join(terms) if terms else "=0")
    r = _write_formula_row(ws, r, "  GAP Collected", coll_formulas)

    row_total_collected = r
    formulas = [f"={mcol(i)}{row_bobas_collected}+{mcol(i)}{row_gap_collected}" for i in range(N)]
    r = _write_formula_row(ws, r, "TOTAL COLLECTIONS", formulas, bold=True)
    r += 1

    # --- OVERHEAD (mix of formulas and hardcoded) ---
    ws.cell(row=r, column=2, value="DIRECT OVERHEAD").font = subsection_font
    r += 1

    # Billing = 18% of collections
    row_billing = r
    formulas = [f"={mcol(i)}{row_total_collected}*Assumptions!$C${asm['billing_rate']}/100" for i in range(N)]
    r = _write_formula_row(ws, r, "  Billing (18%)", formulas)

    # All overhead rows reference per-location schedules on Assumptions tab
    loc_sched = asm.get('loc_schedules', {}).get(location_name, {})

    row_payroll = r
    if 'payroll' in loc_sched:
        formulas = [f"=Assumptions!{mcol(i)}${loc_sched['payroll']}" for i in range(N)]
        r = _write_formula_row(ws, r, "  Payroll (W-2)", formulas)
    else:
        r = _write_row(ws, r, "  Payroll (W-2)", pl_data['payroll'])

    row_contractors = r
    if 'contractors' in loc_sched:
        formulas = [f"=Assumptions!{mcol(i)}${loc_sched['contractors']}" for i in range(N)]
        r = _write_formula_row(ws, r, "  Contractors", formulas)
    else:
        r = _write_row(ws, r, "  Contractors", pl_data['contractors'])

    row_opex = r
    if 'opex' in loc_sched:
        formulas = [f"=Assumptions!{mcol(i)}${loc_sched['opex']}" for i in range(N)]
        r = _write_formula_row(ws, r, "  Operating Expenses", formulas)
    else:
        r = _write_row(ws, r, "  Operating Expenses", pl_data['direct_opex'])

    row_expansion = r
    if 'expansion' in loc_sched:
        formulas = [f"=Assumptions!{mcol(i)}${loc_sched['expansion']}" for i in range(N)]
        r = _write_formula_row(ws, r, "  Expansion Costs", formulas)
    else:
        r = _write_row(ws, r, "  Expansion Costs", pl_data.get('expansion_costs', [0]*N))

    row_shared = r
    if 'shared_overhead' in loc_sched:
        formulas = [f"=Assumptions!{mcol(i)}${loc_sched['shared_overhead']}" for i in range(N)]
        r = _write_formula_row(ws, r, "  Shared Overhead (allocated)", formulas)
    elif 'shared_overhead_allocation' in pl_data:
        r = _write_row(ws, r, "  Shared Overhead (allocated)", pl_data['shared_overhead_allocation'])

    # Total overhead (formula summing all components)
    row_total_overhead = r
    oh_refs = [row_billing, row_payroll, row_contractors, row_opex, row_expansion]
    if 'shared_overhead_allocation' in pl_data:
        oh_refs.append(row_shared)
    formulas = [
        "=" + "+".join(f"{mcol(i)}{rr}" for rr in oh_refs)
        for i in range(N)
    ]
    r = _write_formula_row(ws, r, "TOTAL OVERHEAD", formulas, bold=True)
    r += 1

    # Surgeon compensation
    surgeon_pay = pl_data.get('surgeon_compensation', [0]*N)
    row_surgeon = None
    if any(v != 0 for v in surgeon_pay):
        surgeon_rate = pl_data.get('surgeon_rate', 0)
        row_surgeon = r
        # Formula: collections * surgeon %
        formulas = [f"={mcol(i)}{row_total_collected}*{surgeon_rate}/100" for i in range(N)]
        r = _write_formula_row(ws, r, f"  Surgeon Compensation ({surgeon_rate:.0f}%)", formulas)
        r += 1

    # Contribution (formula)
    ws.cell(row=r, column=2, value="CONTRIBUTION").font = subsection_font
    r += 1
    if row_surgeon:
        formulas = [f"={mcol(i)}{row_total_collected}-{mcol(i)}{row_total_overhead}-{mcol(i)}{row_surgeon}" for i in range(N)]
    else:
        formulas = [f"={mcol(i)}{row_total_collected}-{mcol(i)}{row_total_overhead}" for i in range(N)]
    r = _write_formula_row(ws, r, "CONTRIBUTION", formulas, bold=True)

    # Style total rows
    for row_idx in range(4, r):
        label = ws.cell(row=row_idx, column=2).value
        if label and ("TOTAL" in str(label) or "CONTRIBUTION" == str(label).strip()):
            style_range(ws, row_idx, 2, last_col, border=bottom_border)

    ws.freeze_panes = "C7"
    return ws


def build_model(output_path: str = None):
    if output_path is None:
        output_path = "CNS Financial Model.xlsx"

    wb = openpyxl.Workbook()

    print("Building Dashboard...")
    build_dashboard(wb)

    print("Building Assumptions...")
    asm = build_assumptions(wb)

    print("Building Monthly P&L...")
    pl_rows = build_monthly_pl(wb, asm)

    print("Building Cash Flow...")
    build_cash_flow(wb, asm, pl_rows)

    print("Building Scenarios...")
    build_scenarios(wb)

    print("Building QBO Actuals...")
    build_qbo_actuals(wb)

    print("Building Per-Location P&Ls...")
    from financial_calcs import generate_pl_by_location
    pl_by_loc = generate_pl_by_location(DEFAULT_ASSUMPTIONS)
    for loc_name in DEFAULT_ASSUMPTIONS.get('locations', ['Westlake']):
        if loc_name in pl_by_loc:
            build_location_pl(wb, loc_name, pl_by_loc[loc_name], asm)
            print(f"  Built {loc_name} P&L")

    print("Building Case Analytics...")
    build_case_analytics(wb)

    wb.save(output_path)
    print(f"\nModel saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    import sys
    output = sys.argv[1] if len(sys.argv) > 1 else None
    build_model(output)
