"""
CNS Case Data Analysis
Extracts granular insights from client surgery case data for revenue forecasting.
"""
import openpyxl
from datetime import datetime, timedelta
import re

wb = openpyxl.load_workbook('/Users/chandlerclemons/financial-modeling/data/cns_case_data.xlsx', data_only=True)

def parse_amount(val):
    """Parse dollar amounts from various formats."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    # Handle "$266,354.12+ $50,659" -> sum
    if '+' in s:
        parts = s.split('+')
        total = 0
        for p in parts:
            p = p.strip().replace('$','').replace(',','').replace('k','000').replace('K','000')
            try:
                total += float(p)
            except:
                pass
        return total if total > 0 else None
    s = s.replace('$','').replace(',','').strip()
    if s.lower().endswith('k'):
        try: return float(s[:-1]) * 1000
        except: return None
    try: return float(s)
    except: return None

def parse_date(val):
    """Parse dates from various formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if s.lower() in ('pending', 'loss', 'tbd', 'no surgery', ''):
        return None
    # Handle "1/4/2025, 1/7/25" -> first date
    if ',' in s:
        s = s.split(',')[0].strip()
    # Handle "8/5/25 & 8/5/26" -> first date
    if '&' in s:
        s = s.split('&')[0].strip()
    for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%B %d, %Y']:
        try: return datetime.strptime(s, fmt)
        except: pass
    return None

def normalize_insurance(ins):
    """Normalize insurance company names."""
    if ins is None:
        return 'Unknown'
    ins = ins.strip().upper()
    if 'KAISER' in ins: return 'Kaiser'
    if 'AETNA' in ins: return 'Aetna'
    if 'CIGNA' in ins: return 'Cigna'
    if 'UHC' in ins or 'UNITED' in ins or 'UMR' in ins or 'SUREST' in ins: return 'UHC/United'
    if 'ANTHEM' in ins: return 'Anthem BlueCross'
    if 'BLUE CROSS' in ins or 'BLUECROSS' in ins or 'BCBS' in ins:
        if 'ANTHEM' in ins: return 'Anthem BlueCross'
        return 'BlueCross/BCBS'
    if 'BLUE SHIELD' in ins or 'BLUESHIELD' in ins or 'BS OF CA' in ins:
        return 'Blue Shield of CA'
    if 'MEDICARE' in ins: return 'Medicare'
    if 'GOLDCOAST' in ins or 'MEDICAID' in ins: return 'Medicaid'
    if 'REGAL' in ins: return 'Regal Health'
    if 'ALIGNMENT' in ins: return 'Alignment Health'
    return ins.title().strip()

# ============================================================
# EXTRACT ALL CASES
# ============================================================
cases = []

# BOBA 2024
ws = wb['BOBA 2024']
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8, values_only=True):
    if row[0] is None: continue
    cases.append({
        'type': 'BOBA', 'year': 2024,
        'surgery_date': parse_date(row[0]),
        'procedure': str(row[1])[:60] if row[1] else None,
        'insurance': normalize_insurance(row[2]),
        'amount': parse_amount(row[3]),
        'payment_date': parse_date(row[4]),
        'zip': row[5],
        'place': row[6],
    })

# BOBA 2025
ws = wb[' BOBA 2025']
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8, values_only=True):
    if row[0] is None: continue
    cases.append({
        'type': 'BOBA', 'year': 2025,
        'surgery_date': parse_date(row[0]),
        'procedure': str(row[1])[:60] if row[1] else None,
        'insurance': normalize_insurance(row[2]),
        'payment_date': parse_date(row[3]),
        'place': row[4],
        'amount': parse_amount(row[5]),
        'status': row[6],
        'zip': row[7],
    })

# BOBA 2026
ws = wb['BOBA 2026']
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8, values_only=True):
    if row[0] is None: continue
    cases.append({
        'type': 'BOBA', 'year': 2026,
        'surgery_date': parse_date(row[0]),
        'procedure': str(row[1])[:60] if row[1] else None,
        'insurance': normalize_insurance(row[2]),
        'place': row[3],
        'payment_date': parse_date(row[4]),
        'amount': parse_amount(row[5]),
        'status': row[6],
        'zip': row[7],
    })

# GAP 2024
ws = wb['GAP 2024']
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8, values_only=True):
    if row[0] is None: continue
    cases.append({
        'type': 'GAP', 'year': 2024,
        'surgery_date': parse_date(row[0]),
        'procedure': str(row[1])[:60] if row[1] else None,
        'insurance': normalize_insurance(row[2]),
        'payment_date': parse_date(row[3]),
        'place': row[4],
        'amount': parse_amount(row[5]),
        'status': row[6],
        'zip': row[7],
    })

# GAP 2025
ws = wb['GAP 2025']
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8, values_only=True):
    if row[0] is None: continue
    s = str(row[0]).strip().lower()
    if s in ('no surgery',): continue
    cases.append({
        'type': 'GAP', 'year': 2025,
        'surgery_date': parse_date(row[0]),
        'procedure': str(row[1])[:60] if row[1] else None,
        'insurance': normalize_insurance(row[2]),
        'payment_date': parse_date(row[3]),
        'place': row[4],
        'amount': parse_amount(row[5]),
        'status': row[6],
        'zip': row[7],
    })

# GAP 2026
ws = wb['GAP 2026']
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8, values_only=True):
    if row[0] is None: continue
    cases.append({
        'type': 'GAP', 'year': 2026,
        'surgery_date': parse_date(row[0]),
        'procedure': str(row[1])[:60] if row[1] else None,
        'insurance': normalize_insurance(row[2]),
        'payment_date': parse_date(row[3]),
        'place': row[4],
        'amount': parse_amount(row[5]),
        'zip': row[6],
    })

# Filter out cases without surgery dates
cases = [c for c in cases if c['surgery_date'] is not None]

# ============================================================
# ANALYSIS
# ============================================================

print("=" * 70)
print("CNS CASE DATA ANALYSIS")
print("=" * 70)

# --- Case counts ---
boba = [c for c in cases if c['type'] == 'BOBA']
gap = [c for c in cases if c['type'] == 'GAP']
print(f"\nTotal cases: {len(cases)} (BOBA: {len(boba)}, GAP: {len(gap)})")

# --- Cases with payment data ---
paid = [c for c in cases if c['amount'] is not None and c['amount'] > 0]
paid_with_date = [c for c in paid if c['payment_date'] is not None]
print(f"Cases with amount: {len(paid)}")
print(f"Cases with amount AND payment date: {len(paid_with_date)}")

# --- PAYMENT LAG ANALYSIS ---
print(f"\n{'='*70}")
print("PAYMENT LAG ANALYSIS (surgery date → payment date)")
print(f"{'='*70}")

lags_boba = []
lags_gap = []
for c in paid_with_date:
    days = (c['payment_date'] - c['surgery_date']).days
    months = round(days / 30.44, 1)
    lag_entry = {
        'days': days, 'months': months,
        'insurance': c['insurance'],
        'amount': c['amount'],
        'surgery': c['surgery_date'].strftime('%Y-%m-%d'),
        'paid': c['payment_date'].strftime('%Y-%m-%d'),
    }
    if c['type'] == 'BOBA':
        lags_boba.append(lag_entry)
    else:
        lags_gap.append(lag_entry)

print(f"\nBOBA cases with payment lag data: {len(lags_boba)}")
if lags_boba:
    for l in sorted(lags_boba, key=lambda x: x['days']):
        print(f"  {l['surgery']} → {l['paid']}  = {l['days']}d ({l['months']}mo)  "
              f"${l['amount']:>12,.2f}  {l['insurance']}")
    avg_lag = sum(l['days'] for l in lags_boba) / len(lags_boba)
    med_lag = sorted(l['days'] for l in lags_boba)[len(lags_boba)//2]
    wtd_lag = sum(l['days']*l['amount'] for l in lags_boba) / sum(l['amount'] for l in lags_boba)
    print(f"\n  Avg lag: {avg_lag:.0f} days ({avg_lag/30.44:.1f} months)")
    print(f"  Median lag: {med_lag} days ({med_lag/30.44:.1f} months)")
    print(f"  $ Weighted avg lag: {wtd_lag:.0f} days ({wtd_lag/30.44:.1f} months)")

print(f"\nGAP cases with payment lag data: {len(lags_gap)}")
if lags_gap:
    for l in sorted(lags_gap, key=lambda x: x['days']):
        print(f"  {l['surgery']} → {l['paid']}  = {l['days']}d ({l['months']}mo)  "
              f"${l['amount']:>12,.2f}  {l['insurance']}")
    avg_lag = sum(l['days'] for l in lags_gap) / len(lags_gap)
    med_lag = sorted(l['days'] for l in lags_gap)[len(lags_gap)//2]
    wtd_lag = sum(l['days']*l['amount'] for l in lags_gap) / sum(l['amount'] for l in lags_gap)
    print(f"\n  Avg lag: {avg_lag:.0f} days ({avg_lag/30.44:.1f} months)")
    print(f"  Median lag: {med_lag} days ({med_lag/30.44:.1f} months)")
    print(f"  $ Weighted avg lag: {wtd_lag:.0f} days ({wtd_lag/30.44:.1f} months)")

# --- REVENUE BY INSURANCE ---
print(f"\n{'='*70}")
print("REVENUE BY INSURANCE COMPANY")
print(f"{'='*70}")

from collections import defaultdict
rev_by_ins = defaultdict(lambda: {'count': 0, 'paid_count': 0, 'total_paid': 0, 'types': set()})
for c in cases:
    ins = c['insurance']
    rev_by_ins[ins]['count'] += 1
    rev_by_ins[ins]['types'].add(c['type'])
    if c['amount'] and c['amount'] > 0:
        rev_by_ins[ins]['paid_count'] += 1
        rev_by_ins[ins]['total_paid'] += c['amount']

print(f"\n{'Insurance':<25} {'Cases':>6} {'Paid':>6} {'Total Paid':>14} {'Avg Paid':>12} {'Types'}")
print("-" * 80)
for ins in sorted(rev_by_ins.keys(), key=lambda x: rev_by_ins[x]['total_paid'], reverse=True):
    d = rev_by_ins[ins]
    avg = d['total_paid'] / d['paid_count'] if d['paid_count'] > 0 else 0
    types = ','.join(sorted(d['types']))
    print(f"{ins:<25} {d['count']:>6} {d['paid_count']:>6} ${d['total_paid']:>12,.2f} ${avg:>10,.2f}  {types}")

# --- AR STATUS ---
print(f"\n{'='*70}")
print("AR STATUS (Cases with vs. without payment)")
print(f"{'='*70}")

for case_type in ['BOBA', 'GAP']:
    type_cases = [c for c in cases if c['type'] == case_type]
    paid_cases = [c for c in type_cases if c['amount'] and c['amount'] > 0 and c['payment_date']]
    unpaid = [c for c in type_cases if not (c['amount'] and c['amount'] > 0 and c['payment_date'])]
    
    print(f"\n{case_type}:")
    print(f"  Total cases: {len(type_cases)}")
    print(f"  PAID (with payment date): {len(paid_cases)}  (${sum(c['amount'] for c in paid_cases):,.2f})")
    print(f"  UNPAID / IN AR: {len(unpaid)}")
    
    # Some have amounts but no payment date (awarded but not yet received)
    awarded_no_date = [c for c in type_cases if c['amount'] and c['amount'] > 0 and not c['payment_date']]
    if awarded_no_date:
        print(f"    Awarded but unpaid: {len(awarded_no_date)} (${sum(c['amount'] for c in awarded_no_date):,.2f})")
    
    no_amount = [c for c in type_cases if not c['amount'] or c['amount'] == 0]
    if no_amount:
        print(f"    No amount yet: {len(no_amount)}")
    
    for c in unpaid:
        amt_str = f"${c['amount']:,.2f}" if c['amount'] else "TBD"
        status = c.get('status', '') or ''
        print(f"    {c['surgery_date'].strftime('%Y-%m-%d')} | {c['insurance']:<25} | {amt_str:>14} | {status}")

# --- REVENUE BY CASE TYPE ---
print(f"\n{'='*70}")
print("AVERAGE REVENUE BY CASE TYPE (from paid cases only)")
print(f"{'='*70}")

for case_type in ['BOBA', 'GAP']:
    paid_type = [c for c in cases if c['type'] == case_type and c['amount'] and c['amount'] > 0]
    if paid_type:
        amounts = [c['amount'] for c in paid_type]
        print(f"\n{case_type}: {len(paid_type)} cases with amounts")
        print(f"  Mean:   ${sum(amounts)/len(amounts):>12,.2f}")
        print(f"  Median: ${sorted(amounts)[len(amounts)//2]:>12,.2f}")
        print(f"  Min:    ${min(amounts):>12,.2f}")
        print(f"  Max:    ${max(amounts):>12,.2f}")
        print(f"  Total:  ${sum(amounts):>12,.2f}")

# --- MONTHLY VOLUME BY TYPE ---
print(f"\n{'='*70}")
print("MONTHLY CASE VOLUME (BOBA vs GAP)")
print(f"{'='*70}")

vol = defaultdict(lambda: {'BOBA': 0, 'GAP': 0})
for c in cases:
    key = c['surgery_date'].strftime('%Y-%m')
    vol[key][c['type']] += 1

print(f"\n{'Month':<10} {'BOBA':>6} {'GAP':>6} {'Total':>6}")
print("-" * 30)
for month in sorted(vol.keys()):
    b, g = vol[month]['BOBA'], vol[month]['GAP']
    print(f"{month:<10} {b:>6} {g:>6} {b+g:>6}")

