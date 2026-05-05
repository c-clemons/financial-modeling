"""
Flexible QBO Profit & Loss Excel parser for CNS.

Handles the standard QuickBooks "Profit and Loss" export shape:
    - Multi-row preamble (title, date range)
    - Header row with month names ("January 2025", "Jan 2026", etc.) + optional "Total"
    - Account rows whose first column starts with a 3-digit code ("400 Fee Income")
    - Subtotal rows ("Total for Income"), section headers ("Income", "Expenses"),
      and grand-total rows ("Net Income", "Gross Profit") which are not data rows.

Designed for an evolving chart of accounts: any row whose code/name does not
match `QBO_ACCOUNTS` is returned as `unmapped` for the user to map manually.
"""

from __future__ import annotations

import re
import sys
from datetime import datetime
from pathlib import Path
from typing import IO, Optional

import openpyxl

CNS_ROOT = Path(__file__).resolve().parent.parent
if str(CNS_ROOT) not in sys.path:
    sys.path.insert(0, str(CNS_ROOT))

from baseline_data import QBO_ACCOUNTS  # noqa: E402

MONTH_TOKEN_RE = re.compile(
    r"\b("
    r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
    r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|"
    r"nov(?:ember)?|dec(?:ember)?"
    r")\b",
    re.IGNORECASE,
)
YEAR_RE = re.compile(r"\b(20\d{2})\b")
ACCOUNT_CODE_RE = re.compile(r"^\s*(\d{3,4})\b")
SECTION_HEADERS = {
    "income", "expenses", "other income", "other expenses",
    "cost of goods sold", "cogs", "distribution account",
}
GRAND_TOTAL_LABELS = {
    "gross profit", "net operating income", "net income",
    "net other income", "total for income", "total for expenses",
    "total for other income", "total for other expenses",
}
MONTH_ABBR = {
    "jan": "Jan", "feb": "Feb", "mar": "Mar", "apr": "Apr",
    "may": "May", "jun": "Jun", "jul": "Jul", "aug": "Aug",
    "sep": "Sep", "sept": "Sep", "oct": "Oct", "nov": "Nov", "dec": "Dec",
}


# ---------------------------------------------------------------------------
# Mapping helpers
# ---------------------------------------------------------------------------

def _account_lookup_by_code() -> dict[str, dict]:
    return {code: meta for code, meta in QBO_ACCOUNTS.items()}


def _account_lookup_by_name() -> dict[str, dict]:
    return {meta["name"].strip().lower(): {**meta, "code": code}
            for code, meta in QBO_ACCOUNTS.items()}


def normalize_key(label: str) -> str:
    """Turn an arbitrary account label into a usable snake_case key."""
    s = label.strip().lower()
    # strip leading 3-4 digit code
    s = re.sub(r"^\d{3,4}\s+", "", s)
    # remove "@ ..." compensation tags
    s = re.split(r"\s+@\s+", s)[0]
    # strip parentheticals
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    return s or "unmapped"


def map_account(label: str, extras: Optional[dict] = None) -> dict:
    """Resolve a row label to a canonical key.

    Returns a dict with keys: code, key, name, mapped (bool), source.
    """
    if not label:
        return {"code": None, "key": None, "name": "", "mapped": False, "source": "empty"}

    raw = str(label).strip()
    extras = extras or {}

    # 1) user-provided mapping (full label)
    if raw in extras:
        target = extras[raw]
        return {"code": _extract_code(raw), "key": target, "name": raw,
                "mapped": True, "source": "extras"}

    # 2) lookup by 3-4 digit code prefix
    code = _extract_code(raw)
    by_code = _account_lookup_by_code()
    if code and code in by_code:
        meta = by_code[code]
        return {"code": code, "key": meta["key"], "name": meta["name"],
                "mapped": True, "source": "code"}

    # 3) fuzzy match by name (strip code prefix)
    name_only = re.sub(r"^\s*\d{3,4}\s+", "", raw).lower().strip()
    by_name = _account_lookup_by_name()
    if name_only in by_name:
        meta = by_name[name_only]
        return {"code": meta["code"], "key": meta["key"], "name": meta["name"],
                "mapped": True, "source": "name"}
    # 4) loose contains-match on first significant token
    for name_lc, meta in by_name.items():
        if name_lc and name_lc in name_only:
            return {"code": meta["code"], "key": meta["key"], "name": meta["name"],
                    "mapped": True, "source": "fuzzy"}

    # 5) unmapped — suggest a key derived from the label
    return {"code": code, "key": normalize_key(raw), "name": raw,
            "mapped": False, "source": "suggested"}


def _extract_code(label: str) -> Optional[str]:
    m = ACCOUNT_CODE_RE.match(str(label))
    return m.group(1) if m else None


# ---------------------------------------------------------------------------
# Header parsing
# ---------------------------------------------------------------------------

def _parse_month_header(cell_value) -> Optional[str]:
    """Convert a header cell like 'January 2025' or 'Jan 2026' to 'Jan-25'."""
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.strftime("%b-%y")
    s = str(cell_value).strip()
    if not s or s.lower() == "total":
        return None
    m = MONTH_TOKEN_RE.search(s)
    y = YEAR_RE.search(s)
    if not m or not y:
        return None
    abbr = MONTH_ABBR[m.group(1).lower()[:4] if m.group(1).lower().startswith("sept") else m.group(1).lower()[:3]]
    yr = y.group(1)[2:]
    return f"{abbr}-{yr}"


def _detect_header(ws) -> tuple[int, list[tuple[int, str]]]:
    """Locate the row that contains month headers. Returns (row_idx, [(col_idx, 'Mon-YY'), ...])."""
    max_scan = min(ws.max_row, 12)
    for r in range(1, max_scan + 1):
        candidates = []
        for c in range(2, ws.max_column + 1):
            label = _parse_month_header(ws.cell(row=r, column=c).value)
            if label:
                candidates.append((c, label))
        if len(candidates) >= 1:
            return r, candidates
    raise ValueError("Could not locate a month header row in the workbook.")


# ---------------------------------------------------------------------------
# Main entry
# ---------------------------------------------------------------------------

def parse_pl_workbook(file_obj: IO | str | Path,
                       extras: Optional[dict] = None) -> dict:
    """Parse a QBO P&L .xlsx file (path or file-like).

    Returns:
        {
            'months': ['Jan-26', 'Feb-26'],
            'year': 2026,                          # primary year of data
            'data': {key: [vals_per_month]},       # mapped accounts only
            'unmapped': [{'label': str, 'code': str|None,
                          'suggested_key': str, 'values': [...]}],
            'totals': {'total_income': [...], 'total_expenses': [...],
                       'net_income': [...]},
            'meta': {'sheet': str, 'header_row': int,
                     'parsed_rows': int, 'skipped_rows': int}
        }
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.worksheets[0]

    header_row, month_cols = _detect_header(ws)
    months = [m for _, m in month_cols]
    cols = [c for c, _ in month_cols]
    n_months = len(cols)

    if not months:
        raise ValueError("No monthly columns detected.")

    # Determine primary year (most common 2-digit suffix)
    year_counts: dict[str, int] = {}
    for m in months:
        yr = "20" + m.split("-")[-1]
        year_counts[yr] = year_counts.get(yr, 0) + 1
    primary_year = int(max(year_counts.items(), key=lambda kv: kv[1])[0])

    data: dict[str, list[float]] = {}
    unmapped: list[dict] = []
    totals: dict[str, list[float]] = {}
    parsed = 0
    skipped = 0

    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if not label or not str(label).strip():
            skipped += 1
            continue
        label_str = str(label).strip()
        label_lc = label_str.lower()

        # Skip section headers and grand totals at the row level
        if label_lc in SECTION_HEADERS:
            skipped += 1
            continue
        if any(label_lc == g for g in GRAND_TOTAL_LABELS):
            # capture the grand totals
            vals = _row_values(ws, r, cols)
            if "total for income" in label_lc:
                totals["total_income"] = vals
            elif "total for expenses" in label_lc:
                totals["total_expenses"] = vals
            elif "net operating income" in label_lc:
                totals["net_operating_income"] = vals
            elif "net income" in label_lc and "operating" not in label_lc:
                totals["net_income"] = vals
            elif "gross profit" in label_lc:
                totals["gross_profit"] = vals
            continue

        # Skip subtotal aggregator rows ("Total for 540 Insurance")
        if label_lc.startswith("total for "):
            skipped += 1
            continue

        vals = _row_values(ws, r, cols)
        # Skip rows with no numeric data
        if not any(isinstance(v, (int, float)) and v != 0 for v in vals):
            # but keep them if they would map (preserving zero rows is ok)
            if all(v in (None, "", 0) for v in vals):
                skipped += 1
                continue

        mapping = map_account(label_str, extras)
        # Coerce values
        coerced = [_to_float(v) for v in vals]

        if mapping["mapped"]:
            key = mapping["key"]
            if key in data:
                # account appears twice (rare); sum them
                data[key] = [a + b for a, b in zip(data[key], coerced)]
            else:
                data[key] = coerced
            parsed += 1
        else:
            unmapped.append({
                "label": label_str,
                "code": mapping["code"],
                "suggested_key": mapping["key"],
                "values": coerced,
            })

    # If totals weren't found explicitly, derive net_income from data if possible
    if "total_income" not in totals and "fee_income" in data:
        totals["total_income"] = list(data["fee_income"])
    if "total_expenses" not in totals:
        totals["total_expenses"] = [0.0] * n_months
    if "net_income" not in totals:
        ti = totals.get("total_income", [0.0] * n_months)
        te = totals.get("total_expenses", [0.0] * n_months)
        totals["net_income"] = [ti[i] - te[i] for i in range(n_months)]

    return {
        "months": months,
        "year": primary_year,
        "data": data,
        "unmapped": unmapped,
        "totals": totals,
        "meta": {
            "sheet": ws.title,
            "header_row": header_row,
            "parsed_rows": parsed,
            "skipped_rows": skipped,
        },
    }


def _row_values(ws, row_idx: int, cols: list[int]) -> list[float]:
    out = []
    for c in cols:
        v = ws.cell(row=row_idx, column=c).value
        out.append(v)
    return out


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace(",", "").replace("$", "").strip()
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        return float(s)
    except (TypeError, ValueError):
        return 0.0


# ---------------------------------------------------------------------------
# Convenience: turn parsed payload into a DataStore-ready upload dict
# ---------------------------------------------------------------------------

PAYROLL_KEYS = ("payroll_processing", "salaries_wages", "payroll_taxes")


def to_upload_payload(parsed: dict, source_filename: str) -> dict:
    """Shape parsed output for DataStore.set_uploaded_actuals()."""
    n = len(parsed["months"])
    data = dict(parsed["data"])

    # Synthetic combined payroll line (existing monthly_pl overlay reads this)
    if any(k in data for k in PAYROLL_KEYS):
        combined = [0.0] * n
        for k in PAYROLL_KEYS:
            for i, v in enumerate(data.get(k, [0.0] * n)):
                combined[i] += v
        data["payroll_expenses"] = combined

    # Ensure mandatory totals exist as lists
    totals = dict(parsed["totals"])
    for k, default_source in [
        ("total_income", "fee_income"),
        ("total_expenses", None),
        ("net_income", None),
    ]:
        if k not in totals:
            totals[k] = list(data.get(default_source, [0.0] * n)) if default_source else [0.0] * n

    return {
        "months": parsed["months"],
        "data": data,
        "totals": totals,
        "source_filename": source_filename,
        "uploaded_at": datetime.now().isoformat(timespec="seconds"),
        "meta": parsed.get("meta", {}),
    }
