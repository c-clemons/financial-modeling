"""
Mighty Pilates Financial Model Builder - FORMULA-DRIVEN

Mirrors the Streamlit dashboard at github.com/c-clemons/mighty-pilates.
All calculated cells use Excel formulas referencing the Assumptions tab.
Yellow cells are the only hard-coded inputs.

Coverage: 2026 (Jan-Apr actuals + May-Dec forecast), 2027, 2028 forecast.
"""

import json
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ============================================================
# PATHS
# ============================================================
DASH_DATA = Path("/Users/chandlerclemons/mighty-pilates/dashboard/data")
OUTPUT_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Mighty Pilates/"
    "Mighty Pilates Financial Model.xlsx"
)

# ============================================================
# DATA LOADERS
# ============================================================

def load_dashboard_data():
    """Load all source JSONs from the Streamlit dashboard."""
    with open(DASH_DATA / "committed_actuals.json") as f:
        committed = json.load(f)
    with open(DASH_DATA / "baseline.json") as f:
        baseline = json.load(f)
    try:
        with open(DASH_DATA / "user_overrides.json") as f:
            overrides = json.load(f)
    except FileNotFoundError:
        overrides = {}

    # Merge baseline + overrides (overrides win)
    def deep_merge(a, b):
        if isinstance(a, dict) and isinstance(b, dict):
            out = dict(a)
            for k, v in b.items():
                out[k] = deep_merge(a.get(k), v) if k in a else v
            return out
        return b if b is not None else a

    merged = deep_merge(baseline, overrides)

    return {
        "committed": committed,
        "baseline": baseline,
        "overrides": overrides,
        "merged": merged,
    }


# ============================================================
# CONSTANTS
# ============================================================

# 12 active studios (matches Streamlit ACTIVE_STUDIOS)
STUDIOS = [
    ("BK", "Berkeley"),
    ("CC", "Culver City"),
    ("DN", "Danville"),
    ("LF", "Lafayette"),
    ("MR", "Marin"),
    ("OP", "Ocean Park"),
    ("PH", "Presidio Heights"),
    ("RH", "Russian Hill"),
    ("SB", "Santa Barbara"),
    ("SM", "Santa Monica"),
    ("WP", "West Portal"),
    ("WW", "Westwood"),
]

# Development / not-yet-open studios (used in baseline data but excluded from active P&L)
DEV_STUDIOS = [
    ("CDM", "Corona Del Mar"),
    ("PS", "Pasadena"),
]

# OpEx categories (mirrors Streamlit category buckets)
OPEX_CATEGORIES = [
    ("property", "Property Costs", "rent_lease"),
    ("staff", "Staff Costs", "staff"),
    ("utilities", "Utilities", "operating"),
    ("marketing", "Marketing & Promotion", "operating"),
    ("admin", "Administrative & G&A", "operating"),
    ("professional", "Professional Fees", "operating"),
    ("travel", "Travel & Meals", "operating"),
    ("cogs", "Merchant Fees & COGS", "operating"),
    ("startup", "Studio Start Up Costs", "operating"),
    ("taxes", "Taxes", "operating"),
]

# Months: Jan 2026 through Dec 2028 (36 months)
MONTH_LABELS = []
MONTH_KEYS = []
MONTH_DATES = []
for year in (2026, 2027, 2028):
    for mo in range(1, 13):
        MONTH_LABELS.append(f"{['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][mo-1]} {year}")
        MONTH_KEYS.append(f"{year}-{mo:02d}")
        MONTH_DATES.append(datetime(year, mo, 1))

LAST_ACTUALS_MONTH = "2026-04"  # Apr 2026
LAST_ACTUALS_IDX = MONTH_KEYS.index(LAST_ACTUALS_MONTH)  # 3
FIRST_FORECAST_IDX = LAST_ACTUALS_IDX + 1  # 4 (May 2026)

# ============================================================
# STYLES
# ============================================================
NAVY = "1B2A4A"
ACCENT_BLUE = "4472C4"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"
GREEN_FILL_C = "E2EFDA"
RED_FILL_C = "FCE4EC"
YELLOW_FILL_C = "FFF9E6"
INPUT_FILL_C = "FFFFCC"
ACTUAL_FILL_C = "E8F4E8"  # very light green for actuals columns

title_font = Font(name="Calibri", size=16, bold=True, color=NAVY)
section_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
subsection_font = Font(name="Calibri", size=11, bold=True, color=NAVY)
header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
data_font = Font(name="Calibri", size=10)
data_bold = Font(name="Calibri", size=10, bold=True)
input_font = Font(name="Calibri", size=10, color="0000CC", bold=False)
metric_value_font = Font(name="Calibri", size=14, bold=True, color=NAVY)
metric_label_font = Font(name="Calibri", size=9, color="666666")

section_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
header_fill = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
input_fill = PatternFill(start_color=INPUT_FILL_C, end_color=INPUT_FILL_C, fill_type="solid")
actual_fill = PatternFill(start_color=ACTUAL_FILL_C, end_color=ACTUAL_FILL_C, fill_type="solid")
green_fill = PatternFill(start_color=GREEN_FILL_C, end_color=GREEN_FILL_C, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MED_GRAY), right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY), bottom=Side(style="thin", color=MED_GRAY),
)
double_bottom = Border(bottom=Side(style="double", color=NAVY))

CURR = '#,##0;[Red](#,##0);"-"'
CURR2 = '#,##0.00'
PCT = '0.0%'
PCT2 = '0.00%'
NUM = '#,##0'
center_align = Alignment(horizontal="center", vertical="center")
right_align = Alignment(horizontal="right", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")


# ============================================================
# HELPERS
# ============================================================

def style_range(ws, row, c1, c2, font=None, fill=None, border=None, alignment=None, number_format=None):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if border: cell.border = border
        if alignment: cell.alignment = alignment
        if number_format: cell.number_format = number_format


def section_bar(ws, row, c1, c2, label):
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = section_fill
    cell = ws.cell(row=row, column=c1, value=label)
    cell.font = section_font
    cell.alignment = left_align


def header_row(ws, row, c1, values, fill=None):
    use_fill = fill or header_fill
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=c1 + i, value=v)
        cell.font = header_font
        cell.fill = use_fill
        cell.alignment = center_align


def input_cell(ws, row, col, value, fmt=CURR):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = input_font
    cell.fill = input_fill
    cell.border = thin_border
    cell.number_format = fmt
    return cell


def formula_cell(ws, row, col, formula, fmt=CURR, bold=False, fill=None):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = data_bold if bold else data_font
    if fill:
        cell.fill = fill
    cell.number_format = fmt
    cell.border = thin_border
    return cell


def actual_value_cell(ws, row, col, value, fmt=CURR, bold=False):
    """Hard-coded actuals (closed-month historical data)."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = data_bold if bold else data_font
    cell.fill = actual_fill
    cell.border = thin_border
    cell.number_format = fmt
    return cell


def label_cell(ws, row, col, text, bold=False, indent=0):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = data_bold if bold else data_font
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    return cell


# ============================================================
# TAB BUILDERS (stubs - to be implemented)
# ============================================================

def build_cover(wb, data):
    """Cover / instructions tab."""
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 90

    ws["B2"] = "Mighty Pilates"
    ws["B2"].font = Font(name="Calibri", size=24, bold=True, color=NAVY)
    ws["B3"] = "Financial Model"
    ws["B3"].font = Font(name="Calibri", size=18, color=NAVY)
    ws["B5"] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    ws["B5"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    ws["B7"] = "Tab Guide"
    ws["B7"].font = subsection_font

    tabs = [
        ("Summary & Controls", "Top-level KPIs, studio sales & EBITDA tables, charts"),
        ("Assumptions", "All model inputs: OpEx, sales forecast, rev rec curves, loan terms"),
        ("Cash Flow Forecast", "Monthly cash flow from operations + investing + financing"),
        ("P&L Summary", "Consolidated monthly P&L (Revenue → EBITDA)"),
        ("P&L Detail", "Account-level P&L matching accountant chart of accounts"),
        ("All Studios Summary", "Side-by-side studio comparison"),
        ("[Studio] FCST (×12)", "Per-studio P&L forecast (BK, CC, DN, LF, MR, OP, PH, RH, SB, SM, WP, WW)"),
        ("Cash, Debt & Equity", "Balance sheet, loan amortization schedules, equity tracking"),
        ("Sales Forecast", "Studio × month editable sales grid"),
        ("CapEx", "Capital expenditure project schedule"),
        ("QBO Actuals", "Historical accountant-booked P&L, BS, SCF (read-only)"),
    ]
    r = 9
    for tab, desc in tabs:
        ws.cell(row=r, column=2, value=tab).font = data_bold
        ws.cell(row=r, column=3, value=desc).font = data_font
        r += 1

    ws.column_dimensions["C"].width = 70

    ws["B" + str(r + 2)] = "Conventions"
    ws["B" + str(r + 2)].font = subsection_font
    ws.cell(row=r + 4, column=2, value="Yellow cells = inputs (edit these)").fill = input_fill
    ws.cell(row=r + 5, column=2, value="Light green cells = actuals (closed months, do not edit)").fill = actual_fill


def build_assumptions(wb, data):
    """
    Assumptions tab — global model parameters, rev rec curves, loan terms.

    Sections:
      1. Model Control (last actuals month, forecast horizon)
      2. Revenue Recognition Curves (earned/breakage by month lag)
      3. Refund / Discount / Merchant / COGS rates
      4. Annual Escalation Rates (rent vs other OpEx, sales growth)
      5. Loan Terms
      6. Studio List (active studios)
    """
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 18

    # Title
    ws["B2"] = "Assumptions"
    ws["B2"].font = title_font
    ws["B3"] = "Yellow cells are editable inputs. All other tabs reference these values."
    ws["B3"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    row = 5

    # Initialize refs dict for cross-sheet references
    refs = {}

    # ----- Section 1: Model Control -----
    section_bar(ws, row, 2, 9, "1. Model Control")
    row += 1
    label_cell(ws, row, 2, "Last Actuals Month")
    input_cell(ws, row, 3, "Apr 2026", fmt="@")
    refs["last_actuals_row"] = row
    row += 1
    label_cell(ws, row, 2, "Forecast Horizon (months)")
    input_cell(ws, row, 3, 32, fmt=NUM)
    refs["horizon_row"] = row
    row += 1
    label_cell(ws, row, 2, "Forecast End Month")
    input_cell(ws, row, 3, "Dec 2028", fmt="@")
    refs["end_row"] = row
    row += 2

    # ----- Section 2: Revenue Recognition Curves -----
    section_bar(ws, row, 2, 9, "2. Revenue Recognition Curves")
    row += 1
    label_cell(ws, row, 2, "Lag Month (0 = sale month)", bold=True)
    for i in range(7):
        cell = ws.cell(row=row, column=3 + i, value=i)
        cell.font = data_bold
        cell.fill = light_fill
        cell.alignment = center_align
    label_cell(ws, row, 10, "Sum", bold=True)
    ws.cell(row=row, column=10).fill = light_fill
    ws.cell(row=row, column=10).alignment = center_align
    row += 1

    # Earned curve
    label_cell(ws, row, 2, "Earned %")
    curves = data["committed"]["rev_rec_curves"]
    earned = curves["earned"]
    refs["earned_row"] = row
    for i in range(7):
        input_cell(ws, row, 3 + i, float(earned.get(str(i), 0)) / 100.0, fmt=PCT)
    formula_cell(ws, row, 10, f"=SUM(C{row}:I{row})", fmt=PCT, bold=True, fill=light_fill)
    row += 1

    # Breakage curve
    label_cell(ws, row, 2, "Breakage %")
    breakage = curves["breakage"]
    refs["breakage_row"] = row
    for i in range(7):
        input_cell(ws, row, 3 + i, float(breakage.get(str(i), 0)) / 100.0, fmt=PCT)
    formula_cell(ws, row, 10, f"=SUM(C{row}:I{row})", fmt=PCT, bold=True, fill=light_fill)
    row += 1

    # Combined total
    label_cell(ws, row, 2, "Total (Earned + Breakage)", bold=True)
    for i in range(7):
        col = get_column_letter(3 + i)
        formula_cell(ws, row, 3 + i,
                     f"={col}{refs['earned_row']}+{col}{refs['breakage_row']}",
                     fmt=PCT, bold=True, fill=light_fill)
    formula_cell(ws, row, 10, f"=SUM(C{row}:I{row})", fmt=PCT, bold=True, fill=light_fill)
    row += 2

    # ----- Section 3: Rates -----
    section_bar(ws, row, 2, 9, "3. Rates (% of Gross Revenue)")
    row += 1
    rate_specs = [
        ("refund_row", "Refund Rate", float(curves["refund_pct"]) / 100.0,
         "Refunds as % of gross revenue. Negative reduces net revenue."),
        ("discount_row", "Discount Rate", float(curves["discount_pct"]) / 100.0,
         "Discounts as % of gross revenue. Negative reduces net revenue."),
        ("merchant_row", "Merchant Fee %",
         float(data["committed"]["forecast_ratios"]["merchant_fee_pct"]) / 100.0,
         "Payment processor fees as % of revenue."),
        ("cogs_row", "COGS %",
         float(data["committed"]["forecast_ratios"]["cogs_pct"]) / 100.0,
         "Cost of goods sold as % of revenue (retail)."),
    ]
    for key, label, val, note in rate_specs:
        refs[key] = row
        label_cell(ws, row, 2, label)
        input_cell(ws, row, 3, val, fmt=PCT2)
        cell = ws.cell(row=row, column=4, value=note)
        cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
        cell.alignment = left_align
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
        row += 1
    row += 1

    # ----- Section 4: Annual Escalation Rates -----
    section_bar(ws, row, 2, 9, "4. Annual Escalation Rates")
    row += 1
    esc_specs = [
        ("rent_esc_row", "Rent / Property Escalation", 0.03,
         "3% annual rent step-up applied to forecast months"),
        ("other_esc_row", "Other OpEx Escalation", 0.04,
         "4% annual escalation for all non-rent OpEx categories"),
        ("sales_growth_row", "Sales YoY Growth", 0.05,
         "5% YoY assumed for 2028 (applied to 2027 seasonal pattern)"),
    ]
    for key, label, val, note in esc_specs:
        refs[key] = row
        label_cell(ws, row, 2, label)
        input_cell(ws, row, 3, val, fmt=PCT)
        cell = ws.cell(row=row, column=4, value=note)
        cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
        cell.alignment = left_align
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
        row += 1
    row += 1

    # ----- Section 5: Loan Terms -----
    section_bar(ws, row, 2, 9, "5. Loan Terms (as of Apr 2026)")
    row += 1
    header_row(ws, row, 2, ["Loan", "Original Amt", "Current Balance", "Rate (Annual)",
                             "Avg Monthly Pmt", "Start Date", "Type"])
    row += 1
    loans = data["baseline"].get("loans", [])
    refs["loan_start_row"] = row
    refs["loans"] = []
    for loan in loans:
        loan_row = row
        label_cell(ws, row, 2, loan.get("name", loan.get("id", "?")), bold=True)
        input_cell(ws, row, 3, float(loan.get("original_amount", 0)), fmt=CURR)
        input_cell(ws, row, 4, float(loan.get("current_balance", 0)), fmt=CURR)
        input_cell(ws, row, 5, float(loan.get("rate", 0)), fmt=PCT2)
        input_cell(ws, row, 6, float(loan.get("avg_monthly_payment", 0)), fmt=CURR)
        input_cell(ws, row, 7, str(loan.get("start_date", "")), fmt="@")
        loan_type = "Interest-Only" if loan.get("avg_monthly_payment", 0) == 0 else "Amortizing"
        label_cell(ws, row, 8, loan_type)
        refs["loans"].append({
            "name": loan.get("name", "?"),
            "row": loan_row,
            "is_interest_only": loan.get("avg_monthly_payment", 0) == 0,
        })
        row += 1
    refs["loan_end_row"] = row - 1
    # Totals row
    label_cell(ws, row, 2, "TOTAL DEBT", bold=True)
    formula_cell(ws, row, 4,
                 f"=SUM(D{refs['loan_start_row']}:D{refs['loan_end_row']})",
                 fmt=CURR, bold=True, fill=light_fill)
    formula_cell(ws, row, 6,
                 f"=SUM(F{refs['loan_start_row']}:F{refs['loan_end_row']})",
                 fmt=CURR, bold=True, fill=light_fill)
    refs["loan_total_row"] = row
    row += 2

    # ----- Section 6: Active Studios -----
    section_bar(ws, row, 2, 9, "6. Active Studios")
    row += 1
    header_row(ws, row, 2, ["Code", "Name"])
    row += 1
    for code, name in STUDIOS:
        label_cell(ws, row, 2, code, bold=True)
        label_cell(ws, row, 3, name)
        row += 1

    # Stash refs on workbook for downstream tabs
    if not hasattr(wb, "_mighty_refs"):
        wb._mighty_refs = {}
    wb._mighty_refs["assumptions"] = refs

    return refs


def build_sales_forecast(wb, data):
    """
    Sales Forecast tab: studio × month grid.

    Layout:
      - Col B = Studio code, Col C = Studio name
      - Cols D... = 36 months (Jan 2026 - Dec 2028)
      - First 4 months (Jan-Apr 2026) are actuals (light green, locked)
      - May 2026 onwards are yellow editable forecast cells
      - Bottom row = TOTAL per month (SUM formula)
      - Right side = annual totals + YoY %
    """
    ws = wb.create_sheet("Sales Forecast")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D4"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 22
    for i in range(36):
        ws.column_dimensions[get_column_letter(4 + i)].width = 11

    # Title
    ws["B2"] = "Sales Forecast (Gross)"
    ws["B2"].font = title_font
    ws["B3"] = "Jan-Apr 2026 actuals (locked). May 2026 onwards forecast (editable)."
    ws["B3"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    HEADER_ROW = 5
    # Section bar
    section_bar(ws, HEADER_ROW - 1, 2, 39, "Per-Studio Monthly Sales")

    # Month headers
    label_cell(ws, HEADER_ROW, 2, "Code", bold=True)
    label_cell(ws, HEADER_ROW, 3, "Studio", bold=True)
    ws.cell(row=HEADER_ROW, column=2).fill = header_fill
    ws.cell(row=HEADER_ROW, column=2).font = header_font
    ws.cell(row=HEADER_ROW, column=3).fill = header_fill
    ws.cell(row=HEADER_ROW, column=3).font = header_font
    for i, label in enumerate(MONTH_LABELS):
        cell = ws.cell(row=HEADER_ROW, column=4 + i, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Data rows per studio
    csf = data["committed"]["client_sales_forecast"]
    refs = {"studios": {}, "header_row": HEADER_ROW, "first_data_col": 4}

    row = HEADER_ROW + 1
    for code, name in STUDIOS:
        label_cell(ws, row, 2, code, bold=True)
        label_cell(ws, row, 3, name)
        studio_data = csf.get(code, {})
        for i, mk in enumerate(MONTH_KEYS):
            val = float(studio_data.get(mk, 0))
            if i <= LAST_ACTUALS_IDX:
                # Actuals (Jan-Apr 2026)
                actual_value_cell(ws, row, 4 + i, val, fmt=CURR)
            else:
                # Forecast (editable)
                input_cell(ws, row, 4 + i, val, fmt=CURR)
        refs["studios"][code] = row
        row += 1

    # TOTAL row
    TOTAL_ROW = row
    refs["total_row"] = TOTAL_ROW
    label_cell(ws, row, 2, "TOTAL", bold=True)
    label_cell(ws, row, 3, "All Studios", bold=True)
    for c in range(2, 4):
        ws.cell(row=row, column=c).fill = light_fill
    first_studio_row = HEADER_ROW + 1
    for i in range(36):
        col = get_column_letter(4 + i)
        formula_cell(ws, row, 4 + i,
                     f"=SUM({col}{first_studio_row}:{col}{row - 1})",
                     fmt=CURR, bold=True, fill=light_fill)
    row += 2

    # Annual totals section
    section_bar(ws, row, 2, 8, "Annual Summary")
    row += 1
    header_row(ws, row, 2, ["Code", "Studio", "2026", "2027", "2028", "2027 YoY", "2028 YoY"])
    row += 1
    refs["annual_summary_start"] = row
    for code, name in STUDIOS:
        s_row = refs["studios"][code]
        label_cell(ws, row, 2, code, bold=True)
        label_cell(ws, row, 3, name)
        # 2026 = sum of cols D:O (months 1-12)
        formula_cell(ws, row, 4, f"=SUM(D{s_row}:O{s_row})", fmt=CURR)
        # 2027 = cols P:AA (months 13-24)
        formula_cell(ws, row, 5, f"=SUM(P{s_row}:AA{s_row})", fmt=CURR)
        # 2028 = cols AB:AM (months 25-36)
        formula_cell(ws, row, 6, f"=SUM(AB{s_row}:AM{s_row})", fmt=CURR)
        # 2027 YoY % = (2027-2026)/2026
        formula_cell(ws, row, 7,
                     f"=IFERROR((E{row}-D{row})/D{row},0)",
                     fmt=PCT)
        formula_cell(ws, row, 8,
                     f"=IFERROR((F{row}-E{row})/E{row},0)",
                     fmt=PCT)
        row += 1
    # Total annual row
    refs["annual_total_row"] = row
    label_cell(ws, row, 2, "TOTAL", bold=True)
    label_cell(ws, row, 3, "All Studios", bold=True)
    for c in range(2, 4):
        ws.cell(row=row, column=c).fill = light_fill
    formula_cell(ws, row, 4,
                 f"=SUM(D{refs['annual_summary_start']}:D{row - 1})",
                 fmt=CURR, bold=True, fill=light_fill)
    formula_cell(ws, row, 5,
                 f"=SUM(E{refs['annual_summary_start']}:E{row - 1})",
                 fmt=CURR, bold=True, fill=light_fill)
    formula_cell(ws, row, 6,
                 f"=SUM(F{refs['annual_summary_start']}:F{row - 1})",
                 fmt=CURR, bold=True, fill=light_fill)
    formula_cell(ws, row, 7, f"=IFERROR((E{row}-D{row})/D{row},0)",
                 fmt=PCT, bold=True, fill=light_fill)
    formula_cell(ws, row, 8, f"=IFERROR((F{row}-E{row})/E{row},0)",
                 fmt=PCT, bold=True, fill=light_fill)

    wb._mighty_refs["sales_forecast"] = refs
    return refs


# ============================================================
# OPEX DATA AGGREGATION
# ============================================================

# Maps accountant P&L line totals to Cash Flow OpEx categories
PL_TO_CF_CATEGORY = {
    "property": [
        "Total 700000 Property Costs", "Total for 700000 Property Costs",
    ],
    "staff": [
        "Total 602000 Payroll", "Total for 602000 Payroll",
    ],
    "utilities": [
        "Total 616000 Utilities", "Total for 616000 Utilities",
    ],
    "marketing": [
        "Total 601000 Sales & Marketing", "Total for 601000 Sales & Marketing",
    ],
    "admin": [
        "603000 Software & Web Services",
        "608000 Insurance",
        "610000 Office Supplies & General Expense",
        "610100 Furniture & Equipment",
        "611000 Shipping & postage",
        "613000 Bank fees & Service Charges",
        "615000 Parking Lot Rental",
    ],
    "professional": [
        "Total 604000 Professional Fees", "Total for 604000 Professional Fees",
    ],
    "travel": [
        "605000 Travel (Airfare/hotel/ground trans/etc)",
        "606000 Meals", "607000 Entertainment",
    ],
    "finance": [
        "506000 Merchant Account Fees",
        "Total Cost of Goods Sold", "Total for Cost of goods sold",
    ],
    "startup": [
        "630000 Studio Start Up Costs",
        "Total 630000 Studio Start Up Costs", "Total for 630000 Studio Start Up Costs",
    ],
    "taxes": [
        "902000 Taxes Paid", "903000 Property taxes",
        "Total 902000 Taxes Paid", "Total for 902000 Taxes Paid",
    ],
}

# Display order for Cash Flow OpEx rows
CF_OPEX_DISPLAY = [
    ("property", "Property Costs"),
    ("staff", "Staff Costs"),
    ("utilities", "Utilities"),
    ("marketing", "Marketing & Promotion"),
    ("admin", "Administrative & G&A"),
    ("professional", "Professional Fees"),
    ("travel", "Travel & Meals"),
    ("finance", "Merchant Fees & COGS"),
    ("startup", "Studio Start Up Costs"),
    ("taxes", "Taxes"),
]

# Forecast OpEx categories (from override JSON) → CF category
OPEX_OVERRIDE_TO_CF = {
    "property": "property",
    "staff": "staff",
    "utilities": "utilities",
    "marketing": "marketing",
    "admin": "admin",
    "professional_fees": "professional",
    "travel": "travel",
    "finance": "finance",   # merchant+cogs
    "taxes": "taxes",
}


def compute_actuals_opex(data):
    """Return {month_label: {cf_category: value}} for closed months."""
    pl = data["committed"].get("pl", {})
    result = {}
    for month_label, lines in pl.items():
        cats = {cat: 0.0 for cat, _ in CF_OPEX_DISPLAY}
        for cat, patterns in PL_TO_CF_CATEGORY.items():
            for pat in patterns:
                if pat in lines:
                    v = lines[pat]
                    if isinstance(v, (int, float)):
                        cats[cat] += abs(float(v))
        result[month_label] = cats
    return result


def compute_forecast_opex(data):
    """Return {month_key: {cf_category: value}} for forecast months.

    Sums across studios from merged opex_assumptions.
    Finance is added separately at the CF level (computed from revenue × merchant+cogs%).
    """
    merged_opex = data["merged"].get("opex_assumptions", {})
    forecast_keys = MONTH_KEYS[FIRST_FORECAST_IDX:]  # May 2026 onwards
    result = {mk: {cat: 0.0 for cat, _ in CF_OPEX_DISPLAY} for mk in forecast_keys}

    for studio, cats in merged_opex.items():
        if not isinstance(cats, dict):
            continue
        for cat, months in cats.items():
            cf_cat = OPEX_OVERRIDE_TO_CF.get(cat)
            if not cf_cat or not isinstance(months, dict):
                continue
            for mk, v in months.items():
                if mk in result:
                    try:
                        result[mk][cf_cat] += float(v)
                    except (ValueError, TypeError):
                        pass
    return result


def actuals_total_cash_sales(data):
    """Return {month_label: total cash sales} from monthly_sales (the QBO actuals)."""
    ms = data["committed"].get("monthly_sales", {})
    actuals = {}
    for mk in MONTH_KEYS[:FIRST_FORECAST_IDX]:  # Jan-Apr 2026
        actuals[mk] = float(ms.get(mk, 0))
    return actuals


def actuals_other_lines(data):
    """Extract depreciation, interest, other-income lines from accountant P&L."""
    pl = data["committed"].get("pl", {})
    lines = {
        "depreciation": ["810000 Depreciation", "Total 810000 Depreciation"],
        "interest": ["901000 Interest Expense/(Income)"],
        "other_income": ["900000 Other Expense/(Income)"],
    }
    result = {}
    for month_label, plines in pl.items():
        month_dict = {}
        for key, patterns in lines.items():
            for pat in patterns:
                if pat in plines:
                    v = plines[pat]
                    if isinstance(v, (int, float)):
                        month_dict[key] = month_dict.get(key, 0) + float(v)
        result[month_label] = month_dict
    return result


def month_label_to_key(label):
    """Convert 'Jan 2026' → '2026-01'."""
    mo_map = {"Jan":"01","Feb":"02","Mar":"03","Apr":"04","May":"05","Jun":"06",
              "Jul":"07","Aug":"08","Sep":"09","Oct":"10","Nov":"11","Dec":"12"}
    parts = label.split()
    if len(parts) != 2:
        return None
    return f"{parts[1]}-{mo_map.get(parts[0], '00')}"


def build_cash_flow_forecast(wb, data):
    """
    Cash Flow Forecast tab. Mirrors Streamlit Cash Flow page.

    Sections:
      1. Cash Inflows (Total Cash Sales)
      2. Operating Outflows (10 OpEx categories)
      3. Net Cash from Operations
      4. Investing (Equipment, Leasehold, Deposits)
      5. Financing (Loan Proceeds, Loan Repayments, Intercompany)
      6. Net Change in Cash
      7. Beginning / Ending Cash Balance
    """
    ws = wb.create_sheet("Cash Flow Forecast")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D6"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    for i in range(36):
        ws.column_dimensions[get_column_letter(4 + i)].width = 12

    # Title
    ws["B2"] = "Cash Flow Forecast"
    ws["B2"].font = title_font
    ws["B3"] = ("Actuals: Jan-Apr 2026 (light green). Forecast: May 2026 onwards. "
                "Net Change formulas sum operations + investing + financing.")
    ws["B3"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    HEADER_ROW = 5
    section_bar(ws, HEADER_ROW - 1, 2, 39, "Monthly Cash Flow")
    # Month headers
    label_cell(ws, HEADER_ROW, 2, "Line Item", bold=True)
    ws.cell(row=HEADER_ROW, column=2).fill = header_fill
    ws.cell(row=HEADER_ROW, column=2).font = header_font
    label_cell(ws, HEADER_ROW, 3, "Annual Total")
    ws.cell(row=HEADER_ROW, column=3).fill = header_fill
    ws.cell(row=HEADER_ROW, column=3).font = header_font
    ws.cell(row=HEADER_ROW, column=3).alignment = center_align
    for i, label in enumerate(MONTH_LABELS):
        cell = ws.cell(row=HEADER_ROW, column=4 + i, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    sf_refs = wb._mighty_refs["sales_forecast"]
    actuals_opex = compute_actuals_opex(data)
    forecast_opex = compute_forecast_opex(data)
    cash_actuals = actuals_total_cash_sales(data)

    row = HEADER_ROW + 1

    # ---------- INFLOWS ----------
    section_bar(ws, row, 2, 39, "CASH INFLOWS")
    row += 1
    refs = {"cash_sales_row": row}

    label_cell(ws, row, 2, "Total Cash Sales", bold=True)
    # Each month: actuals for closed months, formula =Sales Forecast TOTAL for forecast
    for i, mk in enumerate(MONTH_KEYS):
        col = get_column_letter(4 + i)
        if i <= LAST_ACTUALS_IDX:
            # Hard-coded actual
            actual_value_cell(ws, row, 4 + i, cash_actuals.get(mk, 0), fmt=CURR, bold=True)
        else:
            # Formula: reference Sales Forecast TOTAL row, same month column (cols match: both use col 4+i)
            formula_cell(ws, row, 4 + i,
                         f"='Sales Forecast'!{col}{sf_refs['total_row']}",
                         fmt=CURR, bold=True)
    # Annual total in col C (placeholder — could be SUMIFS)
    # Skip for now; can add later
    row += 2

    # ---------- OPERATING OUTFLOWS ----------
    section_bar(ws, row, 2, 39, "OPERATING OUTFLOWS")
    row += 1
    refs["opex_rows"] = {}
    refs["opex_start_row"] = row
    for cat, label in CF_OPEX_DISPLAY:
        label_cell(ws, row, 2, label)
        for i, mk in enumerate(MONTH_KEYS):
            if i <= LAST_ACTUALS_IDX:
                # Use actuals month label form
                month_label = MONTH_LABELS[i]
                v = actuals_opex.get(month_label, {}).get(cat, 0)
                actual_value_cell(ws, row, 4 + i, v, fmt=CURR)
            else:
                # Forecast: use the precomputed forecast_opex
                # For "finance" specifically, we compute as % of cash sales (matches Streamlit)
                if cat == "finance":
                    col = get_column_letter(4 + i)
                    # =Total Cash Sales × (Merchant+COGS rate)
                    formula_cell(ws, row, 4 + i,
                                 f"={col}{refs['cash_sales_row']}*"
                                 f"(Assumptions!$C${wb._mighty_refs['assumptions']['merchant_row']}"
                                 f"+Assumptions!$C${wb._mighty_refs['assumptions']['cogs_row']})",
                                 fmt=CURR)
                else:
                    v = forecast_opex.get(mk, {}).get(cat, 0)
                    input_cell(ws, row, 4 + i, v, fmt=CURR)
        refs["opex_rows"][cat] = row
        row += 1

    # Total Operating Expenses row
    refs["opex_total_row"] = row
    label_cell(ws, row, 2, "Total Operating Expenses", bold=True)
    for i in range(36):
        col = get_column_letter(4 + i)
        formula_cell(ws, row, 4 + i,
                     f"=SUM({col}{refs['opex_start_row']}:{col}{row - 1})",
                     fmt=CURR, bold=True, fill=light_fill)
    row += 1

    # Net Cash from Operations
    refs["net_ops_row"] = row
    label_cell(ws, row, 2, "Net Cash from Operations", bold=True)
    for i in range(36):
        col = get_column_letter(4 + i)
        formula_cell(ws, row, 4 + i,
                     f"={col}{refs['cash_sales_row']}-{col}{refs['opex_total_row']}",
                     fmt=CURR, bold=True, fill=green_fill)
    row += 2

    # ---------- INVESTING ----------
    section_bar(ws, row, 2, 39, "INVESTING ACTIVITIES")
    row += 1
    refs["investing_start"] = row
    invest_lines = [
        ("equipment", "Equipment & Furniture"),
        ("leasehold", "Leasehold Improvements"),
        ("deposits", "Deposits"),
    ]
    # For actuals, sum from SCF data. For forecast, default to 0 (CapEx handled separately).
    scf = data["committed"].get("scf", {})
    for key, label in invest_lines:
        label_cell(ws, row, 2, label)
        for i, mk in enumerate(MONTH_KEYS):
            if i <= LAST_ACTUALS_IDX:
                month_label = MONTH_LABELS[i]
                v = 0
                # SCF accounts (negative for outflow)
                scf_patterns = {
                    "equipment": ["151000", "152000", "153000", "154000"],
                    "leasehold": ["155"],
                    "deposits": ["171000"],
                }
                scf_month = scf.get(month_label, {})
                for line_label, val in scf_month.items():
                    for pat in scf_patterns.get(key, []):
                        if pat in line_label and isinstance(val, (int, float)):
                            v += float(val)
                            break
                actual_value_cell(ws, row, 4 + i, v, fmt=CURR)
            else:
                input_cell(ws, row, 4 + i, 0, fmt=CURR)
        row += 1
    refs["investing_end"] = row - 1

    # Net Investing
    refs["net_invest_row"] = row
    label_cell(ws, row, 2, "Net Investing", bold=True)
    for i in range(36):
        col = get_column_letter(4 + i)
        formula_cell(ws, row, 4 + i,
                     f"=SUM({col}{refs['investing_start']}:{col}{refs['investing_end']})",
                     fmt=CURR, bold=True, fill=light_fill)
    row += 2

    # ---------- FINANCING ----------
    section_bar(ws, row, 2, 39, "FINANCING ACTIVITIES")
    row += 1
    refs["financing_start"] = row
    finance_lines = [
        ("proceeds", "Loan Proceeds"),
        ("repayments", "Loan Repayments"),
        ("intercompany", "Intercompany / Owner Contributions"),
    ]
    for key, label in finance_lines:
        label_cell(ws, row, 2, label)
        for i, mk in enumerate(MONTH_KEYS):
            if i <= LAST_ACTUALS_IDX:
                month_label = MONTH_LABELS[i]
                v = 0
                scf_month = scf.get(month_label, {})
                if key in ("proceeds", "repayments"):
                    loan_patterns = ["242", "243000", "244000"]
                    for line_label, val in scf_month.items():
                        for pat in loan_patterns:
                            if pat in line_label and isinstance(val, (int, float)):
                                fv = float(val)
                                if key == "proceeds" and fv > 0:
                                    v += fv
                                elif key == "repayments" and fv < 0:
                                    v += fv
                                break
                else:  # intercompany
                    inter_patterns = ["241000", "251000", "Due to", "Opening balance"]
                    for line_label, val in scf_month.items():
                        for pat in inter_patterns:
                            if pat in line_label and isinstance(val, (int, float)):
                                v += float(val)
                                break
                actual_value_cell(ws, row, 4 + i, v, fmt=CURR)
            else:
                input_cell(ws, row, 4 + i, 0, fmt=CURR)
        row += 1
    refs["financing_end"] = row - 1

    # Net Financing
    refs["net_finance_row"] = row
    label_cell(ws, row, 2, "Net Financing", bold=True)
    for i in range(36):
        col = get_column_letter(4 + i)
        formula_cell(ws, row, 4 + i,
                     f"=SUM({col}{refs['financing_start']}:{col}{refs['financing_end']})",
                     fmt=CURR, bold=True, fill=light_fill)
    row += 2

    # ---------- NET CHANGE / CASH BALANCE ----------
    section_bar(ws, row, 2, 39, "NET CHANGE IN CASH")
    row += 1
    refs["net_change_row"] = row
    label_cell(ws, row, 2, "Net Change in Cash", bold=True)
    for i in range(36):
        col = get_column_letter(4 + i)
        formula_cell(ws, row, 4 + i,
                     f"={col}{refs['net_ops_row']}+{col}{refs['net_invest_row']}+"
                     f"{col}{refs['net_finance_row']}",
                     fmt=CURR, bold=True, fill=green_fill)
    row += 1

    # Beginning + Ending Cash Balance
    # Use BS data for actual beginning balance (Jan 2026 opening)
    bs = data["committed"].get("bs", {})
    # Find opening cash from earliest month's "101000 Operating Cash" or similar
    opening_cash = 0
    if "Jan 2026" in bs:
        jan_bs = bs["Jan 2026"]
        for line, val in jan_bs.items():
            if "101000" in line or ("Cash" in line and "operating" in line.lower()):
                if isinstance(val, (int, float)):
                    opening_cash = float(val)
                    break

    refs["beg_cash_row"] = row
    label_cell(ws, row, 2, "Beginning Cash Balance")
    actual_value_cell(ws, row, 4, opening_cash, fmt=CURR, bold=True)
    for i in range(1, 36):
        prev_col = get_column_letter(4 + i - 1)
        # Beginning = prev ending
        formula_cell(ws, row, 4 + i,
                     f"={prev_col}{row + 1}",
                     fmt=CURR, bold=True)
    row += 1

    refs["end_cash_row"] = row
    label_cell(ws, row, 2, "Ending Cash Balance", bold=True)
    for i in range(36):
        col = get_column_letter(4 + i)
        formula_cell(ws, row, 4 + i,
                     f"={col}{refs['beg_cash_row']}+{col}{refs['net_change_row']}",
                     fmt=CURR, bold=True, fill=green_fill)

    wb._mighty_refs["cash_flow"] = refs
    return refs


# ============================================================
# CONSOLIDATED P&L
# ============================================================

# Mirror Streamlit SUMMARY_ROWS
SUMMARY_ROWS_PL = [
    # (display_name, [pl_label_candidates], is_bold, is_revenue_pos)
    ("Session Revenue", ["Total 401000 Sessions", "Total for 401000 Sessions"], False, True),
    ("Breakage Revenue", ["Total 403000 Breakage Revenue", "Total for 403000 Breakage Revenue"], False, True),
    ("Retail Sales", ["404000 Retail Sales"], False, True),
    ("Refunds", ["406000 Refunds"], False, True),
    ("Discounts", ["407000 Discounts"], False, True),
    ("Total Revenue", ["Total Income", "Total for Income"], True, True),
    ("COGS & Merchant Fees", ["Total Cost of Goods Sold", "Total for Cost of Goods Sold"], False, False),
    ("Gross Profit", ["Gross Profit"], True, False),
    ("Marketing", ["Total 601000 Sales & Marketing", "Total for 601000 Sales & Marketing"], False, False),
    ("Payroll", ["Total 602000 Payroll", "Total for 602000 Payroll"], False, False),
    ("Software", ["603000 Software & Web Services"], False, False),
    ("Professional Fees", ["Total 604000 Professional Fees", "Total for 604000 Professional Fees"], False, False),
    ("Utilities", ["Total 616000 Utilities", "Total for 616000 Utilities"], False, False),
    ("Property Costs", ["Total 700000 Property Costs", "Total for 700000 Property Costs"], False, False),
    ("Total Operating Expenses", ["Total Expenses", "Total for Expenses"], True, False),
    ("Net Operating Income", ["Net Operating Income"], True, False),
    ("Depreciation", ["810000 Depreciation"], False, False),
    ("Interest Expense", ["901000 Interest Expense/(Income)"], False, False),
    ("Taxes Paid", ["902000 Taxes Paid"], False, False),
    ("Property Taxes", ["903000 Property taxes"], False, False),
    ("Total Other Expenses", ["Total Other Expenses", "Total for Other Expenses"], True, False),
    ("Net Income", ["Net Income"], True, False),
]


def get_pl_actual(pl_dict, label_candidates):
    """Return actual value for a P&L line by trying each candidate label."""
    for lbl in label_candidates:
        if lbl in pl_dict:
            v = pl_dict[lbl]
            if isinstance(v, (int, float)):
                return float(v)
    return 0.0


def build_pl_consolidated(wb, data):
    """
    P&L Summary tab. Consolidated P&L by month.

    Layout:
      - Col B: Line Item
      - Cols D..AM: 36 months (Jan 2026 - Dec 2028)
      - Rows: SUMMARY_ROWS_PL structure
      - Actuals (Jan-Apr 2026): hard-coded from c["pl"]
      - Forecast (May 2026+): formulas referencing Sales Forecast, Assumptions, and CF tab
    """
    ws = wb.create_sheet("P&L")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D6"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    for i in range(36):
        ws.column_dimensions[get_column_letter(4 + i)].width = 12

    ws["B2"] = "Profit & Loss (Consolidated)"
    ws["B2"].font = title_font
    ws["B3"] = "Actuals Jan-Apr 2026. Forecast May 2026+ uses formulas from Sales Forecast and Assumptions."
    ws["B3"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    HEADER_ROW = 5
    section_bar(ws, HEADER_ROW - 1, 2, 39, "Monthly P&L")
    label_cell(ws, HEADER_ROW, 2, "Line Item", bold=True)
    ws.cell(row=HEADER_ROW, column=2).fill = header_fill
    ws.cell(row=HEADER_ROW, column=2).font = header_font
    for i, lbl in enumerate(MONTH_LABELS):
        cell = ws.cell(row=HEADER_ROW, column=4 + i, value=lbl)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    pl_data = data["committed"].get("pl", {})
    a_refs = wb._mighty_refs["assumptions"]
    sf_refs = wb._mighty_refs["sales_forecast"]
    cf_refs = wb._mighty_refs["cash_flow"]

    row = HEADER_ROW + 1
    refs = {"rows": {}}

    # Precompute average retail and depreciation/interest from actuals
    retail_avg = 0
    depr_avg = 0
    int_sched = data["committed"].get("interest_schedule", {})
    other_avgs = {}

    if pl_data:
        retails = [get_pl_actual(pl_data[m], ["404000 Retail Sales"]) for m in pl_data]
        retail_avg = sum(retails) / max(len(retails), 1)
        deprs = [get_pl_actual(pl_data[m], ["810000 Depreciation"]) for m in pl_data]
        depr_avg = sum(deprs) / max(len(deprs), 1)

    for display, candidates, is_bold, is_rev in SUMMARY_ROWS_PL:
        label_cell(ws, row, 2, display, bold=is_bold)
        if is_bold:
            for c in range(2, 40):
                ws.cell(row=row, column=c).fill = light_fill

        for i, mk in enumerate(MONTH_KEYS):
            col = get_column_letter(4 + i)
            if i <= LAST_ACTUALS_IDX:
                # Actuals
                month_label = MONTH_LABELS[i]
                v = get_pl_actual(pl_data.get(month_label, {}), candidates)
                actual_value_cell(ws, row, 4 + i, v, fmt=CURR, bold=is_bold)
            else:
                # Forecast formulas
                formula = build_pl_forecast_formula(
                    display, i, col, sf_refs, a_refs, cf_refs,
                    retail_avg, depr_avg, int_sched.get(mk, 0),
                    pl_data, refs,
                )
                if formula is not None:
                    formula_cell(ws, row, 4 + i, formula, fmt=CURR, bold=is_bold,
                                 fill=(light_fill if is_bold else None))
                else:
                    # Default to zero
                    formula_cell(ws, row, 4 + i, 0, fmt=CURR, bold=is_bold,
                                 fill=(light_fill if is_bold else None))
        refs["rows"][display] = row
        row += 1

    wb._mighty_refs["pl"] = refs
    return refs


def build_pl_forecast_formula(display, month_idx, col, sf_refs, a_refs, cf_refs,
                                retail_avg, depr_avg, interest_val, pl_data, pl_refs):
    """Return Excel formula string for a forecast P&L cell."""
    # Sales Forecast TOTAL row gives gross cash sales for this month
    sf_tot = f"'Sales Forecast'!{col}{sf_refs['total_row']}"

    # Earned & breakage curves require summing over 7-month lag
    # earned_sum = SUMPRODUCT of {sales[m-0..6]} × {earned[0..6]}
    # For simplicity: build the lag sum
    earned_lags = []
    breakage_lags = []
    a_earned_row = a_refs["earned_row"]
    a_breakage_row = a_refs["breakage_row"]
    # For each lag (0-6), the source column = current month idx - lag
    # That's column (4 + month_idx - lag)
    for lag in range(7):
        src_idx = month_idx - lag
        if src_idx < 0:
            continue  # No data that far back
        src_col = get_column_letter(4 + src_idx)
        lag_col = get_column_letter(3 + lag)  # Assumptions cols C..I for lags 0..6
        earned_lags.append(
            f"'Sales Forecast'!{src_col}{sf_refs['total_row']}*Assumptions!${lag_col}${a_earned_row}"
        )
        breakage_lags.append(
            f"'Sales Forecast'!{src_col}{sf_refs['total_row']}*Assumptions!${lag_col}${a_breakage_row}"
        )
    earned_formula = "+".join(earned_lags) if earned_lags else "0"
    breakage_formula = "+".join(breakage_lags) if breakage_lags else "0"

    if display == "Session Revenue":
        return f"={earned_formula}"
    if display == "Breakage Revenue":
        return f"={breakage_formula}"
    if display == "Retail Sales":
        return f"={retail_avg}"
    if display == "Refunds":
        # Refunds = (Session + Breakage + Retail) × refund_pct
        return (f"=({earned_formula}+{breakage_formula}+{retail_avg})*"
                f"Assumptions!$C${a_refs['refund_row']}")
    if display == "Discounts":
        return (f"=({earned_formula}+{breakage_formula}+{retail_avg})*"
                f"Assumptions!$C${a_refs['discount_row']}")
    if display == "Total Revenue":
        # Sum of Session + Breakage + Retail + Refunds + Discounts
        rows = pl_refs["rows"]
        return (f"={col}{rows.get('Session Revenue', '#REF!')}"
                f"+{col}{rows.get('Breakage Revenue', '#REF!')}"
                f"+{col}{rows.get('Retail Sales', '#REF!')}"
                f"+{col}{rows.get('Refunds', '#REF!')}"
                f"+{col}{rows.get('Discounts', '#REF!')}")
    if display == "COGS & Merchant Fees":
        rows = pl_refs["rows"]
        # Pull from CF Merchant Fees & COGS row OR compute
        return f"={col}{cf_refs['opex_rows']['finance']}"
    if display == "Gross Profit":
        rows = pl_refs["rows"]
        return f"={col}{rows.get('Total Revenue', '#REF!')}-{col}{rows.get('COGS & Merchant Fees', '#REF!')}"
    # OpEx lines pull from Cash Flow Forecast (consolidated) rows
    opex_map = {
        "Marketing": cf_refs["opex_rows"].get("marketing"),
        "Payroll": cf_refs["opex_rows"].get("staff"),
        "Software": None,  # subset of admin; approximate via 30% of admin
        "Professional Fees": cf_refs["opex_rows"].get("professional"),
        "Utilities": cf_refs["opex_rows"].get("utilities"),
        "Property Costs": cf_refs["opex_rows"].get("property"),
    }
    if display in opex_map and opex_map[display]:
        return f"='Cash Flow Forecast'!{col}{opex_map[display]}"
    if display == "Software":
        # Estimate: avg actual software / avg actual admin × admin forecast
        sw_avg = 0
        admin_avg = 0
        for m, lines in pl_data.items():
            sw_avg += get_pl_actual(lines, ["603000 Software & Web Services"])
            admin_avg += (
                get_pl_actual(lines, ["603000 Software & Web Services"])
                + get_pl_actual(lines, ["608000 Insurance"])
                + get_pl_actual(lines, ["610000 Office Supplies & General Expense"])
            )
        if admin_avg > 0:
            ratio = sw_avg / admin_avg
            return f"='Cash Flow Forecast'!{col}{cf_refs['opex_rows']['admin']}*{ratio:.4f}"
        return "0"
    if display == "Total Operating Expenses":
        rows = pl_refs["rows"]
        return (f"={col}{rows['Marketing']}+{col}{rows['Payroll']}+{col}{rows['Software']}"
                f"+{col}{rows['Professional Fees']}+{col}{rows['Utilities']}+{col}{rows['Property Costs']}")
    if display == "Net Operating Income":
        rows = pl_refs["rows"]
        return f"={col}{rows['Gross Profit']}-{col}{rows['Total Operating Expenses']}"
    if display == "Depreciation":
        return f"={depr_avg}"
    if display == "Interest Expense":
        return f"={interest_val}"
    if display == "Taxes Paid":
        return "0"
    if display == "Property Taxes":
        # Use avg actual or 0
        return "0"
    if display == "Total Other Expenses":
        rows = pl_refs["rows"]
        return (f"={col}{rows['Depreciation']}+{col}{rows['Interest Expense']}"
                f"+{col}{rows['Taxes Paid']}+{col}{rows['Property Taxes']}")
    if display == "Net Income":
        rows = pl_refs["rows"]
        return f"={col}{rows['Net Operating Income']}-{col}{rows['Total Other Expenses']}"
    return None


# ============================================================
# CASH, DEBT & EQUITY
# ============================================================

def build_cash_debt_equity(wb, data):
    """
    Cash, Debt & Equity tab.

    Sections:
      1. Cash position (timeline from CF)
      2. Loan amortization schedules (one row per loan, 36 columns)
      3. Total debt timeline
      4. Owner equity (intercompany / opening balance)
      5. Debt service forecast (monthly P+I from amortization)
    """
    ws = wb.create_sheet("Cash, Debt & Equity")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D5"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    for i in range(36):
        ws.column_dimensions[get_column_letter(4 + i)].width = 12

    ws["B2"] = "Cash, Debt & Equity"
    ws["B2"].font = title_font

    HEADER_ROW = 4
    label_cell(ws, HEADER_ROW, 2, "Item", bold=True)
    ws.cell(row=HEADER_ROW, column=2).fill = header_fill
    ws.cell(row=HEADER_ROW, column=2).font = header_font
    for i, lbl in enumerate(MONTH_LABELS):
        cell = ws.cell(row=HEADER_ROW, column=4 + i, value=lbl)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    cf_refs = wb._mighty_refs["cash_flow"]
    a_refs = wb._mighty_refs["assumptions"]

    row = HEADER_ROW + 1

    # ----- Section 1: Cash Position -----
    section_bar(ws, row, 2, 39, "CASH POSITION")
    row += 1
    label_cell(ws, row, 2, "Ending Cash Balance", bold=True)
    for i in range(36):
        col = get_column_letter(4 + i)
        formula_cell(ws, row, 4 + i,
                     f"='Cash Flow Forecast'!{col}{cf_refs['end_cash_row']}",
                     fmt=CURR, bold=True, fill=green_fill)
    row += 2

    # ----- Section 2: Loan Balances (per-loan amortization) -----
    section_bar(ws, row, 2, 39, "DEBT SCHEDULE — Loan Balances by Month")
    row += 1
    refs = {"loan_rows": {}}
    refs["loan_start"] = row

    # For each loan: starting balance (Apr 2026) hard-coded, then formula:
    # If amortizing: max(prev - monthly_pmt, 0)
    # If interest-only: prev (flat)
    loans = data["baseline"].get("loans", [])
    # Get actual balance per loan per actual month from BS data
    bs = data["committed"].get("bs", {})
    # Map loan name → bs key (best-effort)
    bs_key_map = {
        "MindBody Loan - SM": ["242001 MindBody Loan:MindBody Loan - SM"],
        "MindBody Loan - PH": ["242002 MindBody Loan:MindBody Loan - PH"],
        "MindBody Loan - LF": ["242003 MindBody Loan:MindBody Loan - LF"],
        "MindBody Loan - MR": ["242004 MindBody Loan:MindBody Loan - MR"],
        "Samson Loan": ["243000 Samson Loan"],
        "Specialty Capital Loan": ["244000 Specialty Capital Loan"],
        "Norbrook Inc Loan": ["246000 Norbrook"],
    }

    def extract_bs_value(month_label, patterns):
        if month_label not in bs:
            return None
        for line, val in bs[month_label].items():
            for p in patterns:
                if p in line and isinstance(val, (int, float)):
                    return float(val)
        return None

    for loan in loans:
        name = loan.get("name", "?")
        label_cell(ws, row, 2, name, bold=True)
        avg_pmt = float(loan.get("avg_monthly_payment", 0))
        is_amortizing = avg_pmt > 0
        # Actuals: pull from BS by month
        prev_col = None
        for i, mk in enumerate(MONTH_KEYS):
            col = get_column_letter(4 + i)
            month_label = MONTH_LABELS[i]
            if i <= LAST_ACTUALS_IDX:
                bs_val = extract_bs_value(month_label, bs_key_map.get(name, []))
                if bs_val is None:
                    bs_val = float(loan.get("current_balance", 0)) if i == LAST_ACTUALS_IDX else 0
                actual_value_cell(ws, row, 4 + i, bs_val, fmt=CURR)
            else:
                # Forecast
                if is_amortizing:
                    formula_cell(ws, row, 4 + i,
                                 f"=MAX({prev_col}{row}-{avg_pmt},0)",
                                 fmt=CURR)
                else:
                    # Interest-only: flat
                    formula_cell(ws, row, 4 + i,
                                 f"={prev_col}{row}",
                                 fmt=CURR)
            prev_col = col
        refs["loan_rows"][name] = row
        row += 1

    # Total Debt
    refs["total_debt_row"] = row
    label_cell(ws, row, 2, "TOTAL DEBT", bold=True)
    for i in range(36):
        col = get_column_letter(4 + i)
        formula_cell(ws, row, 4 + i,
                     f"=SUM({col}{refs['loan_start']}:{col}{row - 1})",
                     fmt=CURR, bold=True, fill=light_fill)
    row += 2

    # ----- Section 3: Debt Service Forecast -----
    section_bar(ws, row, 2, 39, "MONTHLY DEBT SERVICE")
    row += 1
    refs["debt_service_start"] = row
    for loan in loans:
        name = loan.get("name", "?")
        avg_pmt = float(loan.get("avg_monthly_payment", 0))
        label_cell(ws, row, 2, name)
        for i in range(36):
            col = get_column_letter(4 + i)
            if i <= LAST_ACTUALS_IDX:
                # No actuals broken out per loan from SCF; use 0 or approximate
                actual_value_cell(ws, row, 4 + i, 0, fmt=CURR)
            else:
                if avg_pmt > 0:
                    # Payment = avg_pmt, but only if loan balance > 0 in that month
                    loan_balance_row = refs["loan_rows"][name]
                    formula_cell(ws, row, 4 + i,
                                 f"=IF({col}{loan_balance_row}>0,{avg_pmt},0)",
                                 fmt=CURR)
                else:
                    # Interest-only: balance × annual rate / 12
                    loan_balance_row = refs["loan_rows"][name]
                    rate = float(loan.get("rate", 0))
                    formula_cell(ws, row, 4 + i,
                                 f"={col}{loan_balance_row}*{rate}/12",
                                 fmt=CURR)
        row += 1
    refs["debt_service_end"] = row - 1

    # Total Debt Service
    refs["total_debt_service_row"] = row
    label_cell(ws, row, 2, "TOTAL DEBT SERVICE", bold=True)
    for i in range(36):
        col = get_column_letter(4 + i)
        formula_cell(ws, row, 4 + i,
                     f"=SUM({col}{refs['debt_service_start']}:{col}{row - 1})",
                     fmt=CURR, bold=True, fill=light_fill)
    row += 2

    # ----- Section 4: Owner Equity / Intercompany -----
    section_bar(ws, row, 2, 39, "OWNER EQUITY / INTERCOMPANY")
    row += 1
    label_cell(ws, row, 2, "Intercompany Net Position", bold=True)
    # From CF intercompany row
    inter_row_idx = cf_refs["financing_end"]  # Last financing line = intercompany
    for i in range(36):
        col = get_column_letter(4 + i)
        formula_cell(ws, row, 4 + i,
                     f"='Cash Flow Forecast'!{col}{inter_row_idx}",
                     fmt=CURR)

    wb._mighty_refs["cash_debt_equity"] = refs
    return refs


# ============================================================
# CAPEX
# ============================================================

def build_capex(wb, data):
    """CapEx project schedule tab."""
    ws = wb.create_sheet("CapEx")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    ws["B2"] = "CapEx & Studio Buildout"
    ws["B2"].font = title_font
    ws["B3"] = "Add planned capital expenditures here. Outflows feed Cash Flow Forecast investing section."
    ws["B3"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    HEADER_ROW = 5
    section_bar(ws, HEADER_ROW - 1, 2, 7, "Projects")
    header_row(ws, HEADER_ROW, 2, ["Project", "Type", "Location", "Total Budget",
                                     "Start Month", "Duration (mo)"])

    row = HEADER_ROW + 1
    capex_projects = data["committed"].get("capex_projects", [])
    if not capex_projects:
        # Show empty placeholder rows
        for _ in range(10):
            for c in range(2, 8):
                cell = ws.cell(row=row, column=c, value=None)
                cell.fill = input_fill
                cell.border = thin_border
            row += 1
    else:
        for proj in capex_projects:
            input_cell(ws, row, 2, proj.get("name", ""), fmt="@")
            input_cell(ws, row, 3, proj.get("location", ""), fmt="@")
            input_cell(ws, row, 4, proj.get("location", ""), fmt="@")
            input_cell(ws, row, 5, float(proj.get("total_budget", 0)), fmt=CURR)
            input_cell(ws, row, 6, proj.get("start_month", ""), fmt="@")
            input_cell(ws, row, 7, int(proj.get("duration_months", 1)), fmt=NUM)
            row += 1


# ============================================================
# QBO ACTUALS (raw reference)
# ============================================================

def build_qbo_actuals(wb, data):
    """Dump raw P&L, BS, SCF actuals from accountant for reference."""
    ws = wb.create_sheet("QBO Actuals")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 50
    for i in range(8):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14

    ws["B2"] = "QBO Actuals (Accountant-Booked)"
    ws["B2"].font = title_font
    ws["B3"] = "Source data from accountant. Read-only reference. Used in formulas via the P&L tab."
    ws["B3"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    row = 5
    for section_name, key in [("P&L", "pl"), ("Balance Sheet", "bs"), ("Cash Flow Statement", "scf")]:
        section_bar(ws, row, 2, 8, section_name)
        row += 1
        section_data = data["committed"].get(key, {})
        if not section_data:
            row += 2
            continue
        # Get all months sorted
        months = sorted(section_data.keys(),
                        key=lambda m: month_label_to_key(m) or "")
        # Headers
        label_cell(ws, row, 2, "Account", bold=True)
        for i, m in enumerate(months):
            cell = ws.cell(row=row, column=3 + i, value=m)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        row += 1
        # All unique line items
        all_lines = set()
        for m_data in section_data.values():
            all_lines.update(m_data.keys())
        # Sort lines (totals last would be ideal, but just alpha for now)
        for line in sorted(all_lines):
            label_cell(ws, row, 2, line)
            any_value = False
            for i, m in enumerate(months):
                v = section_data[m].get(line)
                if isinstance(v, (int, float)):
                    actual_value_cell(ws, row, 3 + i, float(v), fmt=CURR)
                    any_value = True
                else:
                    actual_value_cell(ws, row, 3 + i, 0, fmt=CURR)
            row += 1
        row += 2


# ============================================================
# PER-STUDIO P&L TABS
# ============================================================

def get_studio_pl_actual(studio_data, label_candidates, month_label):
    """Pull actual value for a per-studio P&L line."""
    month_dict = studio_data.get("data", {}).get(month_label, {})
    for lbl in label_candidates:
        if lbl in month_dict:
            v = month_dict[lbl]
            if isinstance(v, (int, float)):
                return float(v)
    return 0.0


def build_studio_pl(wb, data, studio_code, studio_name):
    """Build a single per-studio P&L tab matching the consolidated P&L structure."""
    ws = wb.create_sheet(f"{studio_code} P&L")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D6"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    for i in range(36):
        ws.column_dimensions[get_column_letter(4 + i)].width = 12

    ws["B2"] = f"{studio_name} ({studio_code}) — Profit & Loss"
    ws["B2"].font = title_font

    HEADER_ROW = 5
    section_bar(ws, HEADER_ROW - 1, 2, 39, "Monthly P&L")
    label_cell(ws, HEADER_ROW, 2, "Line Item", bold=True)
    ws.cell(row=HEADER_ROW, column=2).fill = header_fill
    ws.cell(row=HEADER_ROW, column=2).font = header_font
    for i, lbl in enumerate(MONTH_LABELS):
        cell = ws.cell(row=HEADER_ROW, column=4 + i, value=lbl)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    a_refs = wb._mighty_refs["assumptions"]
    sf_refs = wb._mighty_refs["sales_forecast"]
    studios = data["committed"].get("studios", {})
    studio_data = studios.get(studio_code, {"data": {}})
    # Get studio sales row in Sales Forecast tab
    studio_sales_row = sf_refs["studios"].get(studio_code)

    # OpEx forecast for this studio
    studio_opex = data["merged"].get("opex_assumptions", {}).get(studio_code, {})

    # Approximate retail for this studio (avg actual)
    retail_vals = [get_studio_pl_actual(studio_data, ["404000 Retail Sales"], m)
                   for m in studio_data.get("data", {})]
    retail_avg = sum(retail_vals) / max(len(retail_vals), 1) if retail_vals else 0

    row = HEADER_ROW + 1
    refs = {"rows": {}}

    # OpEx categories applicable per-studio
    studio_opex_map = {
        "Marketing": "marketing",
        "Payroll": "staff",
        "Software": None,  # part of admin
        "Professional Fees": "professional_fees",
        "Utilities": "utilities",
        "Property Costs": "property",
    }

    for display, candidates, is_bold, is_rev in SUMMARY_ROWS_PL:
        label_cell(ws, row, 2, display, bold=is_bold)
        if is_bold:
            for c in range(2, 40):
                ws.cell(row=row, column=c).fill = light_fill

        for i, mk in enumerate(MONTH_KEYS):
            col = get_column_letter(4 + i)
            if i <= LAST_ACTUALS_IDX:
                month_label = MONTH_LABELS[i]
                v = get_studio_pl_actual(studio_data, candidates, month_label)
                actual_value_cell(ws, row, 4 + i, v, fmt=CURR, bold=is_bold)
            else:
                formula = _studio_forecast_formula(
                    display, i, col, studio_sales_row, a_refs, sf_refs,
                    studio_opex, studio_opex_map, retail_avg, refs,
                )
                if formula is not None:
                    formula_cell(ws, row, 4 + i, formula, fmt=CURR, bold=is_bold,
                                 fill=(light_fill if is_bold else None))
                else:
                    formula_cell(ws, row, 4 + i, 0, fmt=CURR, bold=is_bold,
                                 fill=(light_fill if is_bold else None))
        refs["rows"][display] = row
        row += 1

    if not hasattr(wb, "_mighty_studio_refs"):
        wb._mighty_studio_refs = {}
    wb._mighty_studio_refs[studio_code] = refs
    return refs


def _studio_forecast_formula(display, month_idx, col, sales_row, a_refs, sf_refs,
                               studio_opex, opex_map, retail_avg, pl_refs):
    """Per-studio forecast formula generator."""
    # Lag formulas reference the studio's own sales row
    earned_lags = []
    breakage_lags = []
    a_earned_row = a_refs["earned_row"]
    a_breakage_row = a_refs["breakage_row"]
    for lag in range(7):
        src_idx = month_idx - lag
        if src_idx < 0:
            continue
        src_col = get_column_letter(4 + src_idx)
        lag_col = get_column_letter(3 + lag)
        earned_lags.append(
            f"'Sales Forecast'!{src_col}{sales_row}*Assumptions!${lag_col}${a_earned_row}"
        )
        breakage_lags.append(
            f"'Sales Forecast'!{src_col}{sales_row}*Assumptions!${lag_col}${a_breakage_row}"
        )
    earned_formula = "+".join(earned_lags) if earned_lags else "0"
    breakage_formula = "+".join(breakage_lags) if breakage_lags else "0"

    if display == "Session Revenue":
        return f"={earned_formula}"
    if display == "Breakage Revenue":
        return f"={breakage_formula}"
    if display == "Retail Sales":
        return f"={retail_avg}"
    if display == "Refunds":
        return (f"=({earned_formula}+{breakage_formula}+{retail_avg})*"
                f"Assumptions!$C${a_refs['refund_row']}")
    if display == "Discounts":
        return (f"=({earned_formula}+{breakage_formula}+{retail_avg})*"
                f"Assumptions!$C${a_refs['discount_row']}")
    if display == "Total Revenue":
        rows = pl_refs["rows"]
        return (f"={col}{rows['Session Revenue']}+{col}{rows['Breakage Revenue']}"
                f"+{col}{rows['Retail Sales']}+{col}{rows['Refunds']}+{col}{rows['Discounts']}")
    if display == "COGS & Merchant Fees":
        rows = pl_refs["rows"]
        return (f"={col}{rows['Total Revenue']}*(Assumptions!$C${a_refs['merchant_row']}"
                f"+Assumptions!$C${a_refs['cogs_row']})")
    if display == "Gross Profit":
        rows = pl_refs["rows"]
        return f"={col}{rows['Total Revenue']}-{col}{rows['COGS & Merchant Fees']}"
    # OpEx lines: pull from studio_opex JSON for the right category
    if display in opex_map:
        ovr_cat = opex_map[display]
        if ovr_cat:
            mk = MONTH_KEYS[month_idx]
            val = studio_opex.get(ovr_cat, {})
            if isinstance(val, dict):
                v = val.get(mk, 0)
                try:
                    return f"={float(v)}"
                except (ValueError, TypeError):
                    return "0"
        if display == "Software":
            return "0"
    if display == "Total Operating Expenses":
        rows = pl_refs["rows"]
        return (f"={col}{rows['Marketing']}+{col}{rows['Payroll']}+{col}{rows['Software']}"
                f"+{col}{rows['Professional Fees']}+{col}{rows['Utilities']}+{col}{rows['Property Costs']}")
    if display == "Net Operating Income":
        rows = pl_refs["rows"]
        return f"={col}{rows['Gross Profit']}-{col}{rows['Total Operating Expenses']}"
    if display in ("Depreciation", "Interest Expense", "Taxes Paid", "Property Taxes"):
        return "0"  # Not allocated per studio
    if display == "Total Other Expenses":
        rows = pl_refs["rows"]
        return (f"={col}{rows['Depreciation']}+{col}{rows['Interest Expense']}"
                f"+{col}{rows['Taxes Paid']}+{col}{rows['Property Taxes']}")
    if display == "Net Income":
        rows = pl_refs["rows"]
        return f"={col}{rows['Net Operating Income']}-{col}{rows['Total Other Expenses']}"
    return None


def build_all_studios_summary(wb, data):
    """Side-by-side studio comparison: key metrics per studio."""
    ws = wb.create_sheet("All Studios Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D5"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    for i in range(36):
        ws.column_dimensions[get_column_letter(4 + i)].width = 12

    ws["B2"] = "All Studios — Summary"
    ws["B2"].font = title_font

    HEADER_ROW = 4
    label_cell(ws, HEADER_ROW, 2, "Studio / Metric", bold=True)
    ws.cell(row=HEADER_ROW, column=2).fill = header_fill
    ws.cell(row=HEADER_ROW, column=2).font = header_font
    for i, lbl in enumerate(MONTH_LABELS):
        cell = ws.cell(row=HEADER_ROW, column=4 + i, value=lbl)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    row = HEADER_ROW + 1

    # For each studio: Net Revenue row + EBITDA row
    studio_refs = wb._mighty_studio_refs
    for code, name in STUDIOS:
        # Section header
        section_bar(ws, row, 2, 39, f"{name} ({code})")
        row += 1
        rows_studio = studio_refs[code]["rows"]
        for metric in ["Total Revenue", "Gross Profit", "Total Operating Expenses",
                       "Net Operating Income"]:
            label_cell(ws, row, 2, metric, bold=(metric == "Net Operating Income"))
            for i in range(36):
                col = get_column_letter(4 + i)
                formula_cell(ws, row, 4 + i,
                             f"='{code} P&L'!{col}{rows_studio[metric]}",
                             fmt=CURR,
                             bold=(metric == "Net Operating Income"))
            row += 1
        row += 1


# ============================================================
# SUMMARY & CONTROLS (client's required top tab)
# ============================================================

def build_summary_controls(wb, data):
    """
    Top-level dashboard. Mimics client's existing format from
    2026 Mighty Financial Model - 4.21.26-CC-mar.xlsx but without
    the "New Studio Controls" section.
    """
    ws = wb.create_sheet("Summary & Controls")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 22
    for i in range(36):
        ws.column_dimensions[get_column_letter(4 + i)].width = 11

    ws["B2"] = "Summary & Controls"
    ws["B2"].font = title_font
    ws["B3"] = "Top-level dashboard. Studio sales, EBITDA, YoY summary, charts."
    ws["B3"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    sf_refs = wb._mighty_refs["sales_forecast"]
    studio_refs = wb._mighty_studio_refs

    HEADER_ROW = 5
    section_bar(ws, HEADER_ROW - 1, 2, 39, "Monthly Total Session Sales by Studio")
    # Month headers
    label_cell(ws, HEADER_ROW, 2, "Metric", bold=True)
    label_cell(ws, HEADER_ROW, 3, "Studio", bold=True)
    for c in (2, 3):
        ws.cell(row=HEADER_ROW, column=c).fill = header_fill
        ws.cell(row=HEADER_ROW, column=c).font = header_font
    for i, lbl in enumerate(MONTH_LABELS):
        cell = ws.cell(row=HEADER_ROW, column=4 + i, value=lbl)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    row = HEADER_ROW + 1

    # Total Session Sales row
    label_cell(ws, row, 2, "Total Session Sales", bold=True)
    for i in range(36):
        col = get_column_letter(4 + i)
        formula_cell(ws, row, 4 + i,
                     f"='Sales Forecast'!{col}{sf_refs['total_row']}",
                     fmt=CURR, bold=True, fill=light_fill)
    row += 1
    # Per-studio rows
    for code, name in STUDIOS:
        s_row = sf_refs["studios"].get(code)
        label_cell(ws, row, 3, name)
        for i in range(36):
            col = get_column_letter(4 + i)
            formula_cell(ws, row, 4 + i,
                         f"='Sales Forecast'!{col}{s_row}",
                         fmt=CURR)
        row += 1
    row += 1

    # EBITDA section
    section_bar(ws, row, 2, 39, "Monthly EBITDA by Studio")
    row += 1
    # Total EBITDA = sum of per-studio NOI
    label_cell(ws, row, 2, "Total EBITDA", bold=True)
    total_ebitda_row = row
    # Will fill formulas after individual rows so we can SUM
    row += 1
    studio_ebitda_rows = {}
    for code, name in STUDIOS:
        label_cell(ws, row, 3, name)
        rows_studio = studio_refs[code]["rows"]
        noi_row = rows_studio["Net Operating Income"]
        for i in range(36):
            col = get_column_letter(4 + i)
            formula_cell(ws, row, 4 + i,
                         f"='{code} P&L'!{col}{noi_row}",
                         fmt=CURR)
        studio_ebitda_rows[code] = row
        row += 1
    # Fill in total EBITDA = sum of studio rows
    for i in range(36):
        col = get_column_letter(4 + i)
        first_row = min(studio_ebitda_rows.values())
        last_row = max(studio_ebitda_rows.values())
        formula_cell(ws, total_ebitda_row, 4 + i,
                     f"=SUM({col}{first_row}:{col}{last_row})",
                     fmt=CURR, bold=True, fill=light_fill)
    row += 1

    # Annual Summary table
    section_bar(ws, row, 2, 12, "Annual Summary")
    row += 1
    header_row(ws, row, 2, ["Metric", "Studio", "2026", "2027", "2028",
                             "2027 YoY", "2028 YoY"])
    row += 1
    label_cell(ws, row, 2, "Total Session Sales", bold=True)
    formula_cell(ws, row, 4, f"=SUM('Sales Forecast'!D{sf_refs['total_row']}:O{sf_refs['total_row']})",
                 fmt=CURR, bold=True, fill=light_fill)
    formula_cell(ws, row, 5, f"=SUM('Sales Forecast'!P{sf_refs['total_row']}:AA{sf_refs['total_row']})",
                 fmt=CURR, bold=True, fill=light_fill)
    formula_cell(ws, row, 6, f"=SUM('Sales Forecast'!AB{sf_refs['total_row']}:AM{sf_refs['total_row']})",
                 fmt=CURR, bold=True, fill=light_fill)
    formula_cell(ws, row, 7, f"=IFERROR((E{row}-D{row})/D{row},0)", fmt=PCT, bold=True)
    formula_cell(ws, row, 8, f"=IFERROR((F{row}-E{row})/E{row},0)", fmt=PCT, bold=True)
    row += 1
    for code, name in STUDIOS:
        s_row = sf_refs["studios"].get(code)
        label_cell(ws, row, 3, name)
        formula_cell(ws, row, 4, f"=SUM('Sales Forecast'!D{s_row}:O{s_row})", fmt=CURR)
        formula_cell(ws, row, 5, f"=SUM('Sales Forecast'!P{s_row}:AA{s_row})", fmt=CURR)
        formula_cell(ws, row, 6, f"=SUM('Sales Forecast'!AB{s_row}:AM{s_row})", fmt=CURR)
        formula_cell(ws, row, 7, f"=IFERROR((E{row}-D{row})/D{row},0)", fmt=PCT)
        formula_cell(ws, row, 8, f"=IFERROR((F{row}-E{row})/E{row},0)", fmt=PCT)
        row += 1

    row += 1
    # EBITDA annual summary
    label_cell(ws, row, 2, "Total EBITDA", bold=True)
    # For 2026: cols D..O of EBITDA per-studio rows. The total EBITDA row spans cols 4..39.
    # 2026 = D..O = cols 4..15 in total_ebitda_row
    ws.cell(row=row, column=4,
            value=f"=SUM(D{total_ebitda_row}:O{total_ebitda_row})").number_format = CURR
    ws.cell(row=row, column=5,
            value=f"=SUM(P{total_ebitda_row}:AA{total_ebitda_row})").number_format = CURR
    ws.cell(row=row, column=6,
            value=f"=SUM(AB{total_ebitda_row}:AM{total_ebitda_row})").number_format = CURR
    ws.cell(row=row, column=4).font = data_bold
    ws.cell(row=row, column=5).font = data_bold
    ws.cell(row=row, column=6).font = data_bold
    for c in (4, 5, 6):
        ws.cell(row=row, column=c).fill = light_fill
    formula_cell(ws, row, 7, f"=IFERROR((E{row}-D{row})/D{row},0)", fmt=PCT, bold=True)
    formula_cell(ws, row, 8, f"=IFERROR((F{row}-E{row})/E{row},0)", fmt=PCT, bold=True)
    row += 1
    for code, name in STUDIOS:
        s_row = studio_ebitda_rows[code]
        label_cell(ws, row, 3, name)
        formula_cell(ws, row, 4, f"=SUM(D{s_row}:O{s_row})", fmt=CURR)
        formula_cell(ws, row, 5, f"=SUM(P{s_row}:AA{s_row})", fmt=CURR)
        formula_cell(ws, row, 6, f"=SUM(AB{s_row}:AM{s_row})", fmt=CURR)
        formula_cell(ws, row, 7, f"=IFERROR((E{row}-D{row})/D{row},0)", fmt=PCT)
        formula_cell(ws, row, 8, f"=IFERROR((F{row}-E{row})/E{row},0)", fmt=PCT)
        row += 1

    # ----- Charts -----
    from openpyxl.chart import BarChart, LineChart, Reference

    # Revenue chart (total session sales by month)
    chart = BarChart()
    chart.title = "Monthly Total Session Sales"
    chart.style = 11
    chart.height = 8
    chart.width = 24
    # Total Session Sales is at row 6 (HEADER_ROW + 1)
    data_ref = Reference(ws, min_col=4, max_col=39, min_row=HEADER_ROW + 1, max_row=HEADER_ROW + 1)
    cats = Reference(ws, min_col=4, max_col=39, min_row=HEADER_ROW, max_row=HEADER_ROW)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, f"B{row + 2}")

    # EBITDA chart (total EBITDA by month)
    chart2 = LineChart()
    chart2.title = "Monthly Total EBITDA"
    chart2.style = 12
    chart2.height = 8
    chart2.width = 24
    data_ref2 = Reference(ws, min_col=4, max_col=39, min_row=total_ebitda_row, max_row=total_ebitda_row)
    chart2.add_data(data_ref2, titles_from_data=False)
    chart2.set_categories(cats)
    ws.add_chart(chart2, f"B{row + 22}")


# ============================================================
# MAIN
# ============================================================

def main():
    print("Loading dashboard data...")
    data = load_dashboard_data()
    print(f"  - committed_actuals: {len(data['committed'])} top-level keys")
    print(f"  - baseline: {len(data['baseline'])} top-level keys")
    print(f"  - merged: ready")

    print("Building workbook...")
    wb = openpyxl.Workbook()
    build_cover(wb, data)
    build_assumptions(wb, data)
    print("  - Assumptions tab built")
    build_sales_forecast(wb, data)
    print("  - Sales Forecast tab built")
    build_cash_flow_forecast(wb, data)
    print("  - Cash Flow Forecast tab built")
    build_pl_consolidated(wb, data)
    print("  - P&L tab built")
    # Per-studio P&L tabs (must be built before All Studios + Summary which reference them)
    for code, name in STUDIOS:
        build_studio_pl(wb, data, code, name)
    print(f"  - Per-studio P&L tabs built ({len(STUDIOS)} studios)")
    build_all_studios_summary(wb, data)
    print("  - All Studios Summary tab built")
    build_cash_debt_equity(wb, data)
    print("  - Cash, Debt & Equity tab built")
    build_capex(wb, data)
    print("  - CapEx tab built")
    build_qbo_actuals(wb, data)
    print("  - QBO Actuals tab built")
    # Build last (references everything above)
    build_summary_controls(wb, data)
    print("  - Summary & Controls tab built")

    # Reorder: Summary & Controls should be first (after Cover)
    cover = wb["Cover"]
    summary = wb["Summary & Controls"]
    wb.move_sheet(summary, offset=-(len(wb.sheetnames) - 2))

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
