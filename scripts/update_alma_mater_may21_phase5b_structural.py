"""
Phase 5b — Structural 2028 Extension (May 21, 2026)
======================================================

Applied to the CORRECT file at /Users/chandlerclemons/financial-modeling/
models/Alma Mater Financial Model.xlsx

Includes:
  1. Wholesale 2028 — 5 channel rows at R222-R226 (Green Grass doubled from
     2027, all other channels = 0 placeholder). Channel × month layout.
  2. POs 2028 — 3 Beta POs sized to cover 8K WS + 5.5K DTC = 13.5K + buffer.
     Uses R186-R188 placeholders. Alpha POs stay at 0 (no Alpha demand).
  3. Inventory tab — extend cols R-AL (Apr'27 → Dec'28). WS formula uses
     SUM(channel × month) pattern matching the correct file's layout.
  4. Monthly P&L — 2028 forecast section at R114-R134.
  5. Cash Flow — extend cols R-AL (preserves IF(QBO/Forecast) pattern on
     revenue rows).
  6. Dashboard — 2028 Forecast column F + Ending Cash Dec 2028 row R32.

Wholesale assumption rows used by Inventory R9 SUM:
  2026 → R88-R92, 2027 → R93-R97, 2028 → R222-R226
Monthly AOV refs (Inventory R32):
  2026 → R69, 2027 → R80, 2028 → R200
DTC Demand refs (Inventory R11/R21):
  2026 → R66/R67, 2027 → R77/R78, 2028 → R196/R197
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/financial-modeling/models/Alma Mater Financial Model.xlsx"
)

YELLOW = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")
BOLD = Font(bold=True)
MONTHS_SHORT = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# 2027 Green Grass: [0, 250, 1000, 250, 0, 0, 500, 1500, 500, 0, 0, 0] = 4,000u
# 2028 Green Grass = 2× 2027
GG_2028_MONTHLY = [0, 500, 2000, 500, 0, 0, 1000, 3000, 1000, 0, 0, 0]  # 8,000 units total

# 2028 POs — sized for 2028 Beta demand: 5.5K DTC + 8K WS = 13.5K + ~1K buffer
NEW_2028_POS = [
    # (row, name, product, pairs, amount, order_month, order_year)
    (186, "Spring 2028 (Beta)", "Beta", 5000, 225000, 11, 2027),
    (187, "Summer 2028 (Beta)", "Beta", 5000, 225000,  2, 2028),
    (188, "Fall 2028 (Beta)",   "Beta", 5000, 225000,  5, 2028),
    # Alpha POs (R189-R191) left at 0 / (Add PO) — no Alpha demand in correct file's all-Beta model
]


# ============================================================
# 1. WHOLESALE 2028 — append at R222-R226 (channel × month)
# ============================================================
WHOLESALE_2028 = [
    ("Big Box Retail",      [0]*12,           "TBD - placeholder"),
    ("Green Grass (CC)",    GG_2028_MONTHLY,  "2× 2027 GG placeholder (8,000u)"),
    ("Other Wholesale",     [0]*12,           "Placeholder"),
    ("International - 1",   [0]*12,           "TBD"),
    ("International - 2",   [0]*12,           "TBD"),
]
WS_2028_START_ROW = 222  # R222-R226


def apply_wholesale_2028(asm):
    """Append 5 channel rows for 2028 at R222-R226."""
    # Header row first
    asm.cell(row=WS_2028_START_ROW - 2, column=2, value="WHOLESALE FORECAST - 2028 (PLACEHOLDER)").font = BOLD
    # Use same column structure as R87 (channel × month + total + ASP + rev + cogs + notes)
    asm.cell(row=WS_2028_START_ROW - 1, column=2, value="Channel").font = BOLD
    asm.cell(row=WS_2028_START_ROW - 1, column=3, value="Year").font = BOLD
    for i, m in enumerate(MONTHS_SHORT):
        asm.cell(row=WS_2028_START_ROW - 1, column=4 + i, value=m).font = BOLD
    asm.cell(row=WS_2028_START_ROW - 1, column=16, value="Total Units").font = BOLD
    asm.cell(row=WS_2028_START_ROW - 1, column=17, value="ASP").font = BOLD
    asm.cell(row=WS_2028_START_ROW - 1, column=18, value="Annual Rev").font = BOLD
    asm.cell(row=WS_2028_START_ROW - 1, column=19, value="Annual COGS").font = BOLD
    asm.cell(row=WS_2028_START_ROW - 1, column=20, value="Notes").font = BOLD

    for i, (channel, monthly, notes) in enumerate(WHOLESALE_2028):
        row = WS_2028_START_ROW + i
        asm.cell(row=row, column=2, value=channel)
        asm.cell(row=row, column=3, value=2028)
        for m, val in enumerate(monthly):
            c = asm.cell(row=row, column=4 + m, value=val)
            c.fill = YELLOW
            c.number_format = '#,##0'
        asm.cell(row=row, column=16, value=f"=SUM(D{row}:O{row})")
        c = asm.cell(row=row, column=17, value=144)  # ASP
        c.fill = YELLOW
        c.number_format = '$#,##0'
        asm.cell(row=row, column=18, value=f"=P{row}*Q{row}")
        asm.cell(row=row, column=18).number_format = '$#,##0'
        asm.cell(row=row, column=19, value=f"=P{row}*$C$33")
        asm.cell(row=row, column=19).number_format = '$#,##0'
        asm.cell(row=row, column=20, value=notes)


# ============================================================
# 2. POs 2028 — fill R186-R188 (Beta), R189-R191 stay as placeholders
# ============================================================
def apply_pos_2028(asm):
    for row, name, product, pairs, amount, order_month, order_year in NEW_2028_POS:
        asm.cell(row=row, column=2, value=name)
        asm.cell(row=row, column=3, value=product).fill = YELLOW
        asm.cell(row=row, column=4, value=pairs).fill = YELLOW
        asm.cell(row=row, column=5, value=amount).fill = YELLOW
        asm.cell(row=row, column=6, value=order_month).fill = YELLOW
        asm.cell(row=row, column=7, value=order_year).fill = YELLOW


# ============================================================
# 3. INVENTORY EXTENSION — cols R-AL (Apr'27 → Dec'28)
# ============================================================
FIRST_NEW_INV_COL = 18  # R (Apr'27)
LAST_NEW_INV_COL = 38   # AL (Dec'28)


def get_dtc_demand_ref(inv_col, product):
    """Asm refs: Beta R66/R77/R196, Alpha R67/R78/R197."""
    if 15 <= inv_col <= 26:
        row = 77 if product == 'Beta' else 78
        asm_col = inv_col - 12
    elif 27 <= inv_col <= 38:
        row = 196 if product == 'Beta' else 197
        asm_col = inv_col - 24
    else:
        row = 66 if product == 'Beta' else 67
        asm_col = inv_col
    return f"Assumptions!{get_column_letter(asm_col)}{row}"


def get_ws_shipments_beta_formula(inv_col):
    """WS Beta = SUM of (channel × month) in wholesale block for that year."""
    if 3 <= inv_col <= 14:    # 2026: R88-R92
        asm_col = get_column_letter(inv_col + 1)  # Inv col C (3) → Asm col D (4) = Jan units
        return f"=SUM(Assumptions!${asm_col}$88:${asm_col}$92)"
    elif 15 <= inv_col <= 26:  # 2027: R93-R97
        asm_col = get_column_letter(inv_col - 11)  # Inv col O (15) → Asm col D (4) = Jan 2027
        return f"=SUM(Assumptions!${asm_col}$93:${asm_col}$97)"
    elif 27 <= inv_col <= 38:  # 2028: R222-R226
        asm_col = get_column_letter(inv_col - 23)  # Inv col AA (27) → Asm col D (4) = Jan 2028
        return f"=SUM(Assumptions!${asm_col}$222:${asm_col}$226)"


def get_aov_ref(inv_col):
    if 15 <= inv_col <= 26:
        return f"Assumptions!{get_column_letter(inv_col - 12)}$80"
    elif 27 <= inv_col <= 38:
        return f"Assumptions!{get_column_letter(inv_col - 24)}$200"
    else:
        return f"Assumptions!{get_column_letter(inv_col)}$69"


def get_discount_return_refs(inv_col):
    if inv_col <= 14:
        return "Assumptions!$C$12", "Assumptions!$C$13"
    else:
        return "Assumptions!$C$14", "Assumptions!$C$15"


def apply_inventory_extension(wb):
    inv = wb["Inventory"]
    # Extend month headers
    for header_row in [6, 16, 26, 31, 36]:
        for c in range(FIRST_NEW_INV_COL, LAST_NEW_INV_COL + 1):
            month_idx = (c - 3) % 12
            year = 2026 + (c - 3) // 12
            year_suffix = str(year)[-2:]
            inv.cell(row=header_row, column=c, value=f"{MONTHS_SHORT[month_idx]}'{year_suffix}")

    # Beta block R7-R13
    for c in range(FIRST_NEW_INV_COL, LAST_NEW_INV_COL + 1):
        col = get_column_letter(c)
        prev_col = get_column_letter(c - 1)
        inv.cell(row=7, column=c, value=f"={prev_col}13")
        inv.cell(row=8, column=c,
                 value='=SUMPRODUCT((Assumptions!$C$172:$C$191="Beta")*'
                       '((Assumptions!$G$172:$G$191-2026)*12+Assumptions!$F$172:$F$191+Assumptions!$C$166=COLUMN()-2)*'
                       'Assumptions!$D$172:$D$191)')
        inv.cell(row=9, column=c, value=get_ws_shipments_beta_formula(c))
        inv.cell(row=10, column=c, value=f"={col}7+{col}8-{col}9")
        inv.cell(row=11, column=c, value=f"={get_dtc_demand_ref(c, 'Beta')}")
        inv.cell(row=12, column=c, value=f"=MIN({col}11,MAX({col}10,0))")
        inv.cell(row=13, column=c, value=f"={col}10-{col}12")

    # Alpha block R17-R23 (WS Alpha shipments = 0 — all-Beta wholesale model)
    for c in range(FIRST_NEW_INV_COL, LAST_NEW_INV_COL + 1):
        col = get_column_letter(c)
        prev_col = get_column_letter(c - 1)
        inv.cell(row=17, column=c, value=f"={prev_col}23")
        inv.cell(row=18, column=c,
                 value='=SUMPRODUCT((Assumptions!$C$172:$C$191="Alpha")*'
                       '((Assumptions!$G$172:$G$191-2026)*12+Assumptions!$F$172:$F$191+Assumptions!$C$166=COLUMN()-2)*'
                       'Assumptions!$D$172:$D$191)')
        inv.cell(row=19, column=c, value=0)  # WS Alpha shipments = 0
        inv.cell(row=20, column=c, value=f"={col}17+{col}18-{col}19")
        inv.cell(row=21, column=c, value=f"={get_dtc_demand_ref(c, 'Alpha')}")
        inv.cell(row=22, column=c, value=f"=MIN({col}21,MAX({col}20,0))")
        inv.cell(row=23, column=c, value=f"={col}20-{col}22")

    # Totals R27, R28
    for c in range(FIRST_NEW_INV_COL, LAST_NEW_INV_COL + 1):
        col = get_column_letter(c)
        inv.cell(row=27, column=c, value=f"={col}13+{col}23")
        inv.cell(row=28, column=c, value=f"={col}12+{col}22")

    # Constrained DTC Revenue R32 (Gross), R33 (Net)
    for c in range(FIRST_NEW_INV_COL, LAST_NEW_INV_COL + 1):
        col = get_column_letter(c)
        inv.cell(row=32, column=c, value=f"=({col}12+{col}22)*{get_aov_ref(c)}")
        disc, ret = get_discount_return_refs(c)
        inv.cell(row=33, column=c, value=f"={col}32*(1-{disc})*(1-{ret})")

    # PO Payment Schedule R37-R56
    for po_row in range(37, 57):
        asm_row = 172 + (po_row - 37)
        for c in range(FIRST_NEW_INV_COL, LAST_NEW_INV_COL + 1):
            inv.cell(row=po_row, column=c,
                     value=f'=IF(AND(Assumptions!$E${asm_row}>0,'
                           f'COLUMN()-2=(Assumptions!$G${asm_row}-2026)*12+Assumptions!$F${asm_row}+'
                           f'Assumptions!$C$166+Assumptions!$C$167),'
                           f'Assumptions!$E${asm_row},0)')

    # Total Inventory Payments R57
    for c in range(FIRST_NEW_INV_COL, LAST_NEW_INV_COL + 1):
        col = get_column_letter(c)
        inv.cell(row=57, column=c, value=f"=SUM({col}37:{col}56)")


# ============================================================
# 4. MONTHLY P&L — 2028 forecast section at R114-R134
# ============================================================
def apply_monthly_pl_2028(wb):
    ws = wb['Monthly P&L']
    ws.cell(row=114, column=2, value="2028 FORECAST").font = BOLD
    for i, m in enumerate(MONTHS_SHORT):
        ws.cell(row=116, column=3 + i, value=m).font = BOLD
    ws.cell(row=116, column=15, value="FY 2028").font = BOLD

    ws.cell(row=117, column=2, value="REVENUE").font = BOLD

    # R118: DTC Revenue (Net) ← Inventory!AA33:AL33
    ws.cell(row=118, column=2, value="DTC Revenue (Net)")
    for i in range(12):
        inv_col = get_column_letter(27 + i)
        ws.cell(row=118, column=3 + i, value=f"=Inventory!{inv_col}33")
    ws.cell(row=118, column=15, value="=SUM(C118:N118)")

    # R119: Wholesale Revenue — uses SUMPRODUCT on 2028 wholesale rows R222-R226
    ws.cell(row=119, column=2, value="Wholesale Revenue")
    for i in range(12):
        asm_col = get_column_letter(4 + i)  # D-O = Jan-Dec
        ws.cell(row=119, column=3 + i,
                value=f"=SUMPRODUCT(Assumptions!${asm_col}$222:${asm_col}$226,Assumptions!$Q$222:$Q$226)")
    ws.cell(row=119, column=15, value="=SUM(C119:N119)")

    # R120: Total Revenue
    ws.cell(row=120, column=2, value="Total Revenue").font = BOLD
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=120, column=3 + i, value=f"={col}118+{col}119")
    ws.cell(row=120, column=15, value="=SUM(C120:N120)")

    ws.cell(row=122, column=2, value="COST OF GOODS SOLD").font = BOLD

    # R123: DTC COGS ← Inventory!AA32:AL32 × Assumptions!$C$24
    ws.cell(row=123, column=2, value="DTC COGS")
    for i in range(12):
        inv_col = get_column_letter(27 + i)
        ws.cell(row=123, column=3 + i, value=f"=Inventory!{inv_col}32*Assumptions!$C$24")
    ws.cell(row=123, column=15, value="=SUM(C123:N123)")

    # R124: Wholesale COGS — sum of units × C33 unit cost
    ws.cell(row=124, column=2, value="Wholesale COGS")
    for i in range(12):
        asm_col = get_column_letter(4 + i)
        ws.cell(row=124, column=3 + i,
                value=f"=SUM(Assumptions!${asm_col}$222:${asm_col}$226)*Assumptions!$C$33")
    ws.cell(row=124, column=15, value="=SUM(C124:N124)")

    # R125: Total COGS
    ws.cell(row=125, column=2, value="Total COGS").font = BOLD
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=125, column=3 + i, value=f"={col}123+{col}124")
    ws.cell(row=125, column=15, value="=SUM(C125:N125)")

    # R126: Gross Profit
    ws.cell(row=126, column=2, value="Gross Profit").font = BOLD
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=126, column=3 + i, value=f"={col}120-{col}125")
    ws.cell(row=126, column=15, value="=SUM(C126:N126)")

    # R127: Gross Margin %
    ws.cell(row=127, column=2, value="Gross Margin %")
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=127, column=3 + i, value=f"=IF({col}120>0,{col}126/{col}120,0)")
        ws.cell(row=127, column=3 + i).number_format = '0.0%'
    ws.cell(row=127, column=15, value="=IF(O120>0,O126/O120,0)")
    ws.cell(row=127, column=15).number_format = '0.0%'

    ws.cell(row=129, column=2, value="OPERATING EXPENSES").font = BOLD

    # R130: Team Costs — flat across years (same as 2026/2027 model behavior)
    ws.cell(row=130, column=2, value="Team Costs (Fully Burdened)")
    for i in range(12):
        tc_col = get_column_letter(7 + i)
        ws.cell(row=130, column=3 + i, value=f"='Team Costs'!{tc_col}16")
    ws.cell(row=130, column=15, value="=SUM(C130:N130)")

    # R131: Other OpEx ← Assumptions!C220:N220 (new 2028 TOTAL from Phase 5 redo)
    ws.cell(row=131, column=2, value="Other OpEx")
    for i in range(12):
        asm_col = get_column_letter(3 + i)
        ws.cell(row=131, column=3 + i, value=f"=Assumptions!{asm_col}220")
    ws.cell(row=131, column=15, value="=SUM(C131:N131)")

    # R132: Total OpEx
    ws.cell(row=132, column=2, value="Total OpEx").font = BOLD
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=132, column=3 + i, value=f"={col}130+{col}131")
    ws.cell(row=132, column=15, value="=SUM(C132:N132)")

    # R133: EBITDA
    ws.cell(row=133, column=2, value="EBITDA").font = BOLD
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=133, column=3 + i, value=f"={col}126-{col}132")
    ws.cell(row=133, column=15, value="=SUM(C133:N133)")

    # R134: EBITDA Margin %
    ws.cell(row=134, column=2, value="EBITDA Margin %")
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=134, column=3 + i, value=f"=IF({col}120>0,{col}133/{col}120,0)")
        ws.cell(row=134, column=3 + i).number_format = '0.0%'
    ws.cell(row=134, column=15, value="=IF(O120>0,O133/O120,0)")
    ws.cell(row=134, column=15).number_format = '0.0%'


# ============================================================
# 5. CASH FLOW EXTENSION — cols R-AL (Apr'27 → Dec'28)
# ============================================================
def cf_col_to_pl_ref_with_qbo(cf_col, pl_row_2026, pl_row_2027, pl_row_2028):
    """Cash Flow refs P&L. For 2026 cols, uses IF(QBO/Forecast) pattern.
    For 2027/2028 cols, just pulls forecast directly."""
    if 18 <= cf_col <= 26:
        pl_col = get_column_letter(cf_col - 12)
        return f"='Monthly P&L'!{pl_col}{pl_row_2027}"
    elif 27 <= cf_col <= 38:
        pl_col = get_column_letter(cf_col - 24)
        return f"='Monthly P&L'!{pl_col}{pl_row_2028}"


def apply_cash_flow_extension(wb):
    cf = wb['Cash Flow']

    # Extend month headers R6
    for c in range(FIRST_NEW_INV_COL, LAST_NEW_INV_COL + 1):
        month_idx = (c - 3) % 12
        year = 2026 + (c - 3) // 12
        year_suffix = str(year)[-2:]
        cf.cell(row=6, column=c, value=f"{MONTHS_SHORT[month_idx]}'{year_suffix}")

    for c in range(FIRST_NEW_INV_COL, LAST_NEW_INV_COL + 1):
        col = get_column_letter(c)
        prev_col = get_column_letter(c - 1)
        month_num = (c - 3) % 12 + 1
        year = 2026 + (c - 3) // 12

        # R8 DTC Revenue ← P&L 2027 row 95, 2028 row 118
        cf.cell(row=8, column=c, value=cf_col_to_pl_ref_with_qbo(c, 43, 95, 118))

        # R9 Wholesale Revenue ← P&L 2027 row 96, 2028 row 119
        cf.cell(row=9, column=c, value=cf_col_to_pl_ref_with_qbo(c, 44, 96, 119))

        # R10 Funding
        cf.cell(row=10, column=c,
                value=f"=SUMPRODUCT((Assumptions!$D$58:$D$60={month_num})*"
                      f"(Assumptions!$E$58:$E$60={year})*Assumptions!$C$58:$C$60)")

        # R11 Total Cash In
        cf.cell(row=11, column=c, value=f"={col}8+{col}9+{col}10")

        # R14 Inventory Purchases ← Inventory!col57
        cf.cell(row=14, column=c, value=f"=Inventory!{col}57")

        # R15 Fulfillment COGS ← Inventory col32 × (cogs_rate - product_cost_rate)
        cf.cell(row=15, column=c,
                value=f"=Inventory!{col}32*(Assumptions!$C$24-Assumptions!$C$20)")

        # R16 Team Costs ← P&L 2027 row 107, 2028 row 130
        cf.cell(row=16, column=c, value=cf_col_to_pl_ref_with_qbo(c, 55, 107, 130))

        # R17 Other OpEx ← P&L 2027 row 108, 2028 row 131
        cf.cell(row=17, column=c, value=cf_col_to_pl_ref_with_qbo(c, 56, 108, 131))

        # R18 Total Cash Out
        cf.cell(row=18, column=c, value=f"={col}14+{col}15+{col}16+{col}17")

        # R21 Monthly Net Cash Flow
        cf.cell(row=21, column=c, value=f"={col}11-{col}18")

        # R23 Beginning Cash
        cf.cell(row=23, column=c, value=f"={prev_col}24")

        # R24 Ending Cash Balance
        cf.cell(row=24, column=c, value=f"={col}23+{col}21")

        # R26 Days of Cash
        cf.cell(row=26, column=c, value=f"=IF({col}18>0,{col}24/({col}18/30),0)")

        # R31 Cash Balance (No NEW Funding)
        cf.cell(row=31, column=c, value=f"={prev_col}31+{col}8+{col}9-{col}18")

    cf.cell(row=4, column=2, value="36-MONTH CASH FLOW").font = BOLD


# ============================================================
# 6. DASHBOARD — 2028 col F + R32 Ending Cash Dec 2028
# ============================================================
def apply_dashboard_2028(wb):
    dash = wb['Dashboard']
    dash.cell(row=7, column=6, value="2028 Forecast").font = BOLD
    dash.cell(row=8, column=6, value="='Monthly P&L'!O120")
    dash.cell(row=9, column=6, value="='Monthly P&L'!O125")
    dash.cell(row=10, column=6, value="='Monthly P&L'!O126")
    dash.cell(row=11, column=6, value="='Monthly P&L'!O127")
    dash.cell(row=11, column=6).number_format = '0.0%'
    dash.cell(row=12, column=6, value="='Monthly P&L'!O132")
    dash.cell(row=13, column=6, value="='Monthly P&L'!O133")
    dash.cell(row=14, column=6, value="='Monthly P&L'!O134")
    dash.cell(row=14, column=6).number_format = '0.0%'

    dash.cell(row=31, column=2, value="Ending Cash - Dec 2027")
    dash.cell(row=31, column=3, value="='Cash Flow'!Z24")
    dash.cell(row=32, column=2, value="Ending Cash - Dec 2028")
    dash.cell(row=32, column=3, value="='Cash Flow'!AL24")


# ============================================================
# Main
# ============================================================
def main():
    wb = openpyxl.load_workbook(MODEL_PATH)
    asm = wb["Assumptions"]

    # Step 1: Wholesale 2028 (must be before Inventory extension that references it)
    apply_wholesale_2028(asm)

    # Step 2: POs 2028
    apply_pos_2028(asm)

    # Step 3: Inventory extension
    apply_inventory_extension(wb)

    # Step 4: Monthly P&L 2028 section
    apply_monthly_pl_2028(wb)

    # Step 5: Cash Flow extension
    apply_cash_flow_extension(wb)

    # Step 6: Dashboard 2028
    apply_dashboard_2028(wb)

    wb.save(MODEL_PATH)
    print(f"✅ Saved: {MODEL_PATH}")
    print(f"\nAll Phase 5b extensions applied:")
    print(f"  Wholesale 2028: R222-R226 (5 channels, GG = 8,000u doubled, others = 0)")
    print(f"  POs 2028: R186-R188 Beta (5K + 5K + 5K = 15K units, $675K)")
    print(f"  Inventory: extended to col AL (Dec'28)")
    print(f"  Monthly P&L: 2028 forecast section R114-R134")
    print(f"  Cash Flow: extended to col AL (Dec'28)")
    print(f"  Dashboard: col F 2028 Forecast + R32 Dec'28 Ending Cash")


if __name__ == "__main__":
    main()
