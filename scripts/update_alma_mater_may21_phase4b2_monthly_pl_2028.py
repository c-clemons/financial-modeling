"""
Phase 4b-2 — Add 2028 P&L forecast section to Monthly P&L tab (May 21, 2026)
=============================================================================

Mirrors the existing 2027 section (R91-R111) at R114-R134.

Pull-throughs:
  DTC Revenue Net  ← Inventory!AA33:AL33  (2028 net DTC revenue)
  DTC COGS         ← Inventory!AA32:AL32 × Assumptions!$C$24 (COGS rate)
  Wholesale Rev    ← SUMPRODUCT over Assumptions!$E$88:$E$93 filtered to year 2028
  Wholesale COGS   ← same pattern, $F$88:$F$93
  Team Costs       ← 'Team Costs'!G16:R16 (same as 2026/2027 — no annual growth in current model)
  Other OpEx       ← Assumptions!C219:N219 (the new 2028 TOTAL row from Phase 4a)

NOTE: Wholesale deals in Assumptions R88:R93 currently end at Fall 2027.
SUMPRODUCT for 2028 will return $0 until 2028 deals are added.
"""

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/"
    "Empirica Financial Model/Alma Mater Financial Model Draft.xlsx"
)

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)
    ws = wb['Monthly P&L']
    bold = Font(bold=True)

    # Inventory col mapping for 2028: Jan=AA(27), Dec=AL(38)
    # Team Costs col mapping: Jan=G(7), Dec=R(18)
    # Assumptions OpEx col mapping for 2028: Jan=C(3), Dec=N(14) — row 219

    # ============================================================
    # R114: 2028 FORECAST header
    # ============================================================
    ws.cell(row=114, column=2, value="2028 FORECAST").font = bold

    # R116: column headers
    for i, m in enumerate(MONTHS):
        ws.cell(row=116, column=3 + i, value=m).font = bold
    ws.cell(row=116, column=15, value="FY 2028").font = bold

    # R117: REVENUE header
    ws.cell(row=117, column=2, value="REVENUE").font = bold

    # R118: DTC Revenue (Net) — Inventory!AA33:AL33
    ws.cell(row=118, column=2, value="DTC Revenue (Net)")
    for i in range(12):
        inv_col = get_column_letter(27 + i)  # AA, AB, ..., AL
        ws.cell(row=118, column=3 + i, value=f"=Inventory!{inv_col}33")
    ws.cell(row=118, column=15, value="=SUM(C118:N118)")

    # R119: Wholesale Revenue
    ws.cell(row=119, column=2, value="Wholesale Revenue")
    for i in range(12):
        month_num = i + 1
        ws.cell(row=119, column=3 + i,
                value=f"=SUMPRODUCT((Assumptions!$J$88:$J$93={month_num})*"
                      f"(Assumptions!$K$88:$K$93=2028)*Assumptions!$E$88:$E$93)")
    ws.cell(row=119, column=15, value="=SUM(C119:N119)")

    # R120: Total Revenue
    ws.cell(row=120, column=2, value="Total Revenue").font = bold
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=120, column=3 + i, value=f"={col}118+{col}119")
    ws.cell(row=120, column=15, value="=SUM(C120:N120)")

    # R122: COGS header
    ws.cell(row=122, column=2, value="COST OF GOODS SOLD").font = bold

    # R123: DTC COGS — Inventory!AA32:AL32 × Assumptions!$C$24
    ws.cell(row=123, column=2, value="DTC COGS")
    for i in range(12):
        inv_col = get_column_letter(27 + i)
        ws.cell(row=123, column=3 + i, value=f"=Inventory!{inv_col}32*Assumptions!$C$24")
    ws.cell(row=123, column=15, value="=SUM(C123:N123)")

    # R124: Wholesale COGS
    ws.cell(row=124, column=2, value="Wholesale COGS")
    for i in range(12):
        month_num = i + 1
        ws.cell(row=124, column=3 + i,
                value=f"=SUMPRODUCT((Assumptions!$J$88:$J$93={month_num})*"
                      f"(Assumptions!$K$88:$K$93=2028)*Assumptions!$F$88:$F$93)")
    ws.cell(row=124, column=15, value="=SUM(C124:N124)")

    # R125: Total COGS
    ws.cell(row=125, column=2, value="Total COGS").font = bold
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=125, column=3 + i, value=f"={col}123+{col}124")
    ws.cell(row=125, column=15, value="=SUM(C125:N125)")

    # R126: Gross Profit
    ws.cell(row=126, column=2, value="Gross Profit").font = bold
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=126, column=3 + i, value=f"={col}120-{col}125")
    ws.cell(row=126, column=15, value="=SUM(C126:N126)")

    # R127: Gross Margin %
    ws.cell(row=127, column=2, value="Gross Margin %")
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=127, column=3 + i, value=f"=IF({col}120>0,{col}126/{col}120,0)")
        ws.cell(row=127, column=3 + i).number_format = '0.0%'
    ws.cell(row=127, column=15, value=f"=IF(O120>0,O126/O120,0)")
    ws.cell(row=127, column=15).number_format = '0.0%'

    # R129: OPERATING EXPENSES header
    ws.cell(row=129, column=2, value="OPERATING EXPENSES").font = bold

    # R130: Team Costs (Fully Burdened) — same as 2026/2027 (Team Costs tab is single-year)
    ws.cell(row=130, column=2, value="Team Costs (Fully Burdened)")
    for i in range(12):
        tc_col = get_column_letter(7 + i)  # G..R
        ws.cell(row=130, column=3 + i, value=f"='Team Costs'!{tc_col}16")
    ws.cell(row=130, column=15, value="=SUM(C130:N130)")

    # R131: Other OpEx — Assumptions!C219:N219 (new 2028 TOTAL row)
    ws.cell(row=131, column=2, value="Other OpEx")
    for i in range(12):
        asm_col = get_column_letter(3 + i)  # C..N
        ws.cell(row=131, column=3 + i, value=f"=Assumptions!{asm_col}219")
    ws.cell(row=131, column=15, value="=SUM(C131:N131)")

    # R132: Total OpEx
    ws.cell(row=132, column=2, value="Total OpEx").font = bold
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=132, column=3 + i, value=f"={col}130+{col}131")
    ws.cell(row=132, column=15, value="=SUM(C132:N132)")

    # R133: EBITDA
    ws.cell(row=133, column=2, value="EBITDA").font = bold
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=133, column=3 + i, value=f"={col}126-{col}132")
    ws.cell(row=133, column=15, value="=SUM(C133:N133)")

    # R134: EBITDA Margin %
    ws.cell(row=134, column=2, value="EBITDA Margin %")
    for i in range(12):
        col = get_column_letter(3 + i)
        ws.cell(row=134, column=3 + i, value=f"=IF({col}120>0,{col}133/{col}120,0)")
        ws.cell(row=134, column=3 + i).number_format = '0.0%'
    ws.cell(row=134, column=15, value=f"=IF(O120>0,O133/O120,0)")
    ws.cell(row=134, column=15).number_format = '0.0%'

    wb.save(MODEL_PATH)
    print(f"✅ Saved Excel: {MODEL_PATH}")
    print(f"\n2028 Forecast section added at R114-R134:")
    print(f"  R118 DTC Revenue Net  ← Inventory AA33:AL33")
    print(f"  R119 Wholesale Rev    ← SUMPRODUCT filtered to year 2028 (currently $0)")
    print(f"  R123 DTC COGS         ← Inventory AA32:AL32 × Assumptions!C24")
    print(f"  R130 Team Costs       ← Team Costs G16:R16 (single-year tab, flat for now)")
    print(f"  R131 Other OpEx       ← Assumptions!C219:N219 (2028 TOTAL)")


if __name__ == "__main__":
    main()
