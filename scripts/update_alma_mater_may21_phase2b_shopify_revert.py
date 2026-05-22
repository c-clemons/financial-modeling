"""
Phase 2b — Revert Shopify to old assumption pending Matt confirmation
======================================================================

Matt's new 2026 Budget has Shopify at $2,500/mo for Mar-Aug only
($15K/yr). Our prior assumption was $2,850/mo flat ($34,200/yr).

Per user direction (May 21): use our old Shopify assumption for now,
confirm with Matt later.

This splits the previous combined "General Systems" line into:
  R125: General Systems — Loop+Yotpo (Matt's $4,581 — keeps Loop Returns
        $340/mo Apr-Dec + Yotpo (general) $169/mo Apr-Dec)
  R126: General Systems — Shopify (OUR assumption $2,850 × 12 = $34,200)

Uses the previously reserved R126 slot. No row shifts. R133 TOTAL
formula picks up automatically.

Net change to 2026 OpEx: +$19,200 (Matt's $15K Shopify → our $34,200)
New 2026 Marketing/Agency OpEx total: $482,669
"""

import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/"
    "Empirica Financial Model/Alma Mater Financial Model Draft.xlsx"
)

# Decomposed General Systems: 2 lines instead of 1
# Both go into R125 and R126
GENERAL_SYSTEMS_LOOP_YOTPO = [0, 0, 0, 509, 509, 509, 509, 509, 509, 509, 509, 509]  # 340 + 169
SHOPIFY_OLD_ASSUMPTION = [2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850]


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)
    asm = wb["Assumptions"]
    yellow = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")

    # R125: replace combined General Systems with Loop+Yotpo only
    asm.cell(row=125, column=2, value="General Systems — Loop+Yotpo")
    for m, val in enumerate(GENERAL_SYSTEMS_LOOP_YOTPO):
        cell = asm.cell(row=125, column=3 + m, value=val)
        cell.fill = yellow
        cell.number_format = '#,##0'

    # R126: use reserved slot for Shopify with OLD assumption
    asm.cell(row=126, column=2, value="General Systems — Shopify (old assumption)")
    for m, val in enumerate(SHOPIFY_OLD_ASSUMPTION):
        cell = asm.cell(row=126, column=3 + m, value=val)
        cell.fill = yellow
        cell.number_format = '#,##0'
    asm.cell(row=126, column=15, value="=SUM(C126:N126)")

    wb.save(MODEL_PATH)
    print(f"✅ Saved Excel: {MODEL_PATH}")

    loop_yotpo_total = sum(GENERAL_SYSTEMS_LOOP_YOTPO)
    shopify_total = sum(SHOPIFY_OLD_ASSUMPTION)
    print(f"\n  R125 General Systems — Loop+Yotpo:       ${loop_yotpo_total:>10,.0f}")
    print(f"  R126 General Systems — Shopify (old):    ${shopify_total:>10,.0f}")
    print(f"  Combined General Systems total:          ${loop_yotpo_total + shopify_total:>10,.0f}")
    print(f"  (vs Matt's combined $19,581)")
    print(f"\n  Δ vs Phase 2 baseline: +${loop_yotpo_total + shopify_total - 19581:,.0f}")


if __name__ == "__main__":
    main()
