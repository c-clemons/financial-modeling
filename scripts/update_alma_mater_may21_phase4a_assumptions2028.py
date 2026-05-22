"""
Alma Mater Phase 4a - Add 2028 to Assumptions tab (May 21, 2026)
==================================================================

Appends 2028 sections at R193+ (after the existing PO section ends at R191).
No row shifts — formulas elsewhere remain intact.

Adds:
  R193: header "DTC UNIT PROJECTIONS - 2028"
  R195: col headers (Jan-Dec)
  R196: Beta 2028 units
  R197: Alpha 2028 units
  R198: Total units formula
  R200: header "MONTHLY BLENDED AOV (2028)"
  R201: 12 monthly AOV values
  R203: header "OTHER OPERATING EXPENSES - 2028"
  R205: col headers
  R206-R212: 7 marketing/agency OpEx lines (from Matt's 2028 monthly file)
  R213-R218: 6 annual-input items (placeholder = 2027 values pending Matt)
  R219: TOTAL OTHER OPEX (2028) formula
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/"
    "Empirica Financial Model/Alma Mater Financial Model Draft.xlsx"
)

BETA_UNITS_2028 = [79, 103, 150, 254, 339, 339, 337, 274, 224, 175, 340, 255]
ALPHA_UNITS_2028 = [68, 89, 135, 230, 309, 309, 311, 254, 208, 165, 320, 240]
AOV_2028 = [400, 400, 450, 450, 450, 450, 450, 450, 425, 425, 400, 400]

OPEX_2028 = [
    ("Brand Creative — Creative",
     [18500, 32000, 17000, 18500, 32000, 17000, 18500, 32000, 17000, 33500, 17000, 17000]),
    ("Marketing Channels — Mgmt",
     [40000] * 12),
    ("Marketing Channels — Spend",
     [15000, 20000, 26000, 33000, 38000, 38000, 38000, 31000, 31000, 20000, 44000, 44000]),
    ("Channel — Mgmt",
     [5000] * 12),
    ("Channel — Creative+Systems",
     [2000, 4000, 2000, 4000, 2000, 4000, 2000, 4000, 2000, 4000, 2000, 4000]),
    ("General Systems — Loop+Yotpo",
     [0, 0, 0, 509, 509, 509, 509, 509, 509, 509, 509, 509]),
    ("General Systems — Shopify (old assumption)",
     [2850] * 12),
]

ANNUAL_INPUT_2028 = [
    ("Travel & Entertainment", 40000),
    ("Development & Innovation", 50000),
    ("Postage & Shipping", 40000),
    ("Service Charges", 7500),
    ("Phone Services", 2000),
    ("Other Operating", 15000),
]


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)
    asm = wb["Assumptions"]
    bold = Font(bold=True)
    yellow = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    # ============================================================
    # R193: 2028 Unit Projections header
    # ============================================================
    asm.cell(row=193, column=2, value="DTC UNIT PROJECTIONS - 2028").font = bold

    # R195: header row
    asm.cell(row=195, column=2, value="Product").font = bold
    for i, m in enumerate(months):
        asm.cell(row=195, column=3 + i, value=m).font = bold
    asm.cell(row=195, column=15, value="Total").font = bold

    # R196: Beta 2028
    asm.cell(row=196, column=2, value="Beta (V2) Units")
    for i, u in enumerate(BETA_UNITS_2028):
        c = asm.cell(row=196, column=3 + i, value=u)
        c.fill = yellow
    asm.cell(row=196, column=15, value="=SUM(C196:N196)")

    # R197: Alpha 2028
    asm.cell(row=197, column=2, value="Alpha Units")
    for i, u in enumerate(ALPHA_UNITS_2028):
        c = asm.cell(row=197, column=3 + i, value=u)
        c.fill = yellow
    asm.cell(row=197, column=15, value="=SUM(C197:N197)")

    # R198: Total units
    asm.cell(row=198, column=2, value="Total Units").font = bold
    for i in range(12):
        col = chr(ord('C') + i)
        asm.cell(row=198, column=3 + i, value=f"=SUM({col}196:{col}197)")
    asm.cell(row=198, column=15, value="=SUM(C198:N198)")

    # ============================================================
    # R200: 2028 Monthly Blended AOV
    # ============================================================
    asm.cell(row=200, column=2, value="Monthly Blended AOV (2028)").font = bold
    for i, aov in enumerate(AOV_2028):
        c = asm.cell(row=200, column=3 + i, value=aov)
        c.fill = yellow
        c.number_format = '$#,##0'

    # ============================================================
    # R203: 2028 OpEx header
    # ============================================================
    asm.cell(row=203, column=2, value="OTHER OPERATING EXPENSES - 2028").font = bold

    # R205: col headers
    asm.cell(row=205, column=2, value="Expense").font = bold
    for i, m in enumerate(months):
        asm.cell(row=205, column=3 + i, value=m).font = bold
    asm.cell(row=205, column=15, value="Annual").font = bold
    asm.cell(row=205, column=16, value="Ann. Input").font = bold

    # R206-R212: 7 marketing/agency OpEx lines
    for i, (label, monthly) in enumerate(OPEX_2028):
        row = 206 + i
        asm.cell(row=row, column=2, value=label)
        for m, val in enumerate(monthly):
            c = asm.cell(row=row, column=3 + m, value=val)
            c.fill = yellow
            c.number_format = '#,##0'
        asm.cell(row=row, column=15, value=f"=SUM(C{row}:N{row})")

    # R213-R218: 6 annual-input items
    for i, (label, annual) in enumerate(ANNUAL_INPUT_2028):
        row = 213 + i
        asm.cell(row=row, column=2, value=label)
        for m in range(12):
            asm.cell(row=row, column=3 + m, value=f"=$P${row}/12")
        asm.cell(row=row, column=15, value=f"=SUM(C{row}:N{row})")
        c = asm.cell(row=row, column=16, value=annual)
        c.fill = yellow
        c.number_format = '#,##0'

    # R219: TOTAL OTHER OPEX (2028)
    asm.cell(row=219, column=2, value="TOTAL OTHER OPEX (2028)").font = bold
    for i in range(12):
        col = chr(ord('C') + i)
        asm.cell(row=219, column=3 + i, value=f"=SUM({col}206:{col}218)")
    asm.cell(row=219, column=15, value="=SUM(O206:O218)")

    wb.save(MODEL_PATH)
    print(f"✅ Saved Excel: {MODEL_PATH}")

    opex_total = sum(sum(vals) for _, vals in OPEX_2028)
    annual_total = sum(v for _, v in ANNUAL_INPUT_2028)
    print(f"\n2028 OpEx breakdown:")
    for label, monthly in OPEX_2028:
        print(f"  {label:<48} ${sum(monthly):>10,.0f}")
    print(f"  {'─'*48}")
    print(f"  Marketing/Agency total:                          ${opex_total:>10,.0f}")
    print(f"  Annual-input items (placeholder = 2027):         ${annual_total:>10,.0f}")
    print(f"  TOTAL 2028 Other OpEx:                           ${opex_total + annual_total:>10,.0f}")
    print(f"\n2028 Units: Beta {sum(BETA_UNITS_2028):,} + Alpha {sum(ALPHA_UNITS_2028):,} = {sum(BETA_UNITS_2028)+sum(ALPHA_UNITS_2028):,}")
    print(f"2028 AOV: {AOV_2028}")


if __name__ == "__main__":
    main()
