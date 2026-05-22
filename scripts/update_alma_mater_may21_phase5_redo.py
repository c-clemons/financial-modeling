"""
Phase 5 — Re-apply Phases 2 through 4b to the CORRECT Excel file
==================================================================

The May 21 work was applied to a stale copy at
  /Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/
  Empirica Financial Model/Alma Mater Financial Model Draft.xlsx (Mar 23)

The operative file is at
  /Users/chandlerclemons/financial-modeling/models/Alma Mater Financial Model.xlsx
  (May 11) — already has Phase 1 done (Matt's orders + monthly AOV) and
  wholesale layout in channel × month format.

This script applies the remaining work:
  Phase 2:  2026 OpEx restructure to 9-line Section × Cost Type
  Phase 2b: Shopify split (Loop+Yotpo + our $2,850/mo flat)
  Phase 3:  2027 OpEx restructure (correct values from Matt's monthly file cols 5-16)
  Phase 4a: 2028 Assumptions section (units + AOV + OpEx) — appended R193+

Inventory / P&L / Cash Flow / Dashboard 2028 extension done in separate
follow-on scripts (Phase 5b) because they need to be adapted for the correct
file's structure (Cash Flow has QBO-vs-forecast IF logic, Wholesale in P&L
uses channel × month SUMPRODUCTs etc).

Phase 1 (revenue) is NOT re-applied — already in correct file with
all-Beta allocation that's intentional (matches wholesale Beta-only model).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/financial-modeling/models/Alma Mater Financial Model.xlsx"
)

YELLOW = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")
BOLD = Font(bold=True)

# ============================================================
# Phase 2: 2026 OpEx restructure (9 lines)
# ============================================================
OPEX_2026 = [
    ("Brand Creative — Creative",        [375, 8750, 3325, 17175, 0, 0, 0, 0, 0, 0, 0, 0]),
    ("Marketing Channels — Mgmt",        [14150, 12375, 11323, 21525, 22333.75, 19600, 22280, 22280, 22280, 20680, 20680, 20680]),
    ("Marketing Channels — Creative",    [0, 0, 1425, 1000, 1600, 4540, 5040, 4540, 5040, 4540, 5040, 4540]),
    ("Marketing Channels — Spend",       [0, 0, 0, 0, 0, 10640, 15768, 15896, 10896, 8768, 21152, 21152]),
    ("Marketing Channels — Systems",     [0, 0, 0, 0, 0, 90, 90, 90, 90, 90, 90, 90]),
    ("Channel — Mgmt",                   [0, 6100, 4750, 3291, 2250, 1000, 1000, 1000, 1000, 1000, 1000, 1000]),
    ("Channel — Creative",               [0, 0, 3900, 3125, 1100, 1100, 1100, 1100, 0, 0, 0, 0]),
    ("Channel — Systems",                [0, 2200, 150, 2403, 1600, 100, 100, 100, 100, 100, 100, 100]),
    ("General Systems — Loop+Yotpo",     [0, 0, 0, 509, 509, 509, 509, 509, 509, 509, 509, 509]),
    ("General Systems — Shopify (old assumption)", [2850]*12),
]

# ============================================================
# Phase 3: 2027 OpEx restructure (7 lines)
# ============================================================
OPEX_2027 = [
    ("Brand Creative — Creative",        [34000, 12500, 12500, 29000, 12500, 12500, 14000, 12500, 27500, 14000, 12500, 12500]),
    ("Marketing Channels — Mgmt",        [34000]*12),
    ("Marketing Channels — Spend",       [10000, 13000, 16000, 21000, 26000, 26000, 26000, 20000, 20000, 13000, 32000, 32000]),
    ("Channel — Mgmt",                   [4000]*12),
    ("Channel — Creative+Systems",       [1000, 4000, 1000, 4000, 1000, 4000, 1000, 4000, 1000, 4000, 1000, 4000]),
    ("General Systems — Loop+Yotpo",     [0, 0, 0, 509, 509, 509, 509, 509, 509, 509, 509, 509]),
    ("General Systems — Shopify (old assumption)", [2850]*12),
]

# ============================================================
# Phase 4a: 2028 (Beta-only allocation to match correct file pattern)
# ============================================================
# Matt's 2028 orders all go to Beta (consistent with correct file's 2026/2027 = all Beta)
MATT_ORDERS_2028 = [147, 192, 285, 484, 648, 648, 648, 528, 432, 340, 660, 495]
AOV_2028 = [400, 400, 450, 450, 450, 450, 450, 450, 425, 425, 400, 400]

OPEX_2028 = [
    ("Brand Creative — Creative",        [18500, 32000, 17000, 18500, 32000, 17000, 18500, 32000, 17000, 33500, 17000, 17000]),
    ("Marketing Channels — Mgmt",        [40000]*12),
    ("Marketing Channels — Spend",       [15000, 20000, 26000, 33000, 38000, 38000, 38000, 31000, 31000, 20000, 44000, 44000]),
    ("Channel — Mgmt",                   [5000]*12),
    ("Channel — Creative+Systems",       [2000, 4000, 2000, 4000, 2000, 4000, 2000, 4000, 2000, 4000, 2000, 4000]),
    ("General Systems — Loop+Yotpo",     [0, 0, 0, 509, 509, 509, 509, 509, 509, 509, 509, 509]),
    ("General Systems — Shopify (old assumption)", [2850]*12),
]

ANNUAL_INPUT_2028 = [
    ("Travel & Entertainment", 40000),
    ("Development & Innovation", 50000),
    ("Postage & Shipping", 40000),
    ("Service Charges", 7500),
    ("Phone Services", 2000),
    ("Other Operating", 15000),
]


def apply_2026_opex(asm):
    """Replace R117-R126 with new 9-line + Shopify split structure (10 rows)."""
    for i, (label, monthly) in enumerate(OPEX_2026):
        row = 117 + i
        asm.cell(row=row, column=2, value=label)
        for m, val in enumerate(monthly):
            c = asm.cell(row=row, column=3 + m, value=val)
            c.fill = YELLOW
            c.number_format = '#,##0'
        asm.cell(row=row, column=15, value=f"=SUM(C{row}:N{row})")


def apply_2027_opex(asm):
    """Replace R138-R144 with new 7-line structure."""
    # First clear cells (since old structure was 7 rows already)
    for i, (label, monthly) in enumerate(OPEX_2027):
        row = 138 + i
        asm.cell(row=row, column=2, value=label)
        for m, val in enumerate(monthly):
            c = asm.cell(row=row, column=3 + m, value=val)
            c.fill = YELLOW
            c.number_format = '#,##0'
        asm.cell(row=row, column=15, value=f"=SUM(C{row}:N{row})")


def apply_2028_section(asm):
    """Append 2028 section at R193+."""
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    # R193: 2028 Unit Projections header
    asm.cell(row=193, column=2, value="DTC UNIT PROJECTIONS - 2028").font = BOLD

    asm.cell(row=195, column=2, value="Product").font = BOLD
    for i, m in enumerate(months):
        asm.cell(row=195, column=3 + i, value=m).font = BOLD
    asm.cell(row=195, column=15, value="Total").font = BOLD

    # R196: Beta 2028 (all units, matching correct file pattern of all-Beta)
    asm.cell(row=196, column=2, value="Beta (V2) Units")
    for i, u in enumerate(MATT_ORDERS_2028):
        c = asm.cell(row=196, column=3 + i, value=u)
        c.fill = YELLOW
    asm.cell(row=196, column=15, value="=SUM(C196:N196)")

    # R197: Alpha 2028 (0 — matches all-Beta pattern)
    asm.cell(row=197, column=2, value="Alpha Units")
    for i in range(12):
        c = asm.cell(row=197, column=3 + i, value=0)
        c.fill = YELLOW
    asm.cell(row=197, column=15, value="=SUM(C197:N197)")

    # R198: Total
    asm.cell(row=198, column=2, value="Total Units").font = BOLD
    for i in range(12):
        col = chr(ord('C') + i)
        asm.cell(row=198, column=3 + i, value=f"=SUM({col}196:{col}197)")
    asm.cell(row=198, column=15, value="=SUM(C198:N198)")

    # R200: Monthly AOV 2028
    asm.cell(row=200, column=2, value="DTC Monthly Blended AOV (2028)").font = BOLD
    for i, aov in enumerate(AOV_2028):
        c = asm.cell(row=200, column=3 + i, value=aov)
        c.fill = YELLOW
        c.number_format = '$#,##0'

    # R201: Gross DTC Revenue 2028
    asm.cell(row=201, column=2, value="Gross DTC Revenue (2028)")
    for i in range(12):
        col = chr(ord('C') + i)
        asm.cell(row=201, column=3 + i, value=f"={col}196*{col}200")
    asm.cell(row=201, column=15, value="=SUM(C201:N201)")

    # R202: Net DTC Revenue 2028 (use 2027 discount/return rates from C14/C15)
    asm.cell(row=202, column=2, value="Net DTC Revenue (2028)")
    for i in range(12):
        col = chr(ord('C') + i)
        asm.cell(row=202, column=3 + i, value=f"={col}201*(1-$C$14)*(1-$C$15)")
    asm.cell(row=202, column=15, value="=SUM(C202:N202)")

    # R204: 2028 OpEx header
    asm.cell(row=204, column=2, value="OTHER OPERATING EXPENSES - 2028").font = BOLD

    # R206: col headers
    asm.cell(row=206, column=2, value="Expense").font = BOLD
    for i, m in enumerate(months):
        asm.cell(row=206, column=3 + i, value=m).font = BOLD
    asm.cell(row=206, column=15, value="Annual").font = BOLD
    asm.cell(row=206, column=16, value="Ann. Input").font = BOLD

    # R207-R213: 7 marketing/agency OpEx lines
    for i, (label, monthly) in enumerate(OPEX_2028):
        row = 207 + i
        asm.cell(row=row, column=2, value=label)
        for m, val in enumerate(monthly):
            c = asm.cell(row=row, column=3 + m, value=val)
            c.fill = YELLOW
            c.number_format = '#,##0'
        asm.cell(row=row, column=15, value=f"=SUM(C{row}:N{row})")

    # R214-R219: 6 annual-input items
    for i, (label, annual) in enumerate(ANNUAL_INPUT_2028):
        row = 214 + i
        asm.cell(row=row, column=2, value=label)
        for m in range(12):
            asm.cell(row=row, column=3 + m, value=f"=$P${row}/12")
        asm.cell(row=row, column=15, value=f"=SUM(C{row}:N{row})")
        c = asm.cell(row=row, column=16, value=annual)
        c.fill = YELLOW
        c.number_format = '#,##0'

    # R220: TOTAL OTHER OPEX (2028)
    asm.cell(row=220, column=2, value="TOTAL OTHER OPEX (2028)").font = BOLD
    for i in range(12):
        col = chr(ord('C') + i)
        asm.cell(row=220, column=3 + i, value=f"=SUM({col}207:{col}219)")
    asm.cell(row=220, column=15, value="=SUM(O207:O219)")


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)
    asm = wb["Assumptions"]

    apply_2026_opex(asm)
    apply_2027_opex(asm)
    apply_2028_section(asm)

    wb.save(MODEL_PATH)

    # Sanity: verify totals
    print(f"✅ Saved Excel: {MODEL_PATH}")

    total_2026 = sum(sum(vals) for _, vals in OPEX_2026)
    total_2027 = sum(sum(vals) for _, vals in OPEX_2027)
    total_2028 = sum(sum(vals) for _, vals in OPEX_2028)
    annual_2028 = sum(v for _, v in ANNUAL_INPUT_2028)

    print(f"\n2026 OpEx (replaces FSG legacy): ${total_2026:,.0f} (+ ${112000} annual = ${total_2026+112000:,})")
    print(f"2027 OpEx (replaces FSG legacy): ${total_2027:,.0f} (+ ${154500} annual = ${total_2027+154500:,})")
    print(f"2028 OpEx (NEW):                 ${total_2028:,.0f} (+ ${annual_2028:,} annual = ${total_2028+annual_2028:,})")
    print(f"\n2028 DTC: Beta {sum(MATT_ORDERS_2028):,} units × monthly AOV → ${sum(o*a for o,a in zip(MATT_ORDERS_2028, AOV_2028)):,}")


if __name__ == "__main__":
    main()
