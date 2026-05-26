"""
May 22 — Fix DTC discount double-count

Per Matt: "This is gross revenue, prior to returns but it would include
discounts." His Forecast tab AOV is post-discount/pre-return. Model was
applying ANOTHER 5% discount on top, double-counting.

Sets Assumptions!C12 (2026 discount) and C14 (2027 discount) to 0.
Return rates (C13/C15) remain at 20%.

Net DTC multiplier: was × 0.76 (5% × 20%), now × 0.80 (0% × 20%)

Impact (vs prior numbers):
  2026 May-Dec forecast DTC: +$23,510
  2027 DTC:                  +$60,768
  2028 DTC:                  +$95,366
"""
import openpyxl
from pathlib import Path

MODEL = Path("/Users/chandlerclemons/financial-modeling/models/Alma Mater Financial Model.xlsx")

wb = openpyxl.load_workbook(MODEL)
asm = wb["Assumptions"]
asm.cell(row=12, column=3, value=0)
asm.cell(row=14, column=3, value=0)
asm.cell(row=12, column=2, value="DTC Discount Rate (2026) - $0 (Matt AOV includes discount)")
asm.cell(row=14, column=2, value="DTC Discount Rate (2027) - $0 (Matt AOV includes discount)")
wb.save(MODEL)
print(f"✅ {MODEL}: R12 + R14 discount rates set to 0")
