"""
Phase 3 FIX — 2027 OpEx (May 21, 2026)
========================================

Phase 3 (original) mis-attributed Matt's Budget tab cols 20-31 as 2027
when they're actually 2028. Per row-1 labels in
AM 26-28 Marketing_Ecom Budget with 2028 budget.xlsx:
  - Col D: 2026 (annual single value)
  - Cols E-S: 2027 (monthly)
  - Cols T-AH: 2028 (monthly)

The original Phase 3 used cols T-AE — those numbers belong in Phase 4 (2028).

This script replaces R138-R144 with the CORRECT 2027 values from cols 5-16.

Old (wrong) 2027 total: $1,262,781
New (correct) 2027 total: $985,781

Lines unchanged:
  - Channel — Mgmt: still eCommerce only
  - General Systems — Loop+Yotpo: carry from 2026
  - General Systems — Shopify (old): our $2,850/mo flat
"""

import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/"
    "Empirica Financial Model/Alma Mater Financial Model Draft.xlsx"
)

# CORRECTED 2027 OpEx (from monthly file cols 5-16)
CORRECT_2027_OPEX = [
    # R138: Brand Creative — Creative
    # R30 Studio + R31 Field + R32 UGC + R35 Copywriting + R36 Image + R37 Digital Design
    ("Brand Creative — Creative",
     [34000, 12500, 12500, 29000, 12500, 12500, 14000, 12500, 27500, 14000, 12500, 12500]),

    # R139: Marketing Channels — Mgmt ($34K/mo flat = 7 roles ex eCommerce)
    # Creative Lead $5K + Marketing Lead $5K + CRM $5K + SEO $3K + Org Soc $6K + Perf Mkt $6K + Community $4K
    ("Marketing Channels — Mgmt",
     [34000, 34000, 34000, 34000, 34000, 34000, 34000, 34000, 34000, 34000, 34000, 34000]),

    # R140: Marketing Channels — Spend
    # R48 Perf Mkt + R51 PostPilot
    ("Marketing Channels — Spend",
     [10000, 13000, 16000, 21000, 26000, 26000, 26000, 20000, 20000, 13000, 32000, 32000]),

    # R141: Channel — Mgmt (eCommerce $4K/mo)
    ("Channel — Mgmt",
     [4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000]),

    # R142: Channel — Creative+Systems (UX $1K + Dev alternating)
    ("Channel — Creative+Systems",
     [1000, 4000, 1000, 4000, 1000, 4000, 1000, 4000, 1000, 4000, 1000, 4000]),

    # R143: General Systems — Loop+Yotpo (carry from 2026)
    ("General Systems — Loop+Yotpo",
     [0, 0, 0, 509, 509, 509, 509, 509, 509, 509, 509, 509]),

    # R144: General Systems — Shopify (old $2,850/mo flat)
    ("General Systems — Shopify (old assumption)",
     [2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850]),
]


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)
    asm = wb["Assumptions"]
    yellow = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")

    total = sum(sum(vals) for _, vals in CORRECT_2027_OPEX)
    expected = 985781
    assert abs(total - expected) < 100, f"Total mismatch: ${total:,} vs expected ${expected:,}"

    for i, (label, monthly) in enumerate(CORRECT_2027_OPEX):
        row = 138 + i
        asm.cell(row=row, column=2, value=label)
        for m, val in enumerate(monthly):
            cell = asm.cell(row=row, column=3 + m, value=val)
            cell.fill = yellow
            cell.number_format = '#,##0'

    wb.save(MODEL_PATH)
    print(f"✅ Saved Excel: {MODEL_PATH}")
    print(f"\n2027 OpEx (CORRECTED):")
    for label, monthly in CORRECT_2027_OPEX:
        print(f"  {label:<48} ${sum(monthly):>10,.0f}")
    print(f"  {'─'*48}")
    print(f"  Marketing/Agency total: ${total:>10,.0f}")
    print(f"  Annual-input (R145-R150 unchanged): $154,500")
    print(f"  TOTAL 2027: ${total + 154500:>10,.0f}")


if __name__ == "__main__":
    main()
