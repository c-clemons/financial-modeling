"""
May 22 — Input DTC + Wholesale unit actuals from Shopify; mark cells green
============================================================================

Per user: "Let's input the actuals for DTC and wholesale orders since we
have those via shopify. The assumptions tab should utilize actuals when
available and actuals should display as green just like they do on the
Actuals tab (both on the assumptions and Dashboard tabs)."

Shopify data (closed months Jan-Apr 2026):
  Jan:  DTC=27, WS=0,  Gifting=0    Total orders=27
  Feb:  DTC=35, WS=8,  Gifting=0    Total orders=43
  Mar:  DTC=87, WS=17, Gifting=21   Total orders=125
  Apr:  DTC=120, WS=7, Gifting=31   Total orders=158
  (May:  DTC=88, WS=12, Gifting=38 — partial, not used since Last Actuals = 4)

Changes:
  1. QBO Actuals tab: add 2026 Unit Actuals section (R82-R86)
     DTC Orders, Wholesale Orders, Gifting Orders, Total Orders
  2. Assumptions R66 (Beta DTC units 2026) — Jan-Apr replaced with actuals,
     formula uses IF(month<=Assumptions!$C$47, actual, Matt forecast).
     May-Dec untouched (Matt's forecast).
  3. Assumptions R102 (Other Wholesale 2026) — Jan-Apr populated with
     WS actuals (other channels stay at 0 since we don't have channel
     breakdown from Shopify).
  4. Green fill on actual cells: same color as QBO Actuals tab (FFE2EFDA)
     applied to:
       - Assumptions R66 cols C-F (DTC actuals)
       - Assumptions R102 cols D-G (WS actuals)
       - Dashboard R37-R50 cols C-F (entire 2026 monthly detail block
         for closed months — they already pull QBO via IF formula)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
from pathlib import Path

MODEL = Path("/Users/chandlerclemons/financial-modeling/models/Alma Mater Financial Model.xlsx")

# Color matching existing QBO Actuals green
GREEN = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
BOLD = Font(bold=True)

# Shopify actuals for closed months (Jan-Apr 2026)
DTC_ACTUALS    = [27,  35,  87,  120]   # Jan, Feb, Mar, Apr
WS_ACTUALS     = [0,   8,   17,  7]
GIFTING_ACTUALS = [0,  0,   21,  31]

# Matt's original forecast for R66 (for the IF fallback when month > last actuals)
MATT_FORECAST_DTC_2026 = [27, 35, 91, 133.7, 194.7, 288, 297.5, 180, 156, 180, 396, 267]


def main():
    wb = openpyxl.load_workbook(MODEL)
    qbo = wb['QBO Actuals']
    asm = wb['Assumptions']
    dash = wb['Dashboard']

    # ============================================================
    # 1. QBO Actuals tab: Add 2026 Unit Actuals section at R82+
    # ============================================================
    # Header
    qbo.cell(row=82, column=2, value="2026 UNIT ACTUALS (Shopify)").font = BOLD
    qbo.cell(row=82, column=2).fill = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
    qbo.cell(row=82, column=2).font = Font(bold=True, color="FFFFFFFF")

    # Column headers (same as R42)
    months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    qbo.cell(row=83, column=2, value="Line Item").font = BOLD
    for i, m in enumerate(months):
        qbo.cell(row=83, column=3 + i, value=m).font = BOLD
    qbo.cell(row=83, column=15, value="YTD 2026").font = BOLD

    # Data rows
    for r, label, data in [
        (84, "DTC Orders",      DTC_ACTUALS),
        (85, "Wholesale Orders", WS_ACTUALS),
        (86, "Gifting Orders",   GIFTING_ACTUALS),
    ]:
        qbo.cell(row=r, column=2, value=label)
        for i, val in enumerate(data):
            cell = qbo.cell(row=r, column=3 + i, value=val)
            cell.fill = GREEN
        # YTD formula
        qbo.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})")

    # Total orders row
    qbo.cell(row=87, column=2, value="Total Orders").font = BOLD
    for i in range(12):
        col = chr(67 + i)  # C..N
        qbo.cell(row=87, column=3 + i, value=f"={col}84+{col}85+{col}86")
        if i < 4:  # green for closed months
            qbo.cell(row=87, column=3 + i).fill = GREEN
            qbo.cell(row=87, column=3 + i).font = BOLD
    qbo.cell(row=87, column=15, value="=SUM(C87:N87)").font = BOLD

    # ============================================================
    # 2. Assumptions R66 (Beta DTC units 2026) — use actuals for closed months
    # ============================================================
    # Pattern: =IF(month <= $C$47, 'QBO Actuals'!cell, Matt_forecast_value)
    for i in range(12):
        month_num = i + 1
        col_letter = chr(67 + i)  # C..N
        col = 3 + i
        if month_num <= 4:  # Jan-Apr: actuals
            formula = f"=IF({month_num}<=$C$47,'QBO Actuals'!{col_letter}84,{MATT_FORECAST_DTC_2026[i]})"
            cell = asm.cell(row=66, column=col, value=formula)
            cell.fill = GREEN
        # May-Dec: leave alone (Matt's forecast untouched)

    # ============================================================
    # 3. Assumptions Other Wholesale R102 — actuals for Jan-Apr
    # ============================================================
    # Format: cols D-O = Jan-Dec, col 4 = D = Jan
    for i in range(4):
        col = 4 + i  # D, E, F, G = Jan, Feb, Mar, Apr
        col_letter = chr(64 + col)  # D, E, F, G
        month_num = i + 1
        # IF formula: actual when closed, else 0 (no Other Wholesale forecast)
        qbo_col = chr(67 + i)  # C..F maps to Jan-Apr in QBO Actuals R85
        formula = f"=IF({month_num}<=Assumptions!$C$47,'QBO Actuals'!{qbo_col}85,0)"
        cell = asm.cell(row=102, column=col, value=formula)
        cell.fill = GREEN

    # ============================================================
    # 4. Green fill on Dashboard 2026 Monthly Detail (closed months Jan-Apr)
    # ============================================================
    # Rows 37-50 (DTC Rev, WS Rev, Total Rev, COGS, GP, GM%, Team, OpEx, Total OpEx, EBITDA, EBITDA%)
    # Already use IF(QBO/Forecast) formula — just need green fill for closed months
    dashboard_rows = [37, 38, 39, 41, 42, 43, 45, 46, 47, 49, 50]
    for r in dashboard_rows:
        for i in range(4):  # cols C-F = Jan-Apr
            cell = dash.cell(row=r, column=3 + i)
            cell.fill = GREEN

    # ============================================================
    # 5. Save
    # ============================================================
    wb.save(MODEL)
    print(f"✅ Saved Excel: {MODEL}")
    print(f"\n=== Updates ===")
    print(f"QBO Actuals R82-R87: Added 2026 Unit Actuals (Shopify) section")
    print(f"  DTC Orders:      {DTC_ACTUALS}")
    print(f"  Wholesale Orders: {WS_ACTUALS}")
    print(f"  Gifting Orders:   {GIFTING_ACTUALS}")
    print(f"\nAssumptions R66 cols C-F: Beta DTC units 2026 — IF(month<=4, QBO actual, Matt fcst)")
    print(f"  Pre-update Matt forecast Jan-Apr: {MATT_FORECAST_DTC_2026[:4]}")
    print(f"  Now resolves to actuals: {DTC_ACTUALS}")
    print(f"  Delta: Jan=0, Feb=0, Mar=-4, Apr=-13.7")
    print(f"\nAssumptions R102 (Other Wholesale 2026) cols D-G: WS actuals {WS_ACTUALS}")
    print(f"\nGreen fill applied to:")
    print(f"  Assumptions R66 cols C-F (DTC actuals)")
    print(f"  Assumptions R102 cols D-G (WS actuals)")
    print(f"  Dashboard R37-R50 cols C-F (2026 monthly detail closed months)")


if __name__ == "__main__":
    main()
