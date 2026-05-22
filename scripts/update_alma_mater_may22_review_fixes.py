"""
May 22 Review Fixes
=====================

Four fixes from the comprehensive review of Alma Mater Financial Model.xlsx:

  #1: Inventory R32 cols R-AL — unify to dual-AOV pattern (matches C-Q):
      OLD: =(col12+col22)*MonthlyAOV
      NEW: =col12*MonthlyAOV + col22*$C$9  (Alpha @ $450 fixed)

  #2: Cash Flow R31 cols R-AL — extend sophisticated SUMPRODUCT formula
      (subtracts only "Projected" funding rounds, matches C-Q logic).

  #3: Monthly P&L R130 cols C-N — reference Team Costs S16:AD16 (2027 cols)
      instead of G16:R16 (2026 cols). 2028 Team Costs increases from
      $424,851 → $521,037 (~$96K). 2028 EBITDA drops accordingly.

  #4: Wholesale 2028 channel names — rename:
      R113 "International - 1" → "International - Japan"
      R114 "International - 2" → "International - Canada"
"""

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/financial-modeling/models/Alma Mater Financial Model.xlsx"
)


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)

    # ============================================================
    # FIX #1: Inventory R32 cols R-AL → dual-AOV
    # ============================================================
    inv = wb['Inventory']

    # Cols R-Z (Apr-Dec 2027): monthly AOV is at Assumptions row 80
    # Cols AA-AL (Jan-Dec 2028): monthly AOV is at Assumptions row 91
    # Map inv col → asm col for monthly AOV
    for c in range(18, 39):  # R-AL
        col = get_column_letter(c)
        if 18 <= c <= 26:
            # 2027: monthly AOV row 80
            asm_aov_row = 80
            asm_aov_col = get_column_letter(c - 12)  # R(18) → F(6), Z(26) → N(14)
        else:
            # 2028: monthly AOV row 91
            asm_aov_row = 91
            asm_aov_col = get_column_letter(c - 24)  # AA(27) → C(3), AL(38) → N(14)
        new_formula = f"={col}12*Assumptions!{asm_aov_col}${asm_aov_row}+{col}22*Assumptions!$C$9"
        inv.cell(row=32, column=c, value=new_formula)

    # ============================================================
    # FIX #2: Cash Flow R31 cols R-AL → sophisticated SUMPRODUCT
    # ============================================================
    cf = wb['Cash Flow']
    # Existing C-Q pattern (Mar'27):
    # =Q24 - SUMPRODUCT((Assumptions!$F$58:$F$60="Projected") *
    #                   ((Assumptions!$E$58:$E$60-2026)*12 + Assumptions!$D$58:$D$60 <= 15)
    #                   * Assumptions!$C$58:$C$60)
    # Month index N for each col: col 3 (C) = month 1 (Jan'26); col 38 (AL) = month 36 (Dec'28)
    for c in range(18, 39):
        col = get_column_letter(c)
        month_idx = c - 2  # col 18 → 16 (Apr'27), col 38 → 36 (Dec'28)
        new_formula = (
            f'={col}24-SUMPRODUCT((Assumptions!$F$58:$F$60="Projected")*'
            f'((Assumptions!$E$58:$E$60-2026)*12+Assumptions!$D$58:$D$60<={month_idx})*'
            f'Assumptions!$C$58:$C$60)'
        )
        cf.cell(row=31, column=c, value=new_formula)

    # ============================================================
    # FIX #3: Monthly P&L R130 cols C-N → reference Team Costs S16:AD16
    # ============================================================
    pl = wb['Monthly P&L']
    # 2028 Team Costs should match 2027 (same team, no new hires modeled).
    # Team Costs cols S(19)=Jan, T(20)=Feb, ..., AD(30)=Dec
    for i in range(12):
        tc_col = get_column_letter(19 + i)  # S, T, ..., AD
        pl.cell(row=130, column=3 + i, value=f"='Team Costs'!{tc_col}16")

    # ============================================================
    # FIX #4: Rename wholesale 2028 international channels
    # ============================================================
    asm = wb['Assumptions']
    asm.cell(row=113, column=2, value="International - Japan")
    asm.cell(row=114, column=2, value="International - Canada")

    wb.save(MODEL_PATH)
    print(f"✅ Saved: {MODEL_PATH}\n")
    print("Fixes applied:")
    print("  #1 Inventory R32 cols R-AL: now use dual-AOV =col12*MonthlyAOV + col22*$C$9")
    print("  #2 Cash Flow R31 cols R-AL: now use sophisticated SUMPRODUCT (only subtracts Projected)")
    print("  #3 Monthly P&L R130: 2028 Team Costs now refs Team Costs S16-AD16 (2027 cols)")
    print("     → 2028 Team Costs annual: $424,851 → $521,037 (+$96K)")
    print("     → 2028 EBITDA will be ~$96K worse than before fix")
    print("  #4 R113/R114 renamed to 'International - Japan' / 'International - Canada'")


if __name__ == "__main__":
    main()
