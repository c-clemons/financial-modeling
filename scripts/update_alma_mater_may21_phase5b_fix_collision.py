"""
Phase 5b-fix — Resolve row collision at R220 (May 21, 2026)
=============================================================

Bug: Phase 5b put WS 2028 header at R220 which overwrote the "TOTAL OTHER
OPEX (2028)" label (formulas survived but label got clobbered).

Fix:
  1. Restore R220 label to "TOTAL OTHER OPEX (2028)"
  2. Move WS 2028 section from R220-R226 → R230-R236 (safe gap below OpEx)
  3. Update all references that pointed to R222-R226 → R230-R234
     (Inventory R9 cols AA-AL, Monthly P&L R119/R124 cols C-N)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/financial-modeling/models/Alma Mater Financial Model.xlsx"
)

YELLOW = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")
BOLD = Font(bold=True)
MONTHS_SHORT = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

GG_2028_MONTHLY = [0, 500, 2000, 500, 0, 0, 1000, 3000, 1000, 0, 0, 0]
WHOLESALE_2028 = [
    ("Big Box Retail",      [0]*12,           "TBD - placeholder"),
    ("Green Grass (CC)",    GG_2028_MONTHLY,  "2× 2027 GG placeholder (8,000u)"),
    ("Other Wholesale",     [0]*12,           "Placeholder"),
    ("International - 1",   [0]*12,           "TBD"),
    ("International - 2",   [0]*12,           "TBD"),
]


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)
    asm = wb["Assumptions"]

    # ============================================================
    # 1. Restore R220 = "TOTAL OTHER OPEX (2028)" label
    # ============================================================
    asm.cell(row=220, column=2, value="TOTAL OTHER OPEX (2028)").font = BOLD

    # ============================================================
    # 2. Clear OLD WS 2028 placement (R220-R226 cols B-T were wholesale)
    #    R220 col B already restored. Clear cols 17-20 (Q/R/S/T):
    # ============================================================
    for c in range(17, 21):  # Q, R, S, T
        asm.cell(row=220, column=c, value=None)
    for r in range(221, 227):  # R221-R226 were WS rows
        for c in range(2, 21):
            asm.cell(row=r, column=c, value=None)

    # ============================================================
    # 3. Re-create WS 2028 at R230-R236 (header + col headers + 5 channels)
    # ============================================================
    NEW_WS_HEADER_ROW = 228  # blank row R229, then content R230+
    asm.cell(row=NEW_WS_HEADER_ROW, column=2,
             value="WHOLESALE FORECAST - 2028 (PLACEHOLDER)").font = BOLD

    # Col headers at R229
    asm.cell(row=229, column=2, value="Channel").font = BOLD
    asm.cell(row=229, column=3, value="Year").font = BOLD
    for i, m in enumerate(MONTHS_SHORT):
        asm.cell(row=229, column=4 + i, value=m).font = BOLD
    asm.cell(row=229, column=16, value="Total Units").font = BOLD
    asm.cell(row=229, column=17, value="ASP").font = BOLD
    asm.cell(row=229, column=18, value="Annual Rev").font = BOLD
    asm.cell(row=229, column=19, value="Annual COGS").font = BOLD
    asm.cell(row=229, column=20, value="Notes").font = BOLD

    # R230-R234: 5 channel rows
    for i, (channel, monthly, notes) in enumerate(WHOLESALE_2028):
        row = 230 + i
        asm.cell(row=row, column=2, value=channel)
        asm.cell(row=row, column=3, value=2028)
        for m, val in enumerate(monthly):
            c = asm.cell(row=row, column=4 + m, value=val)
            c.fill = YELLOW
            c.number_format = '#,##0'
        asm.cell(row=row, column=16, value=f"=SUM(D{row}:O{row})")
        c = asm.cell(row=row, column=17, value=144)
        c.fill = YELLOW
        c.number_format = '$#,##0'
        asm.cell(row=row, column=18, value=f"=P{row}*Q{row}")
        asm.cell(row=row, column=18).number_format = '$#,##0'
        asm.cell(row=row, column=19, value=f"=P{row}*$C$33")
        asm.cell(row=row, column=19).number_format = '$#,##0'
        asm.cell(row=row, column=20, value=notes)

    # ============================================================
    # 4. Update Inventory R9 (WS Beta) cols AA-AL: R222-R226 → R230-R234
    # ============================================================
    inv = wb["Inventory"]
    for c in range(27, 39):  # AA-AL = 2028 cols
        col = get_column_letter(c)
        asm_col = get_column_letter(c - 23)  # AA(27) → D(4), AL(38) → O(15)
        inv.cell(row=9, column=c,
                 value=f"=SUM(Assumptions!${asm_col}$230:${asm_col}$234)")

    # ============================================================
    # 5. Update Monthly P&L R119 (WS Rev) + R124 (WS COGS) cols C-N
    # ============================================================
    pl = wb["Monthly P&L"]
    for i in range(12):
        asm_col = get_column_letter(4 + i)
        pl.cell(row=119, column=3 + i,
                value=f"=SUMPRODUCT(Assumptions!${asm_col}$230:${asm_col}$234,"
                      f"Assumptions!$Q$230:$Q$234)")
        pl.cell(row=124, column=3 + i,
                value=f"=SUM(Assumptions!${asm_col}$230:${asm_col}$234)*Assumptions!$C$33")

    wb.save(MODEL_PATH)
    print(f"✅ Saved Excel: {MODEL_PATH}")
    print("Fixes applied:")
    print("  R220 restored to 'TOTAL OTHER OPEX (2028)' label")
    print("  Old WS placement R221-R226 cleared")
    print("  New WS 2028 at R228-R234 (header R228, col headers R229, channels R230-R234)")
    print("  Inventory R9 cols AA-AL updated to ref R230:R234")
    print("  Monthly P&L R119/R124 cols C-N updated to ref R230:R234")


if __name__ == "__main__":
    main()
