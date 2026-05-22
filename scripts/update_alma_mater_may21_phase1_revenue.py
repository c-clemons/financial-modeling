"""
Alma Mater Phase 1 - Revenue Rebuild (May 21, 2026)
====================================================

Updates Excel model with Matt's monthly revenue forecast:
  1. New monthly AOV arrays in Assumptions (R69 for 2026, R80 for 2027)
  2. Updated DTC unit projections (Beta/Alpha allocated from Matt's monthly orders
     using current monthly Beta:Alpha ratios)
  3. Updated Inventory!R32 formula to use monthly AOV instead of single $250/$450

Sources:
  - Matt's Forecast tab (AM 26-28 Marketing_Ecom Budget_Monthly.xlsx)
  - User direction: keep Beta/Alpha unit structure, use Matt's monthly blended AOV

This script does NOT touch:
  - OpEx (Phase 2 + 3 will handle)
  - Team Cost tab (Matt's people are agency, not internal team)
  - Wholesale tab
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/"
    "Empirica Financial Model/Alma Mater Financial Model Draft.xlsx"
)

# ============================================================================
# Matt's monthly forecast values (from Forecast tab)
# ============================================================================

# Monthly blended AOV (gross, from Matt's Shopify-based forecast)
AOV_2026 = [551, 362, 407, 300, 300, 300, 300, 300, 300, 300, 300, 300]
AOV_2027 = [350, 350, 375, 425, 450, 450, 450, 425, 425, 425, 400, 400]

# Matt's monthly orders (rounded to integer; fractional values from Apr/May/Jul 2026)
MATT_ORDERS_2026 = [27, 35, 91, 134, 195, 288, 298, 180, 156, 180, 396, 267]
MATT_ORDERS_2027 = [84, 112, 216, 360, 396, 396, 396, 320, 288, 272, 440, 330]

# Beta/Alpha allocation derived from current model ratios applied to Matt's orders
# Current 2026 Beta: [10, 20, 30, 50, 100, 150, 200, 225, 250, 275, 300, 325]
# Current 2026 Alpha: [0, 0, 0, 0, 0, 0, 50, 100, 200, 300, 300, 0]
# Allocation: months 1-6 = 100% Beta; months 7-11 split by ratio; Dec 100% Beta
BETA_UNITS_2026 = [27, 35, 91, 134, 195, 288, 238, 125, 87, 86, 198, 267]
ALPHA_UNITS_2026 = [0, 0, 0, 0, 0, 0, 60, 55, 69, 94, 198, 0]

# Current 2027 ratios (52-54% Beta, 46-48% Alpha)
BETA_UNITS_2027 = [45, 60, 114, 189, 207, 207, 206, 166, 149, 140, 227, 170]
ALPHA_UNITS_2027 = [39, 52, 102, 171, 189, 189, 190, 154, 139, 132, 213, 160]


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)

    # ---- Sanity: verify allocation sums match Matt's totals ----
    assert sum(BETA_UNITS_2026) + sum(ALPHA_UNITS_2026) == sum(MATT_ORDERS_2026), \
        f"2026 unit allocation mismatch: {sum(BETA_UNITS_2026)+sum(ALPHA_UNITS_2026)} vs {sum(MATT_ORDERS_2026)}"
    assert sum(BETA_UNITS_2027) + sum(ALPHA_UNITS_2027) == sum(MATT_ORDERS_2027), \
        f"2027 unit allocation mismatch: {sum(BETA_UNITS_2027)+sum(ALPHA_UNITS_2027)} vs {sum(MATT_ORDERS_2027)}"

    yellow = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")
    bold = Font(bold=True)

    # ============================================================================
    # 1. Assumptions tab - add monthly AOV arrays (R69 for 2026, R80 for 2027)
    # ============================================================================
    asm = wb["Assumptions"]

    # 2026 Monthly AOV - row 69 (was empty, sits between R68 Total Units and R70 Gross DTC Revenue header)
    asm.cell(row=69, column=2, value="Monthly Blended AOV (2026)").font = bold
    for i, aov in enumerate(AOV_2026):
        c = asm.cell(row=69, column=3 + i, value=aov)
        c.fill = yellow
        c.number_format = "$#,##0"

    # 2027 Monthly AOV - row 80 (was empty, sits between R79 Total Units and R81 Gross DTC Revenue header)
    asm.cell(row=80, column=2, value="Monthly Blended AOV (2027)").font = bold
    for i, aov in enumerate(AOV_2027):
        c = asm.cell(row=80, column=3 + i, value=aov)
        c.fill = yellow
        c.number_format = "$#,##0"

    # ============================================================================
    # 2. Update DTC Unit Projections
    # ============================================================================
    # 2026 Beta units: R66, cols C-N (Jan-Dec)
    for i, units in enumerate(BETA_UNITS_2026):
        c = asm.cell(row=66, column=3 + i, value=units)
        c.fill = yellow
    # 2026 Alpha units: R67
    for i, units in enumerate(ALPHA_UNITS_2026):
        c = asm.cell(row=67, column=3 + i, value=units)
        c.fill = yellow

    # 2027 Beta units: R77
    for i, units in enumerate(BETA_UNITS_2027):
        c = asm.cell(row=77, column=3 + i, value=units)
        c.fill = yellow
    # 2027 Alpha units: R78
    for i, units in enumerate(ALPHA_UNITS_2027):
        c = asm.cell(row=78, column=3 + i, value=units)
        c.fill = yellow

    # ============================================================================
    # 3. Update Inventory tab - Constrained Gross DTC Revenue (R32)
    #    Old formula: =C12*Assumptions!$C$7+C22*Assumptions!$C$9   (Beta×$250 + Alpha×$450)
    #    New formula: =(C12+C22)*Assumptions!C$69                   (Total units × monthly AOV)
    #    Cols C-N = 2026 (use Assumptions!C$69:N$69)
    #    Cols O-Q = 2027 Jan-Mar (use Assumptions!C$80:E$80)
    # ============================================================================
    inv = wb["Inventory"]

    # Column letter helper
    from openpyxl.utils import get_column_letter

    # 2026 columns: C(3) through N(14) -> Jan26..Dec26 -> AOV row 69, cols C-N
    for i in range(12):
        col_inv = get_column_letter(3 + i)  # C..N
        col_aov = get_column_letter(3 + i)  # C..N (same offset since AOV array also starts at col C)
        inv.cell(row=32, column=3 + i,
                 value=f"=({col_inv}12+{col_inv}22)*Assumptions!{col_aov}$69")

    # 2027 columns in Inventory: O(15)=Jan27, P(16)=Feb27, Q(17)=Mar27
    # 2027 AOV row 80 starts at col C (Jan) through N (Dec)
    for i in range(3):  # Jan, Feb, Mar 2027
        col_inv = get_column_letter(15 + i)  # O, P, Q
        col_aov = get_column_letter(3 + i)   # C, D, E (Jan, Feb, Mar of 2027 AOV row)
        inv.cell(row=32, column=15 + i,
                 value=f"=({col_inv}12+{col_inv}22)*Assumptions!{col_aov}$80")

    # ============================================================================
    # Save
    # ============================================================================
    wb.save(MODEL_PATH)
    print(f"✅ Saved: {MODEL_PATH}")

    # ---- Verification math ----
    print("\nExpected 2026 monthly DTC gross revenue (Matt's forecast):")
    expected_2026 = [u * a for u, a in zip(MATT_ORDERS_2026, AOV_2026)]
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    for m, rev, beta, alpha in zip(months, expected_2026, BETA_UNITS_2026, ALPHA_UNITS_2026):
        print(f"  {m}: {beta} Beta + {alpha} Alpha = {beta+alpha} units × ${AOV_2026[months.index(m)]} = ${rev:,}")
    print(f"  2026 Total: ${sum(expected_2026):,} (Matt's forecast: $692,506 — diff: ${sum(expected_2026)-692506:,})")

    print("\nExpected 2027 monthly DTC gross revenue (Matt's forecast):")
    expected_2027 = [u * a for u, a in zip(MATT_ORDERS_2027, AOV_2027)]
    print(f"  2027 Total: ${sum(expected_2027):,} (Matt's forecast: $1,519,200 — diff: ${sum(expected_2027)-1519200:,})")


if __name__ == "__main__":
    main()
