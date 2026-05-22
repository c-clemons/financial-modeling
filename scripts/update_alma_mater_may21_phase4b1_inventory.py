"""
Phase 4b-1 — Excel Inventory tab extension (May 21, 2026)
==========================================================

Extends Inventory tab from 15 months (Jan'26-Mar'27) to 36 months
(Jan'26-Dec'28). Adds cols R through AL (21 new monthly columns).

Source rows that get extended:
  R7-R13:  Beta inventory (Beg, +PO, -WS, Available, Demand, Sales, Ending)
  R17-R23: Alpha inventory (same)
  R27-R28: Total summary (Ending + DTC Sales)
  R32:     Constrained Gross DTC Revenue
  R33:     Constrained Net DTC Revenue
  R37-R56: PO Payment Schedule (20 PO rows)
  R57:     Total Inventory Payments

Header rows R6, R16, R26, R31, R36 also get month labels extended.

Key column mappings for Assumptions references:
  Cols C-N (Jan-Dec 2026): Assumptions row 66 (Beta) / 67 (Alpha) / 69 (AOV)
  Cols O-Z (Jan-Dec 2027): Assumptions row 77 / 78 / 80
  Cols AA-AL (Jan-Dec 2028): Assumptions row 196 / 197 / 200

Discount/return rates: 2026 uses Assumptions C12/C13, 2027+2028 both use C14/C15.

SUMPRODUCT formulas (R8, R9, R37-R56) auto-handle new cols via COLUMN()-2.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/"
    "Empirica Financial Model/Alma Mater Financial Model Draft.xlsx"
)

# Column ranges
FIRST_NEW_COL = 18  # R (Apr'27)
LAST_NEW_COL = 38   # AL (Dec'28)
YEAR_2027_START = 15  # O = Jan'27
YEAR_2028_START = 27  # AA = Jan'28

MONTHS_SHORT = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def get_asm_ref_for_units(inv_col, product):
    """Map Inventory col (15-38) to Assumptions cell for DTC demand units."""
    if 15 <= inv_col <= 26:
        # 2027 — Asm rows 77 (Beta) / 78 (Alpha), col = inv_col - 12
        row = 77 if product == 'Beta' else 78
        asm_col = inv_col - 12
    elif 27 <= inv_col <= 38:
        # 2028 — Asm rows 196 (Beta) / 197 (Alpha), col = inv_col - 24
        row = 196 if product == 'Beta' else 197
        asm_col = inv_col - 24
    else:
        # 2026 — Asm rows 66/67, col = inv_col
        row = 66 if product == 'Beta' else 67
        asm_col = inv_col
    return f"Assumptions!{get_column_letter(asm_col)}{row}"


def get_aov_ref(inv_col):
    """Map Inventory col to Assumptions monthly AOV cell."""
    if 15 <= inv_col <= 26:
        return f"Assumptions!{get_column_letter(inv_col - 12)}$80"
    elif 27 <= inv_col <= 38:
        return f"Assumptions!{get_column_letter(inv_col - 24)}$200"
    else:
        return f"Assumptions!{get_column_letter(inv_col)}$69"


def get_discount_return_refs(inv_col):
    """For 2026 use C12/C13, for 2027+ use C14/C15."""
    if inv_col <= 14:  # 2026
        return "Assumptions!$C$12", "Assumptions!$C$13"
    else:  # 2027 and 2028 share
        return "Assumptions!$C$14", "Assumptions!$C$15"


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)
    inv = wb["Inventory"]

    # ============================================================
    # 1. Extend month headers (R6, R16, R26, R31, R36)
    # ============================================================
    # Pattern: "Jan'26", ..., "Dec'28"
    # Col 3 = Jan'26, col 14 = Dec'26, col 15 = Jan'27, ..., col 26 = Dec'27,
    # col 27 = Jan'28, col 38 = Dec'28
    for header_row in [6, 16, 26, 31, 36]:
        for c in range(FIRST_NEW_COL, LAST_NEW_COL + 1):
            month_idx = (c - 3) % 12  # 0-11
            year = 2026 + (c - 3) // 12
            year_suffix = str(year)[-2:]
            inv.cell(row=header_row, column=c, value=f"{MONTHS_SHORT[month_idx]}'{year_suffix}")

    # ============================================================
    # 2. Beta Inventory block (R7-R13)
    # ============================================================
    for c in range(FIRST_NEW_COL, LAST_NEW_COL + 1):
        col = get_column_letter(c)
        prev_col = get_column_letter(c - 1)

        # R7: Beginning Inventory (Beta) — references prior month's ending
        inv.cell(row=7, column=c, value=f"={prev_col}13")

        # R8: PO Arrivals (Beta) — SUMPRODUCT formula, copy from col Q
        inv.cell(row=8, column=c,
                 value='=SUMPRODUCT((Assumptions!$C$172:$C$191="Beta")*'
                       '((Assumptions!$G$172:$G$191-2026)*12+Assumptions!$F$172:$F$191+Assumptions!$C$166=COLUMN()-2)*'
                       'Assumptions!$D$172:$D$191)')

        # R9: Wholesale Shipments (Beta)
        inv.cell(row=9, column=c,
                 value='=SUMPRODUCT(((Assumptions!$K$88:$K$93-2026)*12+Assumptions!$J$88:$J$93=COLUMN()-2)*'
                       '(Assumptions!$L$88:$L$93="Beta")*Assumptions!$C$88:$C$93)')

        # R10: Available for DTC (Beta)
        inv.cell(row=10, column=c, value=f"={col}7+{col}8-{col}9")

        # R11: DTC Demand (Beta) — varies by year
        inv.cell(row=11, column=c, value=f"={get_asm_ref_for_units(c, 'Beta')}")

        # R12: DTC Sales (Beta) — constrained by available
        inv.cell(row=12, column=c, value=f"=MIN({col}11,MAX({col}10,0))")

        # R13: Ending Inventory (Beta)
        inv.cell(row=13, column=c, value=f"={col}10-{col}12")

    # ============================================================
    # 3. Alpha Inventory block (R17-R23)
    # ============================================================
    for c in range(FIRST_NEW_COL, LAST_NEW_COL + 1):
        col = get_column_letter(c)
        prev_col = get_column_letter(c - 1)

        # R17: Beginning Inventory (Alpha)
        inv.cell(row=17, column=c, value=f"={prev_col}23")

        # R18: PO Arrivals (Alpha)
        inv.cell(row=18, column=c,
                 value='=SUMPRODUCT((Assumptions!$C$172:$C$191="Alpha")*'
                       '((Assumptions!$G$172:$G$191-2026)*12+Assumptions!$F$172:$F$191+Assumptions!$C$166=COLUMN()-2)*'
                       'Assumptions!$D$172:$D$191)')

        # R19: Wholesale Shipments (Alpha)
        inv.cell(row=19, column=c,
                 value='=SUMPRODUCT(((Assumptions!$K$88:$K$93-2026)*12+Assumptions!$J$88:$J$93=COLUMN()-2)*'
                       '(Assumptions!$L$88:$L$93="Alpha")*Assumptions!$C$88:$C$93)')

        # R20: Available for DTC (Alpha)
        inv.cell(row=20, column=c, value=f"={col}17+{col}18-{col}19")

        # R21: DTC Demand (Alpha)
        inv.cell(row=21, column=c, value=f"={get_asm_ref_for_units(c, 'Alpha')}")

        # R22: DTC Sales (Alpha)
        inv.cell(row=22, column=c, value=f"=MIN({col}21,MAX({col}20,0))")

        # R23: Ending Inventory (Alpha)
        inv.cell(row=23, column=c, value=f"={col}20-{col}22")

    # ============================================================
    # 4. Total summary (R27, R28)
    # ============================================================
    for c in range(FIRST_NEW_COL, LAST_NEW_COL + 1):
        col = get_column_letter(c)
        # R27: Total Ending Inventory
        inv.cell(row=27, column=c, value=f"={col}13+{col}23")
        # R28: Total DTC Sales (constrained)
        inv.cell(row=28, column=c, value=f"={col}12+{col}22")

    # ============================================================
    # 5. Constrained DTC Revenue (R32, R33)
    # ============================================================
    for c in range(FIRST_NEW_COL, LAST_NEW_COL + 1):
        col = get_column_letter(c)
        # R32: Constrained Gross DTC Revenue = (Beta + Alpha units) × monthly AOV
        inv.cell(row=32, column=c, value=f"=({col}12+{col}22)*{get_aov_ref(c)}")
        # R33: Constrained Net DTC Revenue
        disc_ref, ret_ref = get_discount_return_refs(c)
        inv.cell(row=33, column=c, value=f"={col}32*(1-{disc_ref})*(1-{ret_ref})")

    # ============================================================
    # 6. PO Payment Schedule (R37-R56)
    # Each row uses IF + COLUMN()-2 pattern referencing a specific PO row
    # (Assumptions!E172 through E191). Extends naturally.
    # ============================================================
    for po_row in range(37, 57):  # 20 PO rows
        asm_row = 172 + (po_row - 37)  # Maps Inv R37 → Asm R172, R56 → R191
        for c in range(FIRST_NEW_COL, LAST_NEW_COL + 1):
            inv.cell(row=po_row, column=c,
                     value=f'=IF(AND(Assumptions!$E${asm_row}>0,'
                           f'COLUMN()-2=(Assumptions!$G${asm_row}-2026)*12+Assumptions!$F${asm_row}+'
                           f'Assumptions!$C$166+Assumptions!$C$167),'
                           f'Assumptions!$E${asm_row},0)')

    # R57: Total Inventory Payments
    for c in range(FIRST_NEW_COL, LAST_NEW_COL + 1):
        col = get_column_letter(c)
        inv.cell(row=57, column=c, value=f"=SUM({col}37:{col}56)")

    # ============================================================
    # Save
    # ============================================================
    wb.save(MODEL_PATH)
    print(f"✅ Saved Excel: {MODEL_PATH}")
    print(f"   Extended Inventory tab from col Q (Mar'27) to col AL (Dec'28)")
    print(f"   21 new columns × ~30 active rows = ~630 cells populated")
    print(f"\nKey extensions:")
    print(f"  Beta inventory chain (R7-R13)")
    print(f"  Alpha inventory chain (R17-R23)")
    print(f"  Total summary (R27-R28)")
    print(f"  Constrained DTC Revenue Gross+Net (R32-R33)")
    print(f"  PO Payment Schedule (R37-R57)")
    print(f"  Month headers (R6, R16, R26, R31, R36)")


if __name__ == "__main__":
    main()
