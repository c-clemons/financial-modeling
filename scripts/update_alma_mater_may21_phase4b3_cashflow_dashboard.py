"""
Phase 4b-3/4 — Cash Flow extension + Dashboard 2028 column (May 21, 2026)
==========================================================================

Cash Flow: extends from Mar'27 (col Q) to Dec'28 (col AL) — 21 new monthly cols.
Each existing row gets extended:
  R6:  Month headers
  R8:  DTC Revenue (pulls from Monthly P&L)
  R9:  Wholesale Revenue
  R10: Funding/Investment (SUMPRODUCT with hardcoded month/year per col)
  R11: Total Cash In
  R14: Inventory Purchases (← Inventory tab)
  R15: Fulfillment COGS (← Inventory × COGS rate)
  R16: Team Costs (← Monthly P&L)
  R17: Other OpEx (← Monthly P&L)
  R18: Total Cash Out
  R21: Monthly Net Cash Flow
  R23: Beginning Cash (← prior col ending)
  R24: Ending Cash Balance
  R26: Days of Cash
  R31: Cash Balance (No NEW Funding)

Dashboard: adds 2028 Forecast column (col F) + extends 2026 monthly detail comment.
  R7-R14: KPI rollup column for 2028
  R31: Update Dec'27 cash to Cash Flow!Z24 (was already wrong — was Z24, that's correct)
  R32: New row for "Ending Cash - Dec 2028" → Cash Flow!AL24

Source row mapping in Monthly P&L:
  2027 section: rev=95, ws=96, team=107, other=108
  2028 section: rev=118, ws=119, team=130, other=131
"""

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/"
    "Empirica Financial Model/Alma Mater Financial Model Draft.xlsx"
)

FIRST_NEW_COL = 18  # R (Apr'27)
LAST_NEW_COL = 38   # AL (Dec'28)
MONTHS_SHORT = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def cf_col_to_pl_ref(cf_col, pl_row_2027, pl_row_2028):
    """Map a Cash Flow col to the corresponding Monthly P&L cell reference."""
    if 18 <= cf_col <= 26:
        # 2027 Apr-Dec
        pl_col = get_column_letter(cf_col - 12)  # cf18→F(6), cf26→N(14)
        return f"'Monthly P&L'!{pl_col}{pl_row_2027}"
    elif 27 <= cf_col <= 38:
        # 2028 Jan-Dec
        pl_col = get_column_letter(cf_col - 24)  # cf27→C(3), cf38→N(14)
        return f"'Monthly P&L'!{pl_col}{pl_row_2028}"


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)
    cf = wb['Cash Flow']

    # ============================================================
    # CASH FLOW: Month headers (R6)
    # ============================================================
    for c in range(FIRST_NEW_COL, LAST_NEW_COL + 1):
        month_idx = (c - 3) % 12
        year = 2026 + (c - 3) // 12
        year_suffix = str(year)[-2:]
        cf.cell(row=6, column=c, value=f"{MONTHS_SHORT[month_idx]}'{year_suffix}")

    # ============================================================
    # Per-column formulas
    # ============================================================
    for c in range(FIRST_NEW_COL, LAST_NEW_COL + 1):
        col = get_column_letter(c)
        prev_col = get_column_letter(c - 1)
        month_num = (c - 3) % 12 + 1
        year = 2026 + (c - 3) // 12

        # R8: DTC Revenue
        cf.cell(row=8, column=c, value=f"={cf_col_to_pl_ref(c, 95, 118)}")

        # R9: Wholesale Revenue
        cf.cell(row=9, column=c, value=f"={cf_col_to_pl_ref(c, 96, 119)}")

        # R10: Funding/Investment (SUMPRODUCT with hardcoded month/year)
        cf.cell(row=10, column=c,
                value=f"=SUMPRODUCT((Assumptions!$D$58:$D$60={month_num})*"
                      f"(Assumptions!$E$58:$E$60={year})*Assumptions!$C$58:$C$60)")

        # R11: Total Cash In
        cf.cell(row=11, column=c, value=f"={col}8+{col}9+{col}10")

        # R14: Inventory Purchases (← Inventory!col57)
        cf.cell(row=14, column=c, value=f"=Inventory!{col}57")

        # R15: Fulfillment COGS (← Inventory!col32 × (cogs_rate - product_cost_rate))
        # Same formula pattern as existing 2026/2027 cols
        cf.cell(row=15, column=c,
                value=f"=Inventory!{col}32*(Assumptions!$C$24-Assumptions!$C$20)")

        # R16: Team Costs (← Monthly P&L row 107 for 2027, row 130 for 2028)
        cf.cell(row=16, column=c, value=f"={cf_col_to_pl_ref(c, 107, 130)}")

        # R17: Other OpEx (← Monthly P&L row 108 for 2027, row 131 for 2028)
        cf.cell(row=17, column=c, value=f"={cf_col_to_pl_ref(c, 108, 131)}")

        # R18: Total Cash Out
        cf.cell(row=18, column=c, value=f"={col}14+{col}15+{col}16+{col}17")

        # R21: Monthly Net Cash Flow
        cf.cell(row=21, column=c, value=f"={col}11-{col}18")

        # R23: Beginning Cash (= prior col's ending)
        cf.cell(row=23, column=c, value=f"={prev_col}24")

        # R24: Ending Cash Balance
        cf.cell(row=24, column=c, value=f"={col}23+{col}21")

        # R26: Days of Cash
        cf.cell(row=26, column=c, value=f"=IF({col}18>0,{col}24/({col}18/30),0)")

        # R31: Cash Balance (No NEW Funding) — chains from prior col, excludes R10 funding
        cf.cell(row=31, column=c, value=f"={prev_col}31+{col}8+{col}9-{col}18")

    cf.cell(row=4, column=2, value="36-MONTH CASH FLOW").font = Font(bold=True)

    # ============================================================
    # DASHBOARD: Add 2028 column (col F = 6) — currently has 2025/2026/2027
    # ============================================================
    dash = wb['Dashboard']

    # R7: column header
    dash.cell(row=7, column=6, value="2028 Forecast").font = Font(bold=True)

    # R8-R14: KPI rows for 2028 — references new 2028 P&L section
    dash.cell(row=8, column=6, value="='Monthly P&L'!O120")   # Total Revenue 2028
    dash.cell(row=9, column=6, value="='Monthly P&L'!O125")   # Total COGS
    dash.cell(row=10, column=6, value="='Monthly P&L'!O126")  # Gross Profit
    dash.cell(row=11, column=6, value="='Monthly P&L'!O127")  # Gross Margin %
    dash.cell(row=11, column=6).number_format = '0.0%'
    dash.cell(row=12, column=6, value="='Monthly P&L'!O132")  # Total OpEx
    dash.cell(row=13, column=6, value="='Monthly P&L'!O133")  # EBITDA
    dash.cell(row=14, column=6, value="='Monthly P&L'!O134")  # EBITDA Margin %
    dash.cell(row=14, column=6).number_format = '0.0%'

    # R31: Ending Cash - Dec 2027 (Cash Flow!Z24 = col 26 = Dec 2027)
    # was set to Z24 originally which is now Dec'27 in extended Cash Flow. Verify.
    # Then add R32 for Dec 2028.
    dash.cell(row=31, column=2, value="Ending Cash - Dec 2027")
    dash.cell(row=31, column=3, value="='Cash Flow'!Z24")
    dash.cell(row=32, column=2, value="Ending Cash - Dec 2028")
    dash.cell(row=32, column=3, value="='Cash Flow'!AL24")

    wb.save(MODEL_PATH)
    print(f"✅ Saved Excel: {MODEL_PATH}")
    print(f"\nCash Flow extended from col Q (Mar'27) to col AL (Dec'28)")
    print(f"  21 new cols × 11 active rows = 231 cells populated")
    print(f"\nDashboard updates:")
    print(f"  Added col F (2028 Forecast) — R7-R14 KPIs from 2028 P&L section")
    print(f"  Added R32 'Ending Cash - Dec 2028' → Cash Flow!AL24")


if __name__ == "__main__":
    main()
