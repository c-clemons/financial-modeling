"""
Phase 4c — Add 2028 Wholesale Deals (doubled from 2027) — May 21, 2026
========================================================================

Per user direction: double each 2027 wholesale deal for 2028 as a placeholder
until the wholesale team provides input.

2027 → 2028:
  Spring 2027 (Beta) 2,500u @ 125 doors → Spring 2028 5,000u @ 250 doors
  Fall 2027 (Beta) 4,000u @ 200 doors    → Fall 2028 8,000u @ 400 doors
  Fall 2027 (Alpha) 800u @ 100 doors     → Fall 2028 1,600u @ 200 doors

Excel changes:
  R93: was '(Add Deal)' placeholder → Spring 2028 Beta
  R94: was blank separator           → Fall 2028 Beta
  R95: was 'Total 2026 WS' summary   → Fall 2028 Alpha
  R96: new '(Add Deal)' placeholder
  R98: 'Total 2026 WS' (moved from R95)
  R99: 'Total 2027 WS' (moved from R96)
  R100: 'Total 2028 WS' (new)

All 144 SUMPRODUCT references using '88:93' updated to '88:96' across
Inventory tab, Monthly P&L tab, and Assumptions summary rows.

Streamlit baseline_data BASELINE_WHOLESALE also gets the 3 new 2028 deals.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path
import re
import json

EXCEL_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/"
    "Empirica Financial Model/Alma Mater Financial Model Draft.xlsx"
)

# New 2028 deals (doubled from 2027)
NEW_2028_DEALS = [
    # (label, units, ws_price, cogs, doors, del_month, year, product)
    ("Spring 2028 (Beta)",  5000, 144, 333300, 250, 3, 2028, "Beta"),
    ("Fall 2028 (Beta)",    8000, 144, 555500, 400, 8, 2028, "Beta"),
    ("Fall 2028 (Alpha)",   1600, 244, 336300, 200, 9, 2028, "Alpha"),
]


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    asm = wb['Assumptions']
    yellow = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")
    bold = Font(bold=True)

    # ============================================================
    # 1. Add 3 new 2028 deals at R93, R94, R95
    # ============================================================
    for i, (label, units, ws_price, cogs, doors, month, year, product) in enumerate(NEW_2028_DEALS):
        row = 93 + i
        asm.cell(row=row, column=2, value=label)
        asm.cell(row=row, column=3, value=units).fill = yellow
        asm.cell(row=row, column=4, value=ws_price).fill = yellow
        asm.cell(row=row, column=5, value=f"=C{row}*D{row}")
        asm.cell(row=row, column=6, value=cogs).fill = yellow
        asm.cell(row=row, column=7, value=f"=E{row}-F{row}")
        asm.cell(row=row, column=8, value=f'=IF(E{row}>0,G{row}/E{row},0)')
        asm.cell(row=row, column=8).number_format = '0.0%'
        asm.cell(row=row, column=9, value=doors).fill = yellow
        asm.cell(row=row, column=10, value=month).fill = yellow
        asm.cell(row=row, column=11, value=year).fill = yellow
        asm.cell(row=row, column=12, value=product).fill = yellow

    # R96: new (Add Deal) placeholder
    asm.cell(row=96, column=2, value="(Add Deal)")
    asm.cell(row=96, column=3, value=0).fill = yellow
    asm.cell(row=96, column=4, value=0).fill = yellow
    asm.cell(row=96, column=5, value="=C96*D96")
    asm.cell(row=96, column=6, value=0).fill = yellow
    asm.cell(row=96, column=7, value="=E96-F96")
    asm.cell(row=96, column=8, value="=IF(E96>0,G96/E96,0)")
    asm.cell(row=96, column=8).number_format = '0.0%'
    asm.cell(row=96, column=9, value=0).fill = yellow
    asm.cell(row=96, column=10, value=0).fill = yellow
    asm.cell(row=96, column=11, value=2028).fill = yellow
    asm.cell(row=96, column=12, value="Beta").fill = yellow

    # ============================================================
    # 2. Update ALL formulas with 88:93 → 88:96 across the workbook
    # ============================================================
    pattern = re.compile(r'(\$?[A-Z]?\$?)88:(\$?[A-Z]?\$?)93')
    update_count = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '88:' in cell.value and '93' in cell.value:
                    new_val = pattern.sub(r'\g<1>88:\g<2>96', cell.value)
                    if new_val != cell.value:
                        cell.value = new_val
                        update_count += 1

    print(f"Updated {update_count} formulas (88:93 → 88:96)")

    # ============================================================
    # 3. Move summary rows: R95→R98 (Total 2026), R96→R99 (Total 2027), new R100 (Total 2028)
    #    NOTE: R95 and R96 were already overwritten by new deals (Fall 2028 Alpha, (Add Deal)).
    #    The OLD summary formulas were captured by the substitution but on wrong rows now.
    #    Need to explicitly write the moved summaries.
    # ============================================================
    # R97 separator (leave blank)

    # R98: Total 2026 WS — note formulas updated to 88:96
    asm.cell(row=98, column=2, value="Total 2026 WS").font = bold
    asm.cell(row=98, column=5, value="=SUMPRODUCT((K88:K96=2026)*E88:E96)")
    asm.cell(row=98, column=5).number_format = '"$"#,##0'
    asm.cell(row=98, column=6, value="=SUMPRODUCT((K88:K96=2026)*F88:F96)")
    asm.cell(row=98, column=6).number_format = '"$"#,##0'

    # R99: Total 2027 WS
    asm.cell(row=99, column=2, value="Total 2027 WS").font = bold
    asm.cell(row=99, column=5, value="=SUMPRODUCT((K88:K96=2027)*E88:E96)")
    asm.cell(row=99, column=5).number_format = '"$"#,##0'
    asm.cell(row=99, column=6, value="=SUMPRODUCT((K88:K96=2027)*F88:F96)")
    asm.cell(row=99, column=6).number_format = '"$"#,##0'

    # R100: Total 2028 WS (NEW)
    asm.cell(row=100, column=2, value="Total 2028 WS").font = bold
    asm.cell(row=100, column=5, value="=SUMPRODUCT((K88:K96=2028)*E88:E96)")
    asm.cell(row=100, column=5).number_format = '"$"#,##0'
    asm.cell(row=100, column=6, value="=SUMPRODUCT((K88:K96=2028)*F88:F96)")
    asm.cell(row=100, column=6).number_format = '"$"#,##0'

    wb.save(EXCEL_PATH)
    print(f"\n✅ Saved Excel: {EXCEL_PATH}")

    # Summary
    print(f"\n2028 Wholesale Deals (doubled from 2027):")
    total_rev_2028 = 0
    total_cogs_2028 = 0
    for label, units, ws_price, cogs, doors, month, year, product in NEW_2028_DEALS:
        rev = units * ws_price
        gp = rev - cogs
        gp_pct = gp / rev * 100
        total_rev_2028 += rev
        total_cogs_2028 += cogs
        print(f"  {label:<30}  {units:>5}u @ ${ws_price}  Rev ${rev:>7,}  COGS ${cogs:>7,}  GP {gp_pct:.1f}%  {doors:>3} doors  Mo {month}")
    print(f"  {'─'*30}")
    print(f"  2028 Total WS Rev:  ${total_rev_2028:,}")
    print(f"  2028 Total WS COGS: ${total_cogs_2028:,}")
    print(f"  2028 Total WS GP:   ${total_rev_2028 - total_cogs_2028:,}")


if __name__ == "__main__":
    main()
