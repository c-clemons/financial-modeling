"""
Alma Mater Phase 2 - 2026 OpEx Restructure (May 21, 2026)
==========================================================

Replaces the 10 existing 2026 OpEx line items (R117-R126) with 9 lines using
the (Section × Cost Type) compromise structure from Matt's new 2026 Budget file
(AM Marketing_Ecom Budget.xlsx).

Source data:
  /Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/
  Alma Mater Client Supplied Models/AM Marketing_Ecom Budget.xlsx
  → "Budget" tab, subtotal rows R16/R76-R79/R90-R92/R103

Total 2026 OpEx (Matt's plan): $463,469 (vs old model $379k)
  Jan-Apr = Q1 actuals from Matt's tab
  May-Dec = Plan values

Q1 actuals overlap caveat: our model continues to use QBO actuals for closed
months (Assumptions!C47 = Last Actuals Month). Matt's Q1 "actuals" values are
marketing-only spend and inform forward planning, not our actuals source.

Annual-input rows (R127-R132) unchanged: Travel, Dev & Innovation,
Postage & Shipping, Service Charges, Phone Services, Other Operating.

R133 TOTAL formula already covers R117:R132, picks up the new lines auto.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/"
    "Empirica Financial Model/Alma Mater Financial Model Draft.xlsx"
)

# Format: (label, monthly_values_jan_to_dec)
# Pulled from new 2026 Budget file subtotal rows
NEW_2026_OPEX = [
    ("Brand Creative — Creative",        [375, 8750, 3325, 17175, 0, 0, 0, 0, 0, 0, 0, 0]),
    ("Marketing Channels — Mgmt",        [14150, 12375, 11323, 21525, 22333.75, 19600, 22280, 22280, 22280, 20680, 20680, 20680]),
    ("Marketing Channels — Creative",    [0, 0, 1425, 1000, 1600, 4540, 5040, 4540, 5040, 4540, 5040, 4540]),
    ("Marketing Channels — Spend",       [0, 0, 0, 0, 0, 10640, 15768, 15896, 10896, 8768, 21152, 21152]),
    ("Marketing Channels — Systems",     [0, 0, 0, 0, 0, 90, 90, 90, 90, 90, 90, 90]),
    ("Channel — Mgmt",                   [0, 6100, 4750, 3291, 2250, 1000, 1000, 1000, 1000, 1000, 1000, 1000]),
    ("Channel — Creative",               [0, 0, 3900, 3125, 1100, 1100, 1100, 1100, 0, 0, 0, 0]),
    ("Channel — Systems",                [0, 2200, 150, 2403, 1600, 100, 100, 100, 100, 100, 100, 100]),
    ("General Systems",                  [0, 0, 2500, 3009, 3009, 3009, 3009, 3009, 509, 509, 509, 509]),
]


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)
    asm = wb["Assumptions"]

    yellow = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")

    # Sanity: verify total matches Matt's $463,469
    total = sum(sum(vals) for _, vals in NEW_2026_OPEX)
    expected = 463468.75
    assert abs(total - expected) < 1, f"Total mismatch: ${total:,} vs expected ${expected:,}"

    # Replace R117-R125 with Matt's 9 new line items
    for i, (label, monthly) in enumerate(NEW_2026_OPEX):
        row = 117 + i
        asm.cell(row=row, column=2, value=label)
        for m, val in enumerate(monthly):
            cell = asm.cell(row=row, column=3 + m, value=val)
            cell.fill = yellow
            cell.number_format = '#,##0'
        # Annual sum formula (col O = 15)
        asm.cell(row=row, column=15, value=f"=SUM(C{row}:N{row})")

    # R126 was UpPromote — repurpose as reserved/placeholder with all zeros
    asm.cell(row=126, column=2, value="(reserved — extra line)")
    for m in range(12):
        cell = asm.cell(row=126, column=3 + m, value=0)
        cell.fill = yellow
        cell.number_format = '#,##0'
    asm.cell(row=126, column=15, value="=SUM(C126:N126)")

    # R133 TOTAL formula unchanged (=SUM(C117:C132)) — picks up new lines
    # automatically since rows didn't shift.

    wb.save(MODEL_PATH)
    print(f"✅ Saved Excel: {MODEL_PATH}")
    print(f"\n2026 OpEx restructured to 9 lines + 1 reserved. Total: ${total:,.2f}")
    print(f"  vs old model: $379k (Old FSG/Perf Mkt/Shopify/Klaviyo/etc.)")
    print(f"  vs Matt's plan: ${expected:,.2f}")

    # Print line items for sanity
    print("\nNew 2026 OpEx breakdown:")
    for label, monthly in NEW_2026_OPEX:
        print(f"  {label}: ${sum(monthly):>10,.0f}")
    print(f"  ── annual-input items (R127-R132 unchanged) ── $112,000")
    print(f"  TOTAL 2026 (Other OpEx, ex-Team Costs): ${total + 112000:,.0f}")


if __name__ == "__main__":
    main()
