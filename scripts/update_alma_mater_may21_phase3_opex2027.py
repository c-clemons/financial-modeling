"""
Alma Mater Phase 3 - 2027 OpEx Restructure (May 21, 2026)
==========================================================

Replaces the 7 existing 2027 OpEx line items (R138-R144) with 7 lines
aggregated from Matt's monthly file 2027 columns. Same conceptual
structure as Phase 2 (Section × Cost Type), but more aggregated since
Matt's 2027 view is less granular than the new 2026 Budget file.

Mapping notes:
  - Creative Lead ($60K) lumped into Marketing Channels — Mgmt with the
    other agency mgmt roles (Marketing Lead, CRM, SEO, Organic Social,
    Perf Mkt Mgmt, Community Mgmt)
  - Marketing Channels — Creative ($0): Matt's 2027 doesn't break out
    marketing creative production separately
  - Channel — Creative+Systems: UX Design + Development combined
  - General Systems — Loop+Yotpo: carried forward from 2026 (Matt didn't
    update for 2027)
  - General Systems — Shopify: our $2,850/mo flat assumption (pending
    Matt confirmation on his $2,500 Mar-Aug only cadence)

Total 2027 marketing/agency OpEx: $1,265,781 (vs old model $730k)
Plus annual-input items R145-R150 unchanged at $154,500 = $1,420,281
"""

import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path

MODEL_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/"
    "Empirica Financial Model/Alma Mater Financial Model Draft.xlsx"
)

# 2027 monthly values per line item (from Matt's monthly file, aggregated)
NEW_2027_OPEX = [
    # R138: Brand Creative — Creative
    # Source: R30 Photo + R31 Field Shoots + R32 UGC + R35 Copywriting + R36 Image Proc + R37 Digital Design
    ("Brand Creative — Creative",
     [18500, 32000, 17000, 18500, 32000, 17000, 18500, 32000, 17000, 33500, 17000, 17000]),

    # R139: Marketing Channels — Mgmt ($40K/mo flat = 7 agency roles)
    # Creative Lead $5K + Marketing Lead $6K + CRM $6K + SEO/AIO $4K + Organic Social $7K + Perf Mkt $7K + Community $5K
    ("Marketing Channels — Mgmt",
     [40000, 40000, 40000, 40000, 40000, 40000, 40000, 40000, 40000, 40000, 40000, 40000]),

    # R140: Marketing Channels — Spend
    # Perf Mkt (R48): [12, 16, 20, 25, 30, 30, 30, 25, 25, 16, 35, 35]
    # PostPilot (R51): [3, 4, 6, 8, 8, 8, 8, 6, 6, 4, 9, 9]
    ("Marketing Channels — Spend",
     [15000, 20000, 26000, 33000, 38000, 38000, 38000, 31000, 31000, 20000, 44000, 44000]),

    # R141: Channel — Mgmt (eCommerce Mgmt $5K/mo)
    ("Channel — Mgmt",
     [5000, 5000, 5000, 5000, 5000, 5000, 5000, 5000, 5000, 5000, 5000, 5000]),

    # R142: Channel — Creative+Systems
    # UX Design (R63): [1000]×12 + Development (R64): [1, 3, 1, 3, 1, 3, 1, 3, 1, 3, 1, 3]
    ("Channel — Creative+Systems",
     [2000, 4000, 2000, 4000, 2000, 4000, 2000, 4000, 2000, 4000, 2000, 4000]),

    # R143: General Systems — Loop+Yotpo (carried from 2026)
    ("General Systems — Loop+Yotpo",
     [0, 0, 0, 509, 509, 509, 509, 509, 509, 509, 509, 509]),

    # R144: General Systems — Shopify (our $2,850/mo flat assumption)
    ("General Systems — Shopify (old assumption)",
     [2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850, 2850]),
]


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)
    asm = wb["Assumptions"]
    yellow = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")

    # Sanity total
    total = sum(sum(vals) for _, vals in NEW_2027_OPEX)
    expected = 1262781  # Brand Creative actual sum = $270K (not $273K from R4 Total col which includes budget)
    assert abs(total - expected) < 100, f"Total mismatch: ${total:,} vs expected ${expected:,}"

    # Replace R138-R144 (7 rows, 7 line items)
    for i, (label, monthly) in enumerate(NEW_2027_OPEX):
        row = 138 + i
        asm.cell(row=row, column=2, value=label)
        for m, val in enumerate(monthly):
            cell = asm.cell(row=row, column=3 + m, value=val)
            cell.fill = yellow
            cell.number_format = '#,##0'
        asm.cell(row=row, column=15, value=f"=SUM(C{row}:N{row})")

    # R145-R150 unchanged (annual-input items)
    # R151 TOTAL formula unchanged: =SUM(C138:C150) — picks up new lines auto

    wb.save(MODEL_PATH)
    print(f"✅ Saved Excel: {MODEL_PATH}")

    print(f"\nNew 2027 OpEx breakdown:")
    for label, monthly in NEW_2027_OPEX:
        print(f"  {label:<48} ${sum(monthly):>10,.0f}")
    print(f"  {'─'*48}")
    print(f"  Marketing/Agency total: ${total:>10,.0f}")
    print(f"  Annual-input items (R145-R150 unchanged): $154,500")
    print(f"  TOTAL 2027 Other OpEx: ${total + 154500:>10,.0f}")
    print(f"  (Old model: ~$914K, was treating 2027=2026)")


if __name__ == "__main__":
    main()
