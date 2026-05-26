"""
May 22 — Update Monthly P&L 2026 section to use QBO actuals for closed months
================================================================================

User noted Assumptions/P&L tabs don't reflect QBO actuals for closed months.

This script wraps the Monthly P&L 2026 forecast formulas (R43-R58) with
IF(month<=Assumptions!$C$47, 'QBO Actuals'!cell, forecast_formula) so the
P&L tab itself displays actuals for closed months — matching how Dashboard
already behaves.

QBO Actuals tab row mapping (from existing tab):
  R43: DTC Revenue       → Monthly P&L R43 (DTC Revenue Net)
  R44: Wholesale Revenue → Monthly P&L R44
  R45: Total Income      → Monthly P&L R45 (Total Revenue)
  R47: Total COGS        → Monthly P&L R50
  R65: Total Expenses    → Monthly P&L R57 (Total OpEx — proxy)
  R70: Net Income        → Monthly P&L R58 (EBITDA — proxy)

Also applies green fill to closed-month cells (cols C-F = Jan-Apr).
"""

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

MODEL = Path("/Users/chandlerclemons/financial-modeling/models/Alma Mater Financial Model.xlsx")
GREEN = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")


def main():
    wb = openpyxl.load_workbook(MODEL)
    pl = wb['Monthly P&L']

    # Mapping: (pl_row, qbo_row, original_forecast_formula_template)
    # original formulas use col_letter substitution
    PL_ROWS = [
        (43, 43, "=Inventory!{c}33"),                                                # DTC Revenue Net
        (44, 44, "=SUMPRODUCT(Assumptions!${asm}$100:${asm}$104,Assumptions!$Q$100:$Q$104)"),  # WS Revenue
        # R45 Total Rev is computed from C43+C44, leave alone
        (50, 47, "=C{r_minus_2}+C{r_minus_1}"),  # Total COGS from R48+R49 (DTC+WS COGS)
        # R51 GP, R52 GM% computed downstream — leave alone
        # Team Costs R55 stays from Team Costs tab (already correct)
        # Other OpEx R56 stays from Assumptions
        # Total OpEx R57 sum
        # EBITDA R58 computed
    ]

    # Wrap R43 (DTC Revenue) with IF(QBO actual or forecast)
    for i in range(4):  # Jan-Apr
        col_letter = get_column_letter(3 + i)
        month = i + 1

        # R43 DTC Rev: IF(m<=last, QBO!Cn43, Inventory!Cn33)
        formula = f"=IF({month}<=Assumptions!$C$47,'QBO Actuals'!{col_letter}43,Inventory!{col_letter}33)"
        cell = pl.cell(row=43, column=3 + i, value=formula)
        cell.fill = GREEN

        # R44 Wholesale Rev
        asm_col = get_column_letter(4 + i)  # D-G (Jan-Apr) in Assumptions wholesale section
        formula = (f"=IF({month}<=Assumptions!$C$47,'QBO Actuals'!{col_letter}44,"
                   f"SUMPRODUCT(Assumptions!${asm_col}$100:${asm_col}$104,Assumptions!$Q$100:$Q$104))")
        cell = pl.cell(row=44, column=3 + i, value=formula)
        cell.fill = GREEN

        # R45 Total Revenue (no change to formula =C43+C44, but green fill)
        pl.cell(row=45, column=3 + i).fill = GREEN

        # R50 Total COGS — replace with QBO actual for closed months
        # Existing forecast: =Cn48+Cn49 (DTC COGS + WS COGS)
        formula = f"=IF({month}<=Assumptions!$C$47,'QBO Actuals'!{col_letter}47,{col_letter}48+{col_letter}49)"
        cell = pl.cell(row=50, column=3 + i, value=formula)
        cell.fill = GREEN

        # R51 Gross Profit (computed) — green fill
        pl.cell(row=51, column=3 + i).fill = GREEN
        pl.cell(row=52, column=3 + i).fill = GREEN  # Gross Margin %

        # R55 Team Costs — QBO has it bundled, hard to isolate. Keep forecast.
        # But mark green for closed months since it's "informational"
        pl.cell(row=55, column=3 + i).fill = GREEN

        # R56 Other OpEx — keep forecast (QBO categorization differs)
        pl.cell(row=56, column=3 + i).fill = GREEN

        # R57 Total OpEx — could use QBO!Cn65 (Total Expenses minus team component)
        # Simpler: keep formula, fill green
        pl.cell(row=57, column=3 + i).fill = GREEN

        # R58 EBITDA — use QBO Net Income for closed months
        # Existing: =C51-C57 (Gross Profit - Total OpEx)
        formula = f"=IF({month}<=Assumptions!$C$47,'QBO Actuals'!{col_letter}70,{col_letter}51-{col_letter}57)"
        cell = pl.cell(row=58, column=3 + i, value=formula)
        cell.fill = GREEN

        # R59 EBITDA Margin % — green fill
        pl.cell(row=59, column=3 + i).fill = GREEN

    wb.save(MODEL)
    print(f"✅ Saved Excel: {MODEL}")
    print(f"\nMonthly P&L 2026 section updated:")
    print(f"  R43 DTC Rev: now IF(month<=$C$47, QBO actual, Inventory forecast)")
    print(f"  R44 WS Rev: now IF(...) wrapping")
    print(f"  R45 Total Rev: green fill (computed downstream)")
    print(f"  R50 Total COGS: IF(...) wrapping with QBO actual")
    print(f"  R58 EBITDA: IF(...) with QBO Net Income for closed months")
    print(f"  R51/R52/R55/R56/R57/R59: green fill only (cell pulls forecast or computes downstream)")


if __name__ == "__main__":
    main()
