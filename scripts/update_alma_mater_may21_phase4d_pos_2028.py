"""
Phase 4d — Add 2028 Purchase Orders (May 21, 2026)
====================================================

Adds 6 placeholder 2028 POs to support the doubled 2028 wholesale deals plus
projected DTC demand. Sized at ~$45/Beta and ~$55/Alpha unit (matches
existing PO unit costs).

Uses existing 6 (Add PO) placeholder rows R186-R191 (3 Beta + 3 Alpha).
No range extension needed.

2028 demand totals:
  Beta:  15,869 units (5,000 Spring WS + 8,000 Fall WS + 2,869 DTC)
  Alpha:  4,238 units (1,600 Fall WS + 2,638 DTC)

2028 PO totals (sized with ~20% buffer for early 2029 carry):
  Beta:  19,000 units across 3 POs ($855K)
  Alpha:  6,000 units across 3 POs ($330K)
  Grand total: $1,185,000 inventory spend in 2028
"""

import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path

EXCEL_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Alma Mater/"
    "Empirica Financial Model/Alma Mater Financial Model Draft.xlsx"
)

# (row, name, product, pairs, amount, order_month, order_year)
NEW_2028_POS = [
    # Beta POs (R186-R188 placeholders)
    (186, "Spring 2028 (Beta)",  "Beta",  6000, 270000, 11, 2027),  # arrive Mar 2028
    (187, "Summer 2028 (Beta)",  "Beta",  5000, 225000,  2, 2028),  # arrive Jun 2028
    (188, "Fall 2028 (Beta)",    "Beta",  8000, 360000,  5, 2028),  # arrive Sep 2028
    # Alpha POs (R189-R191 placeholders)
    (189, "Spring 2028 (Alpha)", "Alpha", 1500,  82500, 11, 2027),  # arrive Mar 2028
    (190, "Summer 2028 (Alpha)", "Alpha", 2000, 110000,  2, 2028),  # arrive Jun 2028
    (191, "Fall 2028 (Alpha)",   "Alpha", 2500, 137500,  5, 2028),  # arrive Sep 2028
]


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    asm = wb["Assumptions"]
    yellow = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")

    for row, name, product, pairs, amount, order_month, order_year in NEW_2028_POS:
        asm.cell(row=row, column=2, value=name)
        asm.cell(row=row, column=3, value=product).fill = yellow
        asm.cell(row=row, column=4, value=pairs).fill = yellow
        asm.cell(row=row, column=5, value=amount).fill = yellow
        asm.cell(row=row, column=6, value=order_month).fill = yellow
        asm.cell(row=row, column=7, value=order_year).fill = yellow

    wb.save(EXCEL_PATH)
    print(f"✅ Saved Excel: {EXCEL_PATH}")

    beta_total_units = sum(p[3] for p in NEW_2028_POS if p[2] == "Beta")
    alpha_total_units = sum(p[3] for p in NEW_2028_POS if p[2] == "Alpha")
    beta_total_amt = sum(p[4] for p in NEW_2028_POS if p[2] == "Beta")
    alpha_total_amt = sum(p[4] for p in NEW_2028_POS if p[2] == "Alpha")

    print(f"\n2028 POs added:")
    for r, name, product, pairs, amount, m, y in NEW_2028_POS:
        print(f"  R{r} {name:<22} {product:<6} {pairs:>5}u ${amount:>6,}  order {m}/{y}")
    print(f"\n  Beta total:  {beta_total_units:>5}u  ${beta_total_amt:>6,}")
    print(f"  Alpha total: {alpha_total_units:>5}u  ${alpha_total_amt:>6,}")
    print(f"  Grand total: {beta_total_units + alpha_total_units:>5}u  ${beta_total_amt + alpha_total_amt:>6,}")


if __name__ == "__main__":
    main()
